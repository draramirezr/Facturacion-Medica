# ARSFLOW Gestion de Factras Medicas

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file, session
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
from datetime import datetime, timedelta
import json
import functools
from dotenv import load_dotenv
import secrets
import re
from markupsafe import escape
from io import BytesIO
import threading
import time
from collections import defaultdict
from threading import Lock
from functools import wraps
import logging
from logging.handlers import RotatingFileHandler

import pymysql
pymysql.install_as_MySQLdb()

try:
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import Mail
    import base64
    SENDGRID_AVAILABLE = True
except ImportError:
    SENDGRID_AVAILABLE = False
    print("AVISO: SendGrid no disponible")

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("AVISO: ReportLab no disponible")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Protection
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("AVISO: OpenPyXL no disponible")

load_dotenv()

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        RotatingFileHandler('app.log', maxBytes=10485760, backupCount=5),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

app.secret_key = os.getenv('SECRET_KEY', secrets.token_hex(32))

app.config['SESSION_COOKIE_SECURE'] = os.getenv('FLASK_ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['SESSION_COOKIE_NAME'] = 'facturacion_session'
app.config['TEMPLATES_AUTO_RELOAD'] = True

@app.after_request
def set_security_headers(response):
    """Agregar headers de seguridad a todas las respuestas"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://www.googletagmanager.com; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com https://cdnjs.cloudflare.com; "
        "img-src 'self' data: https:; "
        "connect-src 'self' https://www.googletagmanager.com; "
        "frame-ancestors 'self';"
    )
    response.headers['Content-Security-Policy'] = csp
    
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=31536000'
    else:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    
    return response

from flask import g
import gzip

@app.after_request
def compress_response(response):
    """Comprimir respuestas para reducir tamaño de transferencia"""
    if response.status_code < 200 or response.status_code >= 300:
        return response
    
    accept_encoding = request.headers.get('Accept-Encoding', '')
    
    if 'gzip' not in accept_encoding.lower():
        return response
    
    if response.direct_passthrough:
        return response
    
    # Comprimir solo si es mayor a 1KB
    if len(response.get_data()) < 1024:
        return response
    
    response.set_data(gzip.compress(response.get_data()))
    response.headers['Content-Encoding'] = 'gzip'
    response.headers['Content-Length'] = len(response.get_data())
    response.headers['Vary'] = 'Accept-Encoding'
    
    return response

def parse_mysql_url(url):
    """Parsear URL de MySQL"""
    if not url:
        return None
    pattern = r'mysql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)'
    match = re.match(pattern, url)
    if match:
        return {
            'user': match.group(1),
            'password': match.group(2),
            'host': match.group(3),
            'port': int(match.group(4)),
            'database': match.group(5),
            'charset': 'utf8mb4'
        }
    return None

mysql_url = os.getenv('MYSQL_URL', '')
if mysql_url:
    parsed_config = parse_mysql_url(mysql_url)
    if parsed_config:
        DATABASE_CONFIG = parsed_config
    else:
        raise Exception("MYSQL_URL inválida")
else:
    DATABASE_CONFIG = {
        'host': os.getenv('MYSQL_HOST', 'localhost'),
        'user': os.getenv('MYSQL_USER', 'root'),
        'password': os.getenv('MYSQL_PASSWORD', ''),
        'database': os.getenv('MYSQL_DATABASE', 'facturacion_medica'),
        'port': int(os.getenv('MYSQL_PORT', '3306')),
        'charset': 'utf8mb4'
    }

print(f"[OK] Configurado para MySQL: {DATABASE_CONFIG['database']}")

# Cache de conexiones para mejor rendimiento (pymysql no tiene pool nativo)
# Usaremos conexiones reutilizables con contexto de aplicación Flask
from flask import g

def get_db_connection():
    """
    Obtener conexión a la base de datos
    Reutiliza conexión en el contexto de la request si existe
    """
    if 'db_conn' not in g:
        config = DATABASE_CONFIG.copy()
        config['cursorclass'] = pymysql.cursors.DictCursor
        # Configuraciones de rendimiento
        config['autocommit'] = False
        g.db_conn = pymysql.connect(**config)
        logger.debug("Nueva conexión a BD creada")
    
    return g.db_conn

@app.teardown_appcontext
def close_db(error):
    """Cerrar conexión al finalizar request"""
    db_conn = g.pop('db_conn', None)
    if db_conn is not None:
        try:
            db_conn.close()
            logger.debug("Conexión a BD cerrada")
        except Exception as e:
            logger.error(f"Error al cerrar conexión: {e}")

def execute_query(query, params=None, fetch='one'):
    """
    Ejecutar query y retornar resultados
    Con manejo mejorado de errores, logging y reutilización de conexiones
    """
    cursor = None
    try:
        # Validar que la query no esté vacía
        if not query or not query.strip():
            logger.error("Intento de ejecutar query vacía")
            return None
        
        # Validar que params sea tupla o lista si se proporciona
        if params is not None and not isinstance(params, (tuple, list)):
            logger.warning(f"Params debe ser tupla o lista, recibido: {type(params)}")
            params = (params,)
        
        # Validar que no haya SQL injection básico (solo advertencia)
        if params and any(isinstance(p, str) and (';' in p or '--' in p or '/*' in p) for p in params):
            logger.warning("Posible intento de SQL injection detectado en params")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(query, params or ())
        
        if fetch == 'one':
            result = cursor.fetchone()
        elif fetch == 'all':
            result = cursor.fetchall()
        else:
            result = None
        
        conn.commit()
        return result
    except pymysql.Error as e:
        conn = get_db_connection()
        try:
            conn.rollback()
        except:
            pass
        logger.error(f"Error SQL en execute_query: {e} - Query: {query[:100]}...")
        # No exponer detalles del error al usuario en producción
        if os.getenv('FLASK_ENV') == 'development':
            raise
        return None
    except Exception as e:
        conn = get_db_connection()
        try:
            conn.rollback()
        except:
            pass
        logger.error(f"Error inesperado en execute_query: {e}", exc_info=True)
        if os.getenv('FLASK_ENV') == 'development':
            raise
        return None
    finally:
        if cursor:
            cursor.close()
        # No cerramos la conexión aquí, se cierra al finalizar la request

def execute_update(query, params=None):
    """
    Ejecutar UPDATE/INSERT/DELETE
    Con manejo mejorado de errores, logging y reutilización de conexiones
    """
    cursor = None
    try:
        # Validar que la query no esté vacía
        if not query or not query.strip():
            logger.error("Intento de ejecutar update con query vacía")
            return None
        
        # Validar que params sea tupla o lista si se proporciona
        if params is not None and not isinstance(params, (tuple, list)):
            logger.warning(f"Params debe ser tupla o lista, recibido: {type(params)}")
            params = (params,)
        
        # Validar que no haya SQL injection básico (solo advertencia)
        if params and any(isinstance(p, str) and (';' in p or '--' in p or '/*' in p) for p in params):
            logger.warning("Posible intento de SQL injection detectado en params")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(query, params or ())
        conn.commit()
        return cursor.lastrowid
    except pymysql.Error as e:
        conn = get_db_connection()
        try:
            conn.rollback()
        except:
            pass
        logger.error(f"Error SQL en execute_update: {e} - Query: {query[:100]}...")
        # No exponer detalles del error al usuario en producción
        if os.getenv('FLASK_ENV') == 'development':
            raise
        return None
    except Exception as e:
        conn = get_db_connection()
        try:
            conn.rollback()
        except:
            pass
        logger.error(f"Error inesperado en execute_update: {e}", exc_info=True)
        if os.getenv('FLASK_ENV') == 'development':
            raise
        return None
    finally:
        if cursor:
            cursor.close()
        # No cerramos la conexión aquí, se cierra al finalizar la request

def sanitize_input(text, max_length=500, allow_html=False):
    """
    Sanitizar entrada de texto para prevenir XSS y otros ataques
    
    Args:
        text: Texto a sanitizar
        max_length: Longitud máxima permitida
        allow_html: Si True, permite HTML (usar con escape en templates)
    
    Returns: Texto sanitizado
    """
    if not text:
        return ""
    
    text = str(text).strip()
    
    # Remover HTML/scripts a menos que se permita explícitamente
    if not allow_html:
        text = re.sub(r'<[^>]*>', '', text)
        # Remover caracteres peligrosos
        text = text.replace('javascript:', '').replace('onerror=', '')
        text = text.replace('onclick=', '').replace('onload=', '')
    
    # Limitar longitud
    if max_length and len(text) > max_length:
        text = text[:max_length]
        logger.warning(f"Texto truncado por exceder longitud máxima: {max_length}")
    
    return text

def validate_int(value, min_value=None, max_value=None, default=None):
    """Validar y convertir a entero de forma segura"""
    try:
        int_value = int(value)
        if min_value is not None and int_value < min_value:
            return default
        if max_value is not None and int_value > max_value:
            return default
        return int_value
    except (ValueError, TypeError):
        return default

def validate_float(value, min_value=None, max_value=None, default=None):
    """Validar y convertir a float de forma segura"""
    try:
        float_value = float(value)
        if min_value is not None and float_value < min_value:
            return default
        if max_value is not None and float_value > max_value:
            return default
        return float_value
    except (ValueError, TypeError):
        return default

def validate_email(email):
    """Validar formato de email"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validar_password_segura(password):
    """Validar contraseña segura"""
    errores = []
    if len(password) < 8:
        errores.append("Mínimo 8 caracteres")
    if not re.search(r'[A-Z]', password):
        errores.append("Al menos una mayúscula")
    if not re.search(r'[a-z]', password):
        errores.append("Al menos una minúscula")
    if not re.search(r'\d', password):
        errores.append("Al menos un número")
    return errores

@app.template_filter('formato_moneda')
def formato_moneda(valor):
    """Formatear números como moneda"""
    try:
        return "{:,.2f}".format(float(valor))
    except (ValueError, TypeError):
        return "0.00"

# Context processor para temas de color
@app.context_processor
def inject_theme():
    """Inyectar tema de color en todos los templates"""
    TEMAS = {
        'cyan': {
            'primary': '#06B6D4',
            'primary_dark': '#0891B2',
            'primary_light': '#22D3EE',
            'background': '#F0FDFA',
            'gradient_start': '#06B6D4',
            'gradient_end': '#0891B2',
            'nombre': 'Verde Azulado',
            'categoria': 'Fresco'
        },
        'ocean': {
            'primary': '#0EA5E9',
            'primary_dark': '#0284C7',
            'primary_light': '#38BDF8',
            'background': '#F0F9FF',
            'gradient_start': '#0EA5E9',
            'gradient_end': '#0284C7',
            'nombre': 'Azul Océano',
            'categoria': 'Fresco'
        },
        'emerald': {
            'primary': '#10B981',
            'primary_dark': '#059669',
            'primary_light': '#34D399',
            'background': '#F0FDF4',
            'gradient_start': '#10B981',
            'gradient_end': '#059669',
            'nombre': 'Verde Esmeralda',
            'categoria': 'Fresco'
        },
        'teal': {
            'primary': '#14B8A6',
            'primary_dark': '#0D9488',
            'primary_light': '#2DD4BF',
            'background': '#F0FDFA',
            'gradient_start': '#14B8A6',
            'gradient_end': '#0D9488',
            'nombre': 'Verde Azulado Oscuro',
            'categoria': 'Fresco'
        },
        'coral': {
            'primary': '#FF6B6B',
            'primary_dark': '#EE5A52',
            'primary_light': '#FF8787',
            'background': '#FFF5F5',
            'gradient_start': '#FF6B6B',
            'gradient_end': '#EE5A52',
            'nombre': 'Coral Cálido',
            'categoria': 'Cálido'
        },
        'sunset': {
            'primary': '#F59E0B',
            'primary_dark': '#D97706',
            'primary_light': '#FBBF24',
            'background': '#FFFBEB',
            'gradient_start': '#F59E0B',
            'gradient_end': '#D97706',
            'nombre': 'Naranja Atardecer',
            'categoria': 'Cálido'
        },
        'rose': {
            'primary': '#F43F5E',
            'primary_dark': '#E11D48',
            'primary_light': '#FB7185',
            'background': '#FFF1F2',
            'gradient_start': '#F43F5E',
            'gradient_end': '#E11D48',
            'nombre': 'Rosa Moderno',
            'categoria': 'Cálido'
        },
        'amber': {
            'primary': '#F59E0B',
            'primary_dark': '#D97706',
            'primary_light': '#FCD34D',
            'background': '#FEFCE8',
            'gradient_start': '#F59E0B',
            'gradient_end': '#D97706',
            'nombre': 'Ámbar Dorado',
            'categoria': 'Cálido'
        },
        'indigo': {
            'primary': '#4F46E5',
            'primary_dark': '#4338CA',
            'primary_light': '#6366F1',
            'background': '#EEF2FF',
            'gradient_start': '#4F46E5',
            'gradient_end': '#6366F1',
            'nombre': 'Azul Profesional',
            'categoria': 'Profesional'
        },
        'purple': {
            'primary': '#7C3AED',
            'primary_dark': '#6D28D9',
            'primary_light': '#8B5CF6',
            'background': '#FAF5FF',
            'gradient_start': '#7C3AED',
            'gradient_end': '#8B5CF6',
            'nombre': 'Púrpura Elegante',
            'categoria': 'Profesional'
        },
        'violet': {
            'primary': '#8B5CF6',
            'primary_dark': '#7C3AED',
            'primary_light': '#A78BFA',
            'background': '#F5F3FF',
            'gradient_start': '#8B5CF6',
            'gradient_end': '#7C3AED',
            'nombre': 'Violeta Suave',
            'categoria': 'Profesional'
        },
        'slate': {
            'primary': '#475569',
            'primary_dark': '#334155',
            'primary_light': '#64748B',
            'background': '#F8FAFC',
            'gradient_start': '#475569',
            'gradient_end': '#334155',
            'nombre': 'Gris Elegante',
            'categoria': 'Neutral'
        },
        'navy': {
            'primary': '#1E3A8A',
            'primary_dark': '#1E40AF',
            'primary_light': '#3B82F6',
            'background': '#EFF6FF',
            'gradient_start': '#1E3A8A',
            'gradient_end': '#1E40AF',
            'nombre': 'Azul Marino',
            'categoria': 'Profesional'
        },
        'forest': {
            'primary': '#15803D',
            'primary_dark': '#166534',
            'primary_light': '#22C55E',
            'background': '#F0FDF4',
            'gradient_start': '#15803D',
            'gradient_end': '#166534',
            'nombre': 'Verde Bosque',
            'categoria': 'Natural'
        },
        'wine': {
            'primary': '#BE123C',
            'primary_dark': '#9F1239',
            'primary_light': '#E11D48',
            'background': '#FFF1F2',
            'gradient_start': '#BE123C',
            'gradient_end': '#9F1239',
            'nombre': 'Vino Elegante',
            'categoria': 'Sofisticado'
        },
        'bronze': {
            'primary': '#92400E',
            'primary_dark': '#78350F',
            'primary_light': '#B45309',
            'background': '#FEF3C7',
            'gradient_start': '#92400E',
            'gradient_end': '#78350F',
            'nombre': 'Bronce Cálido',
            'categoria': 'Cálido'
        }
    }
    
    tema_actual = 'cyan'  # Default
    if current_user.is_authenticated:
        tema_actual = current_user.tema_color or 'cyan'
    
    # Información de empresa para multi-tenant
    empresa_info = {}
    if current_user.is_authenticated and hasattr(current_user, 'tenant_id'):
        empresa_info = {
            'tenant_id': current_user.tenant_id,
            'empresa_nombre': current_user.empresa_nombre or 'Sin empresa'
        }
    
    return {
        'tema': TEMAS.get(tema_actual, TEMAS['cyan']),
        'tema_nombre': tema_actual,
        'temas_disponibles': TEMAS,
        'empresa': empresa_info
    }

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = None

class User(UserMixin):
    def __init__(self, id, nombre, email, perfil, tema_color='cyan', tenant_id=1, empresa_nombre=''):
        self.id = id
        self.nombre = nombre
        self.email = email
        self.perfil = perfil
        self.tema_color = tema_color or 'cyan'
        self.tenant_id = tenant_id
        self.empresa_nombre = empresa_nombre

@login_manager.user_loader
def load_user(user_id):
    """Cargar usuario desde la base de datos con información de empresa"""
    user_data = execute_query('''
        SELECT u.*, e.nombre as empresa_nombre 
        FROM usuarios u
        LEFT JOIN empresas e ON u.tenant_id = e.id
        WHERE u.id = %s AND u.activo = 1
    ''', (user_id,))
    
    if user_data:
        return User(
            id=user_data['id'],
            nombre=user_data['nombre'],
            email=user_data['email'],
            perfil=user_data['perfil'],
            tema_color=user_data.get('tema_color', 'cyan'),
            tenant_id=user_data.get('tenant_id', 1),
            empresa_nombre=user_data.get('empresa_nombre', '')
        )
    return None

def validate_id(value, field_name="ID"):
    """
    Validar que un valor sea un ID válido (entero positivo).
    Previene inyección SQL y errores de tipo.
    """
    try:
        id_val = int(value)
        if id_val <= 0:
            raise ValueError(f"{field_name} debe ser positivo")
        return id_val
    except (ValueError, TypeError):
        raise ValueError(f"{field_name} inválido")

def validate_date(date_string, field_name="Fecha"):
    """Validar formato de fecha"""
    try:
        return datetime.strptime(date_string, '%Y-%m-%d').date()
    except ValueError:
        raise ValueError(f"{field_name} inválida. Formato esperado: YYYY-MM-DD")

def validate_numeric(value, field_name="Valor", min_val=None, max_val=None):
    """Validar valores numéricos con rangos opcionales"""
    try:
        num = float(value)
        if min_val is not None and num < min_val:
            raise ValueError(f"{field_name} debe ser mayor o igual a {min_val}")
        if max_val is not None and num > max_val:
            raise ValueError(f"{field_name} debe ser menor o igual a {max_val}")
        return num
    except (ValueError, TypeError):
        raise ValueError(f"{field_name} debe ser un número válido")

@app.errorhandler(404)
def not_found(error):
    """Página no encontrada"""
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """Error interno del servidor"""
    return render_template('errors/500.html'), 500

@app.errorhandler(403)
def forbidden(error):
    """Acceso prohibido"""
    flash('No tienes permisos para acceder a este recurso', 'error')
    return redirect(url_for('facturacion_menu')), 403

def get_current_tenant_id():
    """
    Obtener el TenantID del usuario actual de forma segura.
    El TenantID NUNCA viene del cliente, siempre de la sesión.
    """
    if current_user.is_authenticated:
        return current_user.tenant_id
    return None

def require_tenant(func):
    """
    Decorador que asegura que el usuario tenga un TenantID válido.
    Uso: @require_tenant antes de @login_required
    """
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        
        if not hasattr(current_user, 'tenant_id') or not current_user.tenant_id:
            flash('Error: Usuario sin empresa asignada', 'error')
            return redirect(url_for('logout'))
        
        return func(*args, **kwargs)
    return wrapper

def add_tenant_filter(query, table_alias=''):
    """
    Agregar filtro de TenantID automáticamente a una query.
    
    Uso:
        query = "SELECT * FROM pacientes WHERE activo = 1"
        query = add_tenant_filter(query)
        # Resultado: "SELECT * FROM pacientes WHERE activo = 1 AND tenant_id = %s"
    """
    tenant_id = get_current_tenant_id()
    if not tenant_id:
        return query
    
    table_ref = f"{table_alias}." if table_alias else ""
    
    # Si la query ya tiene WHERE, agregar AND
    if 'WHERE' in query.upper():
        query += f" AND {table_ref}tenant_id = {tenant_id}"
    else:
        query += f" WHERE {table_ref}tenant_id = {tenant_id}"
    
    return query

# Whitelist de tablas permitidas para prevenir SQL injection
_ALLOWED_TABLES = {
    'usuarios', 'pacientes', 'medicos', 'ars', 'servicios', 'ncf', 
    'facturas', 'factura_detalles', 'pacientes_pendientes', 
    'centros_medicos', 'empresas'
}

_ALLOWED_ID_COLUMNS = {'id', 'paciente_id', 'medico_id', 'ars_id', 'factura_id'}

def validate_tenant_access(table, record_id, id_column='id'):
    """
    Validar que un registro pertenece al Tenant del usuario actual.
    Previene acceso a datos de otras empresas.
    
    Args:
        table: Nombre de la tabla (debe estar en whitelist)
        record_id: ID del registro a validar
        id_column: Nombre de la columna ID (debe estar en whitelist)
    
    Returns: True si el usuario tiene acceso, False si no.
    """
    # Validación de seguridad: whitelist de tablas y columnas
    if table not in _ALLOWED_TABLES:
        logger.warning(f"Intento de acceso a tabla no permitida: {table}")
        return False
    
    if id_column not in _ALLOWED_ID_COLUMNS:
        logger.warning(f"Intento de acceso con columna ID no permitida: {id_column}")
        return False
    
    # Validar que record_id sea un entero
    try:
        record_id = int(record_id)
    except (ValueError, TypeError):
        logger.warning(f"record_id inválido: {record_id}")
        return False
    
    tenant_id = get_current_tenant_id()
    if not tenant_id:
        return False
    
    # Query segura usando parámetros
    result = execute_query(
        "SELECT COUNT(*) as count FROM {} WHERE {} = %s AND tenant_id = %s".format(
            table, id_column
        ),
        (record_id, tenant_id)
    )
    
    return result and result.get('count', 0) > 0

def check_license_available(tenant_id):
    """
    Verificar si una empresa tiene licencias disponibles.
    Returns: (disponible: bool, licencias_restantes: int, mensaje: str)
    """
    empresa = execute_query('''
        SELECT licencias_totales, licencias_usadas, 
               (licencias_totales - licencias_usadas) as disponibles
        FROM empresas 
        WHERE id = %s AND estado = 'activo'
    ''', (tenant_id,))
    
    if not empresa:
        return (False, 0, "Empresa no encontrada o inactiva")
    
    disponibles = empresa['disponibles']
    
    if disponibles <= 0:
        return (False, 0, f"No hay licencias disponibles ({empresa['licencias_usadas']}/{empresa['licencias_totales']} en uso)")
    
    return (True, disponibles, f"{disponibles} licencias disponibles")

def get_empresa_info(tenant_id=None):
    """
    Obtener información de la empresa actual.
    """
    if tenant_id is None:
        tenant_id = get_current_tenant_id()
    
    if not tenant_id:
        return None
    
    return execute_query('SELECT * FROM empresas WHERE id = %s', (tenant_id,))

def verificar_suscripciones_vencidas():
    """
    Verificar y suspender automáticamente empresas con suscripción vencida.
    Esta función es un backup del evento MySQL.
    Se ejecuta en momentos clave (login, dashboard admin, etc.)
    """
    try:
        from datetime import date
        
        # Suspender empresas cuya fecha_fin ya pasó
        result = execute_update('''
            UPDATE empresas
            SET estado = 'suspendido'
            WHERE fecha_fin < CURDATE()
              AND estado = 'activo'
        ''')
        
        # Retornar número de empresas suspendidas
        return result if result else 0
    except Exception as e:
        print(f"Error verificando suscripciones: {e}")
        return 0

def get_dias_restantes_suscripcion(tenant_id=None):
    """
    Obtener días restantes de suscripción de una empresa.
    Returns: (dias_restantes: int, estado: str, mensaje: str)
    """
    empresa = get_empresa_info(tenant_id)
    
    if not empresa:
        return (0, 'error', 'Empresa no encontrada')
    
    if not empresa.get('fecha_fin'):
        return (9999, 'sin_fecha', 'Sin fecha de vencimiento')
    
    from datetime import date
    fecha_fin = empresa['fecha_fin']
    if isinstance(fecha_fin, str):
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    dias = (fecha_fin - date.today()).days
    
    if dias < 0:
        return (dias, 'vencida', f'Suscripción vencida hace {abs(dias)} días')
    elif dias <= 7:
        return (dias, 'urgente', f'Vence en {dias} días - URGENTE')
    elif dias <= 30:
        return (dias, 'proximo', f'Vence en {dias} días')
    else:
        return (dias, 'vigente', f'{dias} días restantes')

def execute_query_tenant(query, params=None, fetch='one'):
    """
    Wrapper para execute_query que automáticamente agrega filtro de tenant_id
    Solo para tablas que tienen tenant_id
    
    IMPORTANTE: Esta función agrega el tenant_id automáticamente al final de params
    """
    tenant_id = get_current_tenant_id()
    if not tenant_id:
        return execute_query(query, params, fetch)
    
    # Convertir params a lista si es tupla
    if params is None:
        params = []
    elif isinstance(params, tuple):
        params = list(params)
    elif not isinstance(params, list):
        params = [params]
    
    # Agregar tenant_id a los parámetros
    params.append(tenant_id)
    
    return execute_query(query, tuple(params), fetch)

def execute_update_tenant(query, params=None):
    """
    Wrapper para execute_update que valida tenant_id en UPDATEs/DELETEs
    """
    return execute_update(query, params)

request_counts = defaultdict(list)
rate_limit_lock = Lock()

def rate_limit(max_requests=10, window=60):
    """Decorador para rate limiting"""
    def decorator(f):
        @functools.wraps(f)
        def wrapper(*args, **kwargs):
            client_ip = request.remote_addr
            current_time = time.time()
            
            with rate_limit_lock:
                request_counts[client_ip] = [
                    req_time for req_time in request_counts[client_ip]
                    if current_time - req_time < window
                ]
                
                if len(request_counts[client_ip]) >= max_requests:
                    return jsonify({'error': 'Rate limit exceeded'}), 429
                
                request_counts[client_ip].append(current_time)
            
            return f(*args, **kwargs)
        return wrapper
    return decorator


@app.route('/')
def index():
    """Página principal - redirige al login"""
    if current_user.is_authenticated:
        return redirect(url_for('facturacion_menu'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Inicio de sesión"""
    if current_user.is_authenticated:
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        # Verificar suscripciones vencidas antes del login
        verificar_suscripciones_vencidas()
        # Rate limit manual
        client_ip = request.remote_addr
        current_time = time.time()
        
        with rate_limit_lock:
            request_counts[f'{client_ip}_login'] = [
                req_time for req_time in request_counts.get(f'{client_ip}_login', [])
                if current_time - req_time < 300
            ]
            
            if len(request_counts.get(f'{client_ip}_login', [])) >= 5:
                flash('Demasiados intentos. Espera 5 minutos.', 'error')
                return redirect(url_for('login'))
            
            if f'{client_ip}_login' not in request_counts:
                request_counts[f'{client_ip}_login'] = []
            request_counts[f'{client_ip}_login'].append(current_time)
        
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        
        if not email or not password:
            flash('Por favor ingresa email y contraseña', 'error')
            return redirect(url_for('login'))
        
        # Obtener usuario con información de empresa
        user_data = execute_query('''
            SELECT u.*, 
                   e.nombre as empresa_nombre, 
                   e.estado as empresa_estado,
                   e.fecha_fin as empresa_fecha_fin
            FROM usuarios u
            LEFT JOIN empresas e ON u.tenant_id = e.id
            WHERE u.email = %s
        ''', (email,))
        
        if user_data and user_data['activo']:
            # Validar que la empresa esté activa
            if user_data.get('empresa_estado') != 'activo':
                if user_data.get('empresa_estado') == 'suspendido':
                    flash('Suscripción vencida o suspendida. Contacta al administrador del sistema.', 'error')
                else:
                    flash('La empresa asociada a este usuario está inactiva', 'error')
                return redirect(url_for('login'))
            
            # Validar fecha de fin de suscripción
            if user_data.get('empresa_fecha_fin'):
                from datetime import date
                fecha_fin = user_data['empresa_fecha_fin']
                if isinstance(fecha_fin, str):
                    fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                
                if fecha_fin < date.today():
                    # Suscripción vencida - suspender empresa automáticamente
                    execute_update('''
                        UPDATE empresas 
                        SET estado = 'suspendido' 
                        WHERE id = %s AND estado = 'activo'
                    ''', (user_data.get('tenant_id'),))
                    
                    flash('La suscripción de tu empresa ha vencido. Contacta al administrador.', 'error')
                    return redirect(url_for('login'))
            
            if check_password_hash(user_data['password_hash'], password):
                if user_data['password_temporal']:
                    session['cambio_password_usuario_id'] = user_data['id']
                    session['cambio_password_email'] = user_data['email']
                    flash('Debes cambiar tu contraseña temporal', 'warning')
                    return redirect(url_for('cambiar_password_obligatorio'))
                
                user = User(
                    id=user_data['id'],
                    nombre=user_data['nombre'],
                    email=user_data['email'],
                    perfil=user_data['perfil'],
                    tema_color=user_data.get('tema_color', 'cyan'),
                    tenant_id=user_data.get('tenant_id', 1),
                    empresa_nombre=user_data.get('empresa_nombre', '')
                )
                
                session.permanent = True
                session['tenant_id'] = user.tenant_id  # Guardar tenant_id en sesión
                session['empresa_nombre'] = user.empresa_nombre
                login_user(user, remember=True)
                
                execute_update('UPDATE usuarios SET last_login = %s WHERE id = %s',
                           (datetime.now(), user_data['id']))
                
                return redirect(url_for('facturacion_menu'))
            else:
                flash('Contraseña incorrecta', 'error')
        else:
            flash('Usuario no encontrado o inactivo', 'error')
        
        return redirect(url_for('login'))
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Cerrar sesión"""
    logout_user()
    flash('Sesión cerrada correctamente', 'success')
    return redirect(url_for('index'))

@app.route('/cambiar-password-obligatorio', methods=['GET', 'POST'])
def cambiar_password_obligatorio():
    """Cambio de contraseña obligatorio"""
    if 'cambio_password_usuario_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')
        
        if not password or not password_confirm:
            flash('Debes completar todos los campos', 'error')
            return redirect(url_for('cambiar_password_obligatorio'))
        
        if password != password_confirm:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('cambiar_password_obligatorio'))
        
        password_errors = validar_password_segura(password)
        if password_errors:
            flash(f'Contraseña no válida: {", ".join(password_errors)}', 'error')
            return redirect(url_for('cambiar_password_obligatorio'))
        
        user_id = session['cambio_password_usuario_id']
        password_hash = generate_password_hash(password)
        
        execute_update('''
            UPDATE usuarios 
            SET password_hash = %s, password_temporal = 0
            WHERE id = %s
        ''', (password_hash, user_id))
        
        user_data = execute_query('SELECT * FROM usuarios WHERE id = %s', (user_id,))
        
        # Limpiar sesión temporal
        session.pop('cambio_password_usuario_id', None)
        session.pop('cambio_password_email', None)
        
        # Login automático
        user = User(
            id=user_data['id'],
            nombre=user_data['nombre'],
            email=user_data['email'],
            perfil=user_data['perfil'],
            tema_color=user_data.get('tema_color', 'cyan')
        )
        login_user(user, remember=True)
        
        flash('Contraseña cambiada exitosamente', 'success')
        return redirect(url_for('facturacion_menu'))
    
    return render_template('cambiar_password_obligatorio.html')

@app.route('/solicitar-recuperacion', methods=['GET', 'POST'])
@rate_limit(max_requests=3, window=300)
def solicitar_recuperacion():
    """Solicitar recuperación de contraseña"""
    if current_user.is_authenticated:
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        
        if not email or not validate_email(email):
            flash('Por favor ingresa un email válido', 'error')
            return redirect(url_for('solicitar_recuperacion'))
        
        usuario = execute_query(
            'SELECT * FROM usuarios WHERE email = %s AND activo = 1',
            (email,)
        )
        
        if usuario:
            # Generar token de recuperación
            token = secrets.token_urlsafe(32)
            expiracion = datetime.now() + timedelta(hours=1)
            
            execute_update('''
                UPDATE usuarios 
                SET reset_token = %s, reset_token_expiracion = %s
                WHERE id = %s
            ''', (token, expiracion, usuario['id']))
            
            # Enviar email si SendGrid está disponible
            if SENDGRID_AVAILABLE:
                try:
                    reset_url = url_for('recuperar_password', token=token, _external=True)
                    
                    message = Mail(
                        from_email=os.getenv('SENDGRID_FROM_EMAIL', 'noreply@facturacion.com'),
                        to_emails=email,
                        subject='Recuperación de Contraseña - ARSFLOW Gestion de Factras Medicas',
                        html_content=f'''
                        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                            <h2 style="color: #CEB0B7;">Recuperación de Contraseña</h2>
                            <p>Hola {usuario['nombre']},</p>
                            <p>Has solicitado recuperar tu contraseña. Haz clic en el siguiente enlace para crear una nueva contraseña:</p>
                            <p style="margin: 30px 0;">
                                <a href="{reset_url}" 
                                   style="background: #CEB0B7; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">
                                    Recuperar Contraseña
                                </a>
                            </p>
                            <p style="color: #666; font-size: 14px;">Este enlace expirará en 1 hora.</p>
                            <p style="color: #666; font-size: 14px;">Si no solicitaste este cambio, ignora este email.</p>
                            <hr style="border: none; border-top: 1px solid #eee; margin: 30px 0;">
                            <p style="color: #999; font-size: 12px;">ARSFLOW Gestion de Factras Medicas</p>
                        </div>
                        '''
                    )
                    
                    sg = SendGridAPIClient(os.getenv('SENDGRID_API_KEY'))
                    sg.send(message)
                    
                    flash('Se ha enviado un email con instrucciones para recuperar tu contraseña', 'success')
                except Exception as e:
                    print(f"Error enviando email: {e}")
                    flash('Error al enviar el email. Contacta al administrador.', 'error')
            else:
                # Si no hay SendGrid, mostrar el token (solo para desarrollo)
                flash(f'Token de recuperación (solo desarrollo): {token}', 'info')
                flash('Usa este enlace para recuperar tu contraseña', 'info')
        else:
            # Por seguridad, mostramos el mismo mensaje aunque el usuario no exista
            flash('Si el email existe, recibirás instrucciones para recuperar tu contraseña', 'info')
        
        return redirect(url_for('login'))
    
    return render_template('solicitar_recuperacion.html')

@app.route('/recuperar-password/<token>', methods=['GET', 'POST'])
def recuperar_password(token):
    """Recuperar contraseña con token"""
    if current_user.is_authenticated:
        return redirect(url_for('facturacion_menu'))
    
    usuario = execute_query('''
        SELECT * FROM usuarios 
        WHERE reset_token = %s 
        AND reset_token_expiracion > %s
        AND activo = 1
    ''', (token, datetime.now()))
    
    if not usuario:
        flash('El enlace de recuperación es inválido o ha expirado', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')
        
        if not password or not password_confirm:
            flash('Debes completar todos los campos', 'error')
            return redirect(url_for('recuperar_password', token=token))
        
        if password != password_confirm:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('recuperar_password', token=token))
        
        password_errors = validar_password_segura(password)
        if password_errors:
            flash(f'Contraseña no válida: {", ".join(password_errors)}', 'error')
            return redirect(url_for('recuperar_password', token=token))
        
        password_hash = generate_password_hash(password)
        
        execute_update('''
            UPDATE usuarios 
            SET password_hash = %s, 
                password_temporal = 0,
                reset_token = NULL,
                reset_token_expiracion = NULL
            WHERE id = %s
        ''', (password_hash, usuario['id']))
        
        flash('Contraseña actualizada exitosamente. Ahora puedes iniciar sesión.', 'success')
        return redirect(url_for('login'))
    
    return render_template('recuperar_password.html', token=token)

@app.route('/admin/empresas')
@login_required
def admin_empresas():
    """Listar empresas - Super Admin ve todas, Administrador ve solo la suya"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos para gestionar empresas', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    
    # Si es Super Admin (sin tenant_id), ver todas las empresas
    if tenant_id is None:
        # Super Admin puede ver todas las empresas
        empresas = execute_query('''
            SELECT e.*, 
                   COUNT(u.id) as total_usuarios
            FROM empresas e
            LEFT JOIN usuarios u ON e.id = u.tenant_id AND u.activo = 1
            GROUP BY e.id
            ORDER BY e.fecha_creacion DESC
        ''', fetch='all')
    else:
        # Administrador de empresa solo ve su propia empresa
        empresas = execute_query('''
            SELECT e.*, 
                   COUNT(u.id) as total_usuarios
            FROM empresas e
            LEFT JOIN usuarios u ON e.id = u.tenant_id AND u.activo = 1
            WHERE e.id = %s
            GROUP BY e.id
            ORDER BY e.fecha_creacion DESC
        ''', (tenant_id,), fetch='all') or []
    
    # Verificar y suspender empresas vencidas
    empresas_suspendidas = verificar_suscripciones_vencidas()
    if empresas_suspendidas > 0:
        flash(f'{empresas_suspendidas} empresa(s) suspendida(s) por vencimiento de suscripción', 'warning')
    
    # Super Admin puede ver todas las empresas
    empresas = execute_query('''
        SELECT e.*, 
               COUNT(u.id) as total_usuarios
        FROM empresas e
        LEFT JOIN usuarios u ON e.id = u.tenant_id AND u.activo = 1
        GROUP BY e.id
        ORDER BY e.fecha_creacion DESC
    ''', fetch='all')
    
    # Calcular días restantes y estado de suscripción para cada empresa
    from datetime import date
    for empresa in empresas:
        if empresa.get('fecha_fin'):
            try:
                fecha_fin = empresa['fecha_fin']
                if isinstance(fecha_fin, str):
                    fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                
                dias_restantes = (fecha_fin - date.today()).days
                empresa['dias_restantes'] = dias_restantes
                
                # Determinar estado y clase CSS
                if dias_restantes < 0:
                    empresa['estado_suscripcion'] = 'vencida'
                    empresa['estado_texto'] = 'VENCIDA'
                    empresa['estado_clase'] = 'danger'
                    empresa['estado_icono'] = 'exclamation-circle'
                elif dias_restantes <= 7:
                    empresa['estado_suscripcion'] = 'urgente'
                    empresa['estado_texto'] = f'{dias_restantes}d - URGENTE'
                    empresa['estado_clase'] = 'danger'
                    empresa['estado_icono'] = 'exclamation-triangle'
                elif dias_restantes <= 30:
                    empresa['estado_suscripcion'] = 'proximo'
                    empresa['estado_texto'] = f'{dias_restantes} días'
                    empresa['estado_clase'] = 'warning'
                    empresa['estado_icono'] = 'clock'
                else:
                    empresa['estado_suscripcion'] = 'vigente'
                    empresa['estado_texto'] = f'{dias_restantes} días'
                    empresa['estado_clase'] = 'success'
                    empresa['estado_icono'] = 'check-circle'
            except:
                empresa['dias_restantes'] = None
                empresa['estado_suscripcion'] = 'error'
                empresa['estado_texto'] = 'Error'
                empresa['estado_clase'] = 'secondary'
                empresa['estado_icono'] = 'question'
        else:
            empresa['dias_restantes'] = None
            empresa['estado_suscripcion'] = 'sin_fecha'
            empresa['estado_texto'] = 'Sin fecha'
            empresa['estado_clase'] = 'secondary'
            empresa['estado_icono'] = 'calendar'
    
    return render_template('admin/empresas/lista.html', empresas=empresas)

@app.route('/admin/empresas/nueva', methods=['GET', 'POST'])
@login_required
def admin_empresas_nueva():
    """Crear nueva empresa - Solo Super Admin (sin tenant_id)"""
    # Solo usuarios sin tenant_id (Super Admin) pueden crear empresas
    # Los administradores de una empresa solo gestionan usuarios dentro de su empresa
    tenant_id = get_current_tenant_id()
    if tenant_id is not None:
        flash('No tienes permisos para crear empresas. Solo Super Administradores pueden crear nuevas empresas.', 'error')
        return redirect(url_for('admin_empresas'))
    
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 255)
        razon_social = sanitize_input(request.form.get('razon_social', ''), 255)
        rnc = sanitize_input(request.form.get('rnc', ''), 20)
        telefono = sanitize_input(request.form.get('telefono', ''), 20)
        email = request.form.get('email', '').strip().lower()
        direccion = sanitize_input(request.form.get('direccion', ''), 500)
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')
        licencias_totales = validate_int(request.form.get('licencias_totales', 5), min_value=1, max_value=1000, default=5)
        plan = request.form.get('plan', 'basico')
        tipo_empresa = request.form.get('tipo_empresa', '').strip()
        
        # Validar tipo_empresa
        if tipo_empresa not in ['medico', 'centro_salud']:
            flash('Debe seleccionar un tipo de empresa válido', 'error')
            return redirect(url_for('admin_empresas_nueva'))
        
        if not nombre or not razon_social or not fecha_inicio or not fecha_fin or not tipo_empresa:
            flash('Nombre, Razón Social, Tipo de Empresa y fechas son obligatorios', 'error')
            return redirect(url_for('admin_empresas_nueva'))
        
        # Validar fechas
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            
            if fecha_fin_obj <= fecha_inicio_obj:
                flash('La fecha de fin debe ser posterior a la fecha de inicio', 'error')
                return redirect(url_for('admin_empresas_nueva'))
        except ValueError:
            flash('Fechas inválidas', 'error')
            return redirect(url_for('admin_empresas_nueva'))
        
        # Verificar que no exista
        existe = execute_query('SELECT id FROM empresas WHERE nombre = %s', (nombre,))
        if existe:
            flash('Ya existe una empresa con ese nombre', 'error')
            return redirect(url_for('admin_empresas_nueva'))
        
        # Crear empresa - Verificar si existe columna tipo_empresa
        try:
            # Intentar insertar con tipo_empresa
            execute_update('''
                INSERT INTO empresas (
                    nombre, razon_social, rnc, telefono, email, direccion,
                    fecha_inicio, fecha_fin,
                    licencias_totales, licencias_usadas, plan, estado, tipo_empresa,
                    creado_por
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0, %s, 'activo', %s, %s)
            ''', (nombre, razon_social, rnc, telefono, email, direccion, 
                  fecha_inicio, fecha_fin, licencias_totales, plan, tipo_empresa, current_user.id))
        except Exception as e:
            error_msg = str(e).lower()
            if 'unknown column' in error_msg and 'tipo_empresa' in error_msg:
                # Si no existe la columna, crear sin tipo_empresa y mostrar advertencia
                logger.warning("Columna tipo_empresa no existe en tabla empresas. Ejecute el script de migración.")
                execute_update('''
                    INSERT INTO empresas (
                        nombre, razon_social, rnc, telefono, email, direccion,
                        fecha_inicio, fecha_fin,
                        licencias_totales, licencias_usadas, plan, estado,
                        creado_por
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0, %s, 'activo', %s)
                ''', (nombre, razon_social, rnc, telefono, email, direccion, 
                      fecha_inicio, fecha_fin, licencias_totales, plan, current_user.id))
                flash(f'Empresa {nombre} creada exitosamente. NOTA: Ejecute el script de migración para agregar el campo tipo_empresa.', 'warning')
            else:
                raise
        
        flash(f'Empresa "{nombre}" creada exitosamente', 'success')
        return redirect(url_for('admin_empresas'))
    
    return render_template('admin/empresas/form.html', empresa=None)

@app.route('/admin/empresas/editar/<int:empresa_id>', methods=['GET', 'POST'])
@login_required
def admin_empresas_editar(empresa_id):
    """Editar empresa - Super Admin puede editar cualquier empresa, Administrador solo la suya"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    
    empresa = execute_query('SELECT * FROM empresas WHERE id = %s', (empresa_id,))
    
    if not empresa:
        flash('Empresa no encontrada', 'error')
        return redirect(url_for('admin_empresas'))
    
    # Si el usuario tiene tenant_id (no es Super Admin), solo puede editar su propia empresa
    if tenant_id is not None and empresa['id'] != tenant_id:
        flash('No tienes permisos para editar esta empresa. Solo puedes editar tu propia empresa.', 'error')
        return redirect(url_for('admin_empresas'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 255)
        razon_social = sanitize_input(request.form.get('razon_social', ''), 255)
        rnc = sanitize_input(request.form.get('rnc', ''), 20)
        telefono = sanitize_input(request.form.get('telefono', ''), 20)
        email = request.form.get('email', '').strip().lower()
        direccion = sanitize_input(request.form.get('direccion', ''), 500)
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')
        licencias_totales = validate_int(request.form.get('licencias_totales', 5), min_value=1, max_value=1000, default=5)
        plan = request.form.get('plan', 'basico')
        estado = request.form.get('estado', 'activo')
        tipo_empresa = request.form.get('tipo_empresa', '').strip()
        
        # Validar tipo_empresa - debe ser obligatorio
        if not tipo_empresa:
            flash('El tipo de empresa es obligatorio', 'error')
            return redirect(url_for('admin_empresas_editar', empresa_id=empresa_id))
        
        # Validar que sea un valor válido
        if tipo_empresa not in ['medico', 'centro_salud']:
            flash('Tipo de empresa inválido', 'error')
            return redirect(url_for('admin_empresas_editar', empresa_id=empresa_id))
        
        # Validar fechas
        if fecha_inicio and fecha_fin:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                
                if fecha_fin_obj <= fecha_inicio_obj:
                    flash('La fecha de fin debe ser posterior a la fecha de inicio', 'error')
                    return redirect(url_for('admin_empresas_editar', empresa_id=empresa_id))
                
                # Si la fecha ya venció y está activo, cambiar a suspendido
                from datetime import date
                if fecha_fin_obj < date.today() and estado == 'activo':
                    estado = 'suspendido'
                    flash('La fecha de fin ya venció. El estado se cambió a "suspendido" automáticamente.', 'warning')
            except ValueError:
                flash('Fechas inválidas', 'error')
                return redirect(url_for('admin_empresas_editar', empresa_id=empresa_id))
        
        # Actualizar empresa - Verificar si existe columna tipo_empresa
        try:
            # Intentar actualizar con tipo_empresa (siempre incluir el campo, incluso si está vacío)
            execute_update('''
                UPDATE empresas SET
                    nombre = %s, razon_social = %s, rnc = %s,
                    telefono = %s, email = %s, direccion = %s,
                    fecha_inicio = %s, fecha_fin = %s,
                    licencias_totales = %s, plan = %s, estado = %s, tipo_empresa = %s
                WHERE id = %s
            ''', (nombre, razon_social, rnc, telefono, email, direccion,
                  fecha_inicio, fecha_fin, licencias_totales, plan, estado, tipo_empresa or None, empresa_id))
        except Exception as e:
            error_msg = str(e).lower()
            if 'unknown column' in error_msg and 'tipo_empresa' in error_msg:
                # Si no existe la columna, actualizar sin tipo_empresa
                logger.warning("Columna tipo_empresa no existe en tabla empresas. Ejecute el script de migración.")
                execute_update('''
                    UPDATE empresas SET
                        nombre = %s, razon_social = %s, rnc = %s,
                        telefono = %s, email = %s, direccion = %s,
                        fecha_inicio = %s, fecha_fin = %s,
                        licencias_totales = %s, plan = %s, estado = %s
                    WHERE id = %s
                ''', (nombre, razon_social, rnc, telefono, email, direccion,
                      fecha_inicio, fecha_fin, licencias_totales, plan, estado, empresa_id))
                flash(f'Empresa {nombre} actualizada exitosamente. NOTA: Ejecute el script de migración para agregar el campo tipo_empresa.', 'warning')
            else:
                logger.error(f"Error al actualizar empresa: {e}")
                raise
        
        flash(f'Empresa "{nombre}" actualizada exitosamente', 'success')
        return redirect(url_for('admin_empresas'))
    
    return render_template('admin/empresas/form.html', empresa=empresa)

@app.route('/admin/verificar-multitenant')
@login_required
def verificar_multitenant():
    """Verificar que el sistema multi-tenant está configurado correctamente"""
    if current_user.perfil != 'Administrador':
        return jsonify({'error': 'No tienes permisos'}), 403
    
    verificacion = {
        'titulo': 'VERIFICACIÓN SISTEMA MULTI-TENANT',
        'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'estado_general': 'OK',
        'errores': [],
        'advertencias': [],
        'detalles': {}
    }
    
    try:
        # 1. Verificar tabla empresas existe
        try:
            empresas = execute_query('SELECT COUNT(*) as total FROM empresas', fetch='one')
            verificacion['detalles']['tabla_empresas'] = {
                'existe': True,
                'total_empresas': empresas['total']
            }
        except Exception as e:
            verificacion['errores'].append(f'Tabla empresas no existe: {str(e)}')
            verificacion['estado_general'] = 'ERROR'
            return jsonify(verificacion)
        
        # 2. Verificar empresa por defecto
        empresa_default = execute_query('SELECT * FROM empresas WHERE id = 1', fetch='one')
        if empresa_default:
            verificacion['detalles']['empresa_default'] = {
                'existe': True,
                'nombre': empresa_default['nombre'],
                'licencias_totales': empresa_default['licencias_totales'],
                'licencias_usadas': empresa_default['licencias_usadas'],
                'licencias_disponibles': empresa_default['licencias_totales'] - empresa_default['licencias_usadas'],
                'plan': empresa_default['plan'],
                'estado': empresa_default['estado']
            }
        else:
            verificacion['advertencias'].append('No existe empresa con ID=1 (empresa por defecto)')
        
        # 3. Verificar columnas tenant_id en tablas
        tablas_verificar = [
            'usuarios', 'ars', 'medicos', 'codigo_ars', 
            'centros_medicos', 'medico_centro', 'servicios', 'ncf',
            'pacientes', 'facturas'
        ]
        
        columnas_ok = []
        columnas_faltantes = []
        
        for tabla in tablas_verificar:
            try:



                
                result = execute_query(f"SHOW COLUMNS FROM {tabla} LIKE 'tenant_id'", fetch='one')
                if result:
                    columnas_ok.append(tabla)
                else:
                    columnas_faltantes.append(tabla)
                    verificacion['errores'].append(f'Tabla {tabla} NO tiene columna tenant_id')
            except Exception as e:
                columnas_faltantes.append(tabla)
                verificacion['errores'].append(f'Error verificando tabla {tabla}: {str(e)}')
        
        verificacion['detalles']['columnas_tenant_id'] = {
            'total_tablas': len(tablas_verificar),
            'tablas_ok': len(columnas_ok),
            'tablas_faltantes': len(columnas_faltantes),
            'lista_ok': columnas_ok,
            'lista_faltantes': columnas_faltantes
        }
        
        # 4. Verificar usuarios con tenant_id
        usuarios_sin_tenant = execute_query('''
            SELECT COUNT(*) as total FROM usuarios 
            WHERE tenant_id IS NULL OR tenant_id = 0
        ''', fetch='one')
        
        if usuarios_sin_tenant and usuarios_sin_tenant['total'] > 0:
            verificacion['advertencias'].append(f'{usuarios_sin_tenant["total"]} usuarios sin tenant_id asignado')
        
        usuarios_por_tenant = execute_query('''
            SELECT tenant_id, COUNT(*) as total
            FROM usuarios
            WHERE activo = 1
            GROUP BY tenant_id
        ''', fetch='all')
        
        verificacion['detalles']['usuarios_por_tenant'] = [
            {'tenant_id': u['tenant_id'], 'total': u['total']}
            for u in (usuarios_por_tenant or [])
        ]
        
        # 5. Verificar triggers
        triggers = execute_query('''
            SHOW TRIGGERS WHERE `Trigger` LIKE 'trg_usuarios_%'
        ''', fetch='all')
        
        verificacion['detalles']['triggers'] = {
            'total': len(triggers) if triggers else 0,
            'esperados': 3,
            'ok': len(triggers) == 3 if triggers else False,
            'lista': [t['Trigger'] for t in (triggers or [])]
        }
        
        if not triggers or len(triggers) < 3:
            verificacion['advertencias'].append(f'Solo {len(triggers) if triggers else 0}/3 triggers encontrados')
        
        # 6. Verificar índices
        indices_tenant = execute_query('''
            SELECT TABLE_NAME, INDEX_NAME
            FROM information_schema.STATISTICS
            WHERE TABLE_SCHEMA = DATABASE()
            AND INDEX_NAME LIKE '%tenant%'
        ''', fetch='all')
        
        verificacion['detalles']['indices'] = {
            'total': len(indices_tenant) if indices_tenant else 0,
            'lista': [{'tabla': i['TABLE_NAME'], 'indice': i['INDEX_NAME']} for i in (indices_tenant or [])]
        }
        
        # 7. Estado del usuario actual
        verificacion['detalles']['usuario_actual'] = {
            'nombre': current_user.nombre,
            'email': current_user.email,
            'tenant_id': current_user.tenant_id if hasattr(current_user, 'tenant_id') else 'NO DISPONIBLE',
            'empresa_nombre': current_user.empresa_nombre if hasattr(current_user, 'empresa_nombre') else 'NO DISPONIBLE'
        }
        
        # Determinar estado general
        if verificacion['errores']:
            verificacion['estado_general'] = 'ERROR'
        elif verificacion['advertencias']:
            verificacion['estado_general'] = 'ADVERTENCIAS'
        else:
            verificacion['estado_general'] = 'PERFECTO ✅'
        
    except Exception as e:
        verificacion['estado_general'] = 'ERROR CRÍTICO'
        verificacion['errores'].append(f'Error durante verificación: {str(e)}')
    
    return jsonify(verificacion)

@app.route('/admin/verificar-multitenant-visual')
@login_required
def verificar_multitenant_visual():
    """Página visual de verificación multi-tenant"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    return render_template('admin/empresas/verificar.html')

@app.route('/admin')
@login_required
def admin():
    """Panel admin - redirige al menú de facturación"""
    return redirect(url_for('facturacion_menu'))

@app.route('/facturacion')
@login_required
def facturacion_menu():
    """Menú principal de facturación"""
    return render_template('facturacion/menu.html')

@app.route('/facturacion/ars')
@login_required
def facturacion_ars():
    """Lista de ARS - Filtrado por tenant"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    ars_list = execute_query(
        'SELECT * FROM ars WHERE tenant_id = %s ORDER BY nombre', 
        (tenant_id,), fetch='all'
    ) or []
    return render_template('facturacion/ars.html', ars_list=ars_list)

@app.route('/facturacion/ars/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_ars_nuevo():
    """Crear nueva ARS"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        nombre_ars = sanitize_input(request.form.get('nombre_ars', ''), 50)
        rnc = sanitize_input(request.form.get('rnc', ''), 50)
        activo = 1 if request.form.get('activo') == '1' else 0
        
        if not nombre_ars:
            flash('El nombre es obligatorio', 'error')
            return redirect(url_for('facturacion_ars_nuevo'))
        
        tenant_id = get_current_tenant_id()
        
        # Generar código automáticamente basado en el nombre (primeras letras + timestamp)
        import time
        codigo = ''.join(filter(str.isalnum, nombre_ars[:6].upper())) + str(int(time.time()))[-4:]
        
        # Asegurar que el código sea único
        contador = 1
        codigo_original = codigo
        while execute_query('SELECT id FROM ars WHERE codigo = %s AND tenant_id = %s', (codigo, tenant_id)):
            codigo = f"{codigo_original}{contador}"
            contador += 1
        
        execute_update('''
            INSERT INTO ars (tenant_id, codigo, nombre, rnc, activo)
            VALUES (%s, %s, %s, %s, %s)
        ''', (tenant_id, codigo, nombre_ars, rnc, activo))
        
        flash(f'ARS {nombre_ars} creada exitosamente', 'success')
        return redirect(url_for('facturacion_ars'))
    
    return render_template('facturacion/ars_form.html', ars=None)

@app.route('/facturacion/ars/<int:ars_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_ars_editar(ars_id):
    """Editar ARS"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    ars = execute_query('SELECT * FROM ars WHERE id = %s AND tenant_id = %s', (ars_id, tenant_id))
    if not ars:
        flash('ARS no encontrada', 'error')
        return redirect(url_for('facturacion_ars'))
    
    if request.method == 'POST':
        nombre_ars = sanitize_input(request.form.get('nombre_ars', ''), 50)
        rnc = sanitize_input(request.form.get('rnc', ''), 50)
        activo = 1 if request.form.get('activo') == '1' else 0
        
        if not nombre_ars:
            flash('El nombre es obligatorio', 'error')
            return redirect(url_for('facturacion_ars_editar', ars_id=ars_id))
        
        # Mantener el código existente (no se modifica en edición)
        
        execute_update('''
            UPDATE ars 
            SET nombre = %s, rnc = %s, activo = %s
            WHERE id = %s AND tenant_id = %s
        ''', (nombre_ars, rnc, activo, ars_id, tenant_id))
        
        flash(f'ARS {nombre_ars} actualizada exitosamente', 'success')
        return redirect(url_for('facturacion_ars'))
    
    return render_template('facturacion/ars_form.html', ars=ars)

@app.route('/facturacion/ars/<int:ars_id>/eliminar', methods=['POST'])
@login_required
def facturacion_ars_eliminar(ars_id):
    """Eliminar ARS"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_ars'))
    
    tenant_id = get_current_tenant_id()
    # Validar que pertenece al tenant antes de eliminar
    if not validate_tenant_access('ars', ars_id):
        flash('No tienes acceso a esta ARS', 'error')
        return redirect(url_for('facturacion_ars'))
    
    execute_update('DELETE FROM ars WHERE id = %s AND tenant_id = %s', (ars_id, tenant_id))
    flash('ARS eliminada exitosamente', 'success')
    return redirect(url_for('facturacion_ars'))

@app.route('/facturacion/medicos')
@login_required  
def facturacion_medicos():
    """Lista de médicos - Filtrado por tenant"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    medicos_list = execute_query(
        'SELECT * FROM medicos WHERE tenant_id = %s ORDER BY nombre', 
        (tenant_id,), fetch='all'
    ) or []
    return render_template('facturacion/medicos.html', medicos_list=medicos_list)

@app.route('/facturacion/medicos/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_medicos_nuevo():
    """Crear nuevo médico"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 200)
        exequatur = sanitize_input(request.form.get('exequatur', ''), 50)
        especialidad = sanitize_input(request.form.get('especialidad', ''), 100)
        telefono = sanitize_input(request.form.get('telefono', ''), 20)
        email = request.form.get('email', '').strip()
        cedula = sanitize_input(request.form.get('cedula', ''), 20)
        factura = 1 if request.form.get('factura') == '1' else 0
        
        if not nombre:
            flash('El nombre es obligatorio', 'error')
            return redirect(url_for('facturacion_medicos_nuevo'))
        
        # Validar que el teléfono sea obligatorio si está habilitado para facturar
        if factura == 1 and not telefono:
            flash('El teléfono es obligatorio cuando el médico está habilitado para facturar', 'error')
            return redirect(url_for('facturacion_medicos_nuevo'))
        
        tenant_id = get_current_tenant_id()
        execute_update('''
            INSERT INTO medicos (tenant_id, nombre, exequatur, especialidad, telefono, email, cedula, activo, factura)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 1, %s)
        ''', (tenant_id, nombre, exequatur, especialidad, telefono, email, cedula, factura))
        
        flash(f'Médico {nombre} creado exitosamente', 'success')
        return redirect(url_for('facturacion_medicos'))
    
    return render_template('facturacion/medicos_form.html', medico=None)

@app.route('/facturacion/medicos/<int:medico_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_medicos_editar(medico_id):
    """Editar médico"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    medico = execute_query('SELECT * FROM medicos WHERE id = %s AND tenant_id = %s', (medico_id, tenant_id))
    if not medico:
        flash('Médico no encontrado', 'error')
        return redirect(url_for('facturacion_medicos'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 200)
        exequatur = sanitize_input(request.form.get('exequatur', ''), 50)
        especialidad = sanitize_input(request.form.get('especialidad', ''), 100)
        telefono = sanitize_input(request.form.get('telefono', ''), 20)
        email = request.form.get('email', '').strip()
        cedula = sanitize_input(request.form.get('cedula', ''), 20)
        activo = 1 if request.form.get('activo') == '1' else 0
        factura = 1 if request.form.get('factura') == '1' else 0
        
        if not nombre:
            flash('El nombre es obligatorio', 'error')
            return redirect(url_for('facturacion_medicos_editar', medico_id=medico_id))
        
        # Validar que el teléfono sea obligatorio si está habilitado para facturar
        if factura == 1 and not telefono:
            flash('El teléfono es obligatorio cuando el médico está habilitado para facturar', 'error')
            return redirect(url_for('facturacion_medicos_editar', medico_id=medico_id))
        
        tenant_id = get_current_tenant_id()
        execute_update('''
            UPDATE medicos 
            SET nombre = %s, exequatur = %s, especialidad = %s, telefono = %s, email = %s, cedula = %s, activo = %s, factura = %s
            WHERE id = %s AND tenant_id = %s
        ''', (nombre, exequatur, especialidad, telefono, email, cedula, activo, factura, medico_id, tenant_id))
        
        flash(f'Médico {nombre} actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_medicos'))
    
    return render_template('facturacion/medicos_form.html', medico=medico)

@app.route('/facturacion/medicos/<int:medico_id>/eliminar', methods=['POST'])
@login_required
def facturacion_medicos_eliminar(medico_id):
    """Eliminar médico"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_medicos'))
    
    tenant_id = get_current_tenant_id()
    if not validate_tenant_access('medicos', medico_id):
        flash('No tienes acceso a este médico', 'error')
        return redirect(url_for('facturacion_medicos'))
    
    execute_update('DELETE FROM medicos WHERE id = %s AND tenant_id = %s', (medico_id, tenant_id))
    flash('Médico eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_medicos'))

@app.route('/facturacion/centros-medicos')
@login_required
def facturacion_centros_medicos():
    """Lista de centros médicos - Filtrado por tenant"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    search = request.args.get('search', '').strip()
    ncf_search = request.args.get('ncf', '').strip()
    
    if search:
        centros_list = execute_query(
            '''SELECT * FROM centros_medicos 
               WHERE tenant_id = %s AND (nombre LIKE %s OR rnc LIKE %s OR direccion LIKE %s)
               ORDER BY nombre''', 
            (tenant_id, f'%{search}%', f'%{search}%', f'%{search}%'), fetch='all'
        ) or []
    else:
        centros_list = execute_query(
            'SELECT * FROM centros_medicos WHERE tenant_id = %s ORDER BY nombre', 
            (tenant_id,), fetch='all'
        ) or []
    
    return render_template('facturacion/centros_medicos.html', centros_list=centros_list, search=search)

@app.route('/facturacion/centros-medicos/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_centros_medicos_nuevo():
    """Crear nuevo centro médico"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 200)
        codigo = sanitize_input(request.form.get('codigo', ''), 50)
        direccion = sanitize_input(request.form.get('direccion', ''), 500)
        rnc = sanitize_input(request.form.get('rnc', ''), 20)
        telefono = sanitize_input(request.form.get('telefono', ''), 20)
        
        if not nombre:
            flash('El nombre es obligatorio', 'error')
            return redirect(url_for('facturacion_centros_medicos_nuevo'))
        
        tenant_id = get_current_tenant_id()
        execute_update('''
            INSERT INTO centros_medicos (tenant_id, nombre, codigo, direccion, rnc, telefono, activo)
            VALUES (%s, %s, %s, %s, %s, %s, 1)
        ''', (tenant_id, nombre, codigo, direccion, rnc, telefono))
        
        flash(f'Centro médico {nombre} creado exitosamente', 'success')
        return redirect(url_for('facturacion_centros_medicos'))
    
    return render_template('facturacion/centro_medico_form.html', centro=None)

@app.route('/facturacion/centros-medicos/<int:centro_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_centros_medicos_editar(centro_id):
    """Editar centro médico"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    centro = execute_query('SELECT * FROM centros_medicos WHERE id = %s AND tenant_id = %s', (centro_id, tenant_id))
    if not centro:
        flash('Centro médico no encontrado', 'error')
        return redirect(url_for('facturacion_centros_medicos'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 200)
        codigo = sanitize_input(request.form.get('codigo', ''), 50)
        direccion = sanitize_input(request.form.get('direccion', ''), 500)
        rnc = sanitize_input(request.form.get('rnc', ''), 20)
        telefono = sanitize_input(request.form.get('telefono', ''), 20)
        activo = 1 if request.form.get('activo') == '1' else 0
        
        if not nombre:
            flash('El nombre es obligatorio', 'error')
            return redirect(url_for('facturacion_centros_medicos_editar', centro_id=centro_id))
        
        tenant_id = get_current_tenant_id()
        execute_update('''
            UPDATE centros_medicos 
            SET nombre = %s, codigo = %s, direccion = %s, rnc = %s, telefono = %s, activo = %s
            WHERE id = %s AND tenant_id = %s
        ''', (nombre, codigo, direccion, rnc, telefono, activo, centro_id, tenant_id))
        
        flash(f'Centro médico {nombre} actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_centros_medicos'))
    
    return render_template('facturacion/centro_medico_form.html', centro=centro)

@app.route('/facturacion/centros-medicos/<int:centro_id>/eliminar', methods=['POST'])
@login_required
def facturacion_centros_medicos_eliminar(centro_id):
    """Eliminar centro médico"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_centros_medicos'))
    
    tenant_id = get_current_tenant_id()
    if not validate_tenant_access('centros_medicos', centro_id):
        flash('No tienes acceso a este centro médico', 'error')
        return redirect(url_for('facturacion_centros_medicos'))
    
    execute_update('DELETE FROM centros_medicos WHERE id = %s AND tenant_id = %s', (centro_id, tenant_id))
    flash('Centro médico eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_centros_medicos'))

@app.route('/facturacion/servicios')
@login_required
def facturacion_servicios():
    """Lista de servicios - Filtrado por tenant"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    servicios_list = execute_query(
        'SELECT * FROM servicios WHERE tenant_id = %s ORDER BY descripcion', 
        (tenant_id,), fetch='all'
    ) or []
    return render_template('facturacion/servicios.html', servicios_list=servicios_list)

@app.route('/facturacion/servicios/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_servicios_nuevo():
    """Crear nuevo servicio"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        descripcion = sanitize_input(request.form.get('descripcion', ''), 15)
        precio_base = request.form.get('precio_base', '0')
        
        if not descripcion:
            flash('La descripción es obligatoria', 'error')
            return redirect(url_for('facturacion_servicios_nuevo'))
        
        try:
            precio_base = float(precio_base)
        except:
            precio_base = 0.0
        
        tenant_id = get_current_tenant_id()
        execute_update('''
            INSERT INTO servicios (tenant_id, nombre, descripcion, precio_base, activo)
            VALUES (%s, %s, %s, %s, 1)
        ''', (tenant_id, descripcion, descripcion, precio_base))
        
        flash(f'Servicio {descripcion} creado exitosamente', 'success')
        return redirect(url_for('facturacion_servicios'))
    
    return render_template('facturacion/servicios_form.html', servicio=None)

@app.route('/facturacion/servicios/<int:servicio_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_servicios_editar(servicio_id):
    """Editar servicio"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    servicio = execute_query('SELECT * FROM servicios WHERE id = %s AND tenant_id = %s', (servicio_id, tenant_id))
    if not servicio:
        flash('Servicio no encontrado', 'error')
        return redirect(url_for('facturacion_servicios'))
    
    if request.method == 'POST':
        descripcion = sanitize_input(request.form.get('descripcion', ''), 15)
        precio_base = request.form.get('precio_base', '0')
        activo = 1 if request.form.get('activo') == '1' else 0
        
        if not descripcion:
            flash('La descripción es obligatoria', 'error')
            return redirect(url_for('facturacion_servicios_editar', servicio_id=servicio_id))
        
        try:
            precio_base = float(precio_base)
        except:
            precio_base = 0.0
        
        tenant_id = get_current_tenant_id()
        execute_update('''
            UPDATE servicios 
            SET nombre = %s, descripcion = %s, precio_base = %s, activo = %s
            WHERE id = %s AND tenant_id = %s
        ''', (descripcion, descripcion, precio_base, activo, servicio_id, tenant_id))
        
        flash(f'Servicio {descripcion} actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_servicios'))
    
    return render_template('facturacion/servicios_form.html', servicio=servicio)

@app.route('/facturacion/servicios/<int:servicio_id>/eliminar', methods=['POST'])
@login_required
def facturacion_servicios_eliminar(servicio_id):
    """Eliminar servicio"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_servicios'))
    
    tenant_id = get_current_tenant_id()
    if not validate_tenant_access('servicios', servicio_id):
        flash('No tienes acceso a este servicio', 'error')
        return redirect(url_for('facturacion_servicios'))
    
    execute_update('DELETE FROM servicios WHERE id = %s AND tenant_id = %s', (servicio_id, tenant_id))
    flash('Servicio eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_servicios'))

@app.route('/facturacion/codigo-ars')
@login_required
def facturacion_codigo_ars():
    """Lista de códigos ARS - Filtrado por tenant"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    codigos_list = execute_query('''
        SELECT ca.*, a.nombre, m.nombre as nombre_medico
        FROM codigo_ars ca 
        JOIN ars a ON ca.ars_id = a.id 
        LEFT JOIN medicos m ON ca.medico_id = m.id
        WHERE ca.tenant_id = %s
        ORDER BY m.nombre, a.nombre, ca.codigo
    ''', (tenant_id,), fetch='all') or []
    return render_template('facturacion/codigo_ars.html', codigos_list=codigos_list)

@app.route('/facturacion/codigo-ars/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_codigo_ars_nuevo():
    """Crear nuevo código ARS"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        medico_id = request.form.get('medico_id')
        ars_id = request.form.get('ars_id')
        codigo = sanitize_input(request.form.get('codigo_ars', ''), 50)
        descripcion = sanitize_input(request.form.get('descripcion', ''), 500)
        precio = request.form.get('precio', '0')
        activo = 1 if request.form.get('activo') == '1' else 0
        
        if not medico_id or not ars_id or not codigo:
            flash('Médico, ARS y código son obligatorios', 'error')
            return redirect(url_for('facturacion_codigo_ars_nuevo'))
        
        try:
            precio = float(precio) if precio else 0.0
        except:
            precio = 0.0
        
        tenant_id = get_current_tenant_id()
        existe = execute_query('SELECT id FROM codigo_ars WHERE medico_id = %s AND ars_id = %s AND tenant_id = %s', (medico_id, ars_id, tenant_id))
        if existe:
            flash('Ya existe un código para este médico y ARS', 'error')
            return redirect(url_for('facturacion_codigo_ars_nuevo'))
        
        execute_update('''
            INSERT INTO codigo_ars (tenant_id, medico_id, ars_id, codigo, descripcion, precio, activo)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (tenant_id, medico_id, ars_id, codigo, descripcion or '', precio, activo))
        
        flash(f'Código ARS {codigo} creado exitosamente', 'success')
        return redirect(url_for('facturacion_codigo_ars'))
    
    tenant_id = get_current_tenant_id()
    ars_list = execute_query('SELECT * FROM ars WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    medicos_list = execute_query('SELECT * FROM medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    return render_template('facturacion/codigo_ars_form.html', codigo=None, ars_list=ars_list, medicos=medicos_list)

@app.route('/facturacion/codigo-ars/<int:codigo_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_codigo_ars_editar(codigo_id):
    """Editar código ARS"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    codigo = execute_query('SELECT * FROM codigo_ars WHERE id = %s AND tenant_id = %s', (codigo_id, tenant_id))
    if not codigo:
        flash('Código ARS no encontrado', 'error')
        return redirect(url_for('facturacion_codigo_ars'))
    
    if request.method == 'POST':
        medico_id = request.form.get('medico_id')
        ars_id = request.form.get('ars_id')
        codigo_texto = sanitize_input(request.form.get('codigo_ars', ''), 50)
        descripcion = sanitize_input(request.form.get('descripcion', ''), 500)
        precio = request.form.get('precio', '0')
        activo = 1 if request.form.get('activo') == '1' else 0
        
        if not medico_id or not ars_id or not codigo_texto:
            flash('Médico, ARS y código son obligatorios', 'error')
            return redirect(url_for('facturacion_codigo_ars_editar', codigo_id=codigo_id))
        
        try:
            precio = float(precio) if precio else 0.0
        except:
            precio = 0.0
        
        tenant_id = get_current_tenant_id()
        existe = execute_query('SELECT id FROM codigo_ars WHERE medico_id = %s AND ars_id = %s AND id != %s AND tenant_id = %s', 
                              (medico_id, ars_id, codigo_id, tenant_id))
        if existe:
            flash('Ya existe un código para este médico y ARS', 'error')
            return redirect(url_for('facturacion_codigo_ars_editar', codigo_id=codigo_id))
        
        execute_update('''
            UPDATE codigo_ars 
            SET medico_id = %s, ars_id = %s, codigo = %s, descripcion = %s, precio = %s, activo = %s
            WHERE id = %s AND tenant_id = %s
        ''', (medico_id, ars_id, codigo_texto, descripcion or '', precio, activo, codigo_id, tenant_id))
        
        flash(f'Código ARS actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_codigo_ars'))
    
    tenant_id = get_current_tenant_id()
    ars_list = execute_query('SELECT * FROM ars WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    medicos_list = execute_query('SELECT * FROM medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    return render_template('facturacion/codigo_ars_form.html', codigo=codigo, ars_list=ars_list, medicos=medicos_list)

@app.route('/facturacion/codigo-ars/<int:codigo_id>/eliminar', methods=['POST'])
@login_required
def facturacion_codigo_ars_eliminar(codigo_id):
    """Eliminar código ARS"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_codigo_ars'))
    
    tenant_id = get_current_tenant_id()
    if not validate_tenant_access('codigo_ars', codigo_id):
        flash('No tienes acceso a este código ARS', 'error')
        return redirect(url_for('facturacion_codigo_ars'))
    
    execute_update('DELETE FROM codigo_ars WHERE id = %s AND tenant_id = %s', (codigo_id, tenant_id))
    flash('Código ARS eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_codigo_ars'))

@app.route('/facturacion/medico-centro')
@login_required
def facturacion_medico_centro():
    """Relación médico-centro - Filtrado por tenant"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    relaciones_list = execute_query('''
        SELECT mc.*, m.nombre as medico_nombre, m.especialidad, c.nombre as centro_nombre
        FROM medico_centro mc
        JOIN medicos m ON mc.medico_id = m.id
        JOIN centros_medicos c ON mc.centro_medico_id = c.id
        WHERE mc.tenant_id = %s
        ORDER BY m.nombre, c.nombre
    ''', (tenant_id,), fetch='all') or []
    return render_template('facturacion/medico_centro.html', relaciones_list=relaciones_list)

@app.route('/facturacion/medico-centro/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_medico_centro_nuevo():
    """Crear nueva relación médico-centro"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        medico_id = request.form.get('medico_id')
        centro_medico_id = request.form.get('centro_medico_id')
        es_defecto = 1 if request.form.get('es_defecto') == '1' else 0
        
        if not medico_id or not centro_medico_id:
            flash('Médico y centro médico son obligatorios', 'error')
            return redirect(url_for('facturacion_medico_centro_nuevo'))
        
        tenant_id = get_current_tenant_id()
        existe = execute_query('SELECT id FROM medico_centro WHERE medico_id = %s AND centro_medico_id = %s AND tenant_id = %s', 
                              (medico_id, centro_medico_id, tenant_id))
        if existe:
            flash('Esta relación ya existe', 'error')
            return redirect(url_for('facturacion_medico_centro_nuevo'))
        
        # Si se marca como por defecto, desmarcar otros centros por defecto de este médico
        if es_defecto:
            execute_update('''
                UPDATE medico_centro 
                SET es_defecto = 0 
                WHERE medico_id = %s AND tenant_id = %s
            ''', (medico_id, tenant_id))
        
        execute_update('''
            INSERT INTO medico_centro (tenant_id, medico_id, centro_medico_id, es_defecto)
            VALUES (%s, %s, %s, %s)
        ''', (tenant_id, medico_id, centro_medico_id, es_defecto))
        
        flash('Relación médico-centro creada exitosamente', 'success')
        return redirect(url_for('facturacion_medico_centro'))
    
    tenant_id = get_current_tenant_id()
    
    # Cargar médicos y centros ACTIVOS del tenant
    medicos = execute_query('SELECT * FROM medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    centros = execute_query('SELECT * FROM centros_medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    
    # Si no hay médicos, verificar si hay inactivos
    if not medicos:
        medicos_inactivos = execute_query('SELECT nombre FROM medicos WHERE activo = 0 AND tenant_id = %s', (tenant_id,), fetch='all') or []
        if medicos_inactivos:
            nombres = ', '.join([m['nombre'] for m in medicos_inactivos])
            flash(f'No hay médicos ACTIVOS. Tienes médicos INACTIVOS: {nombres}. Ve a la lista de médicos para activarlos.', 'warning')
        else:
            flash('No hay médicos registrados. Por favor, crea un médico primero.', 'warning')
    
    if not centros:
        centros_inactivos = execute_query('SELECT nombre FROM centros_medicos WHERE activo = 0 AND tenant_id = %s', (tenant_id,), fetch='all') or []
        if centros_inactivos:
            nombres = ', '.join([c['nombre'] for c in centros_inactivos])
            flash(f'No hay centros médicos ACTIVOS. Tienes centros INACTIVOS: {nombres}. Ve a la lista de centros para activarlos.', 'warning')
        else:
            flash('No hay centros médicos registrados. Por favor, crea un centro médico primero.', 'warning')
    
    return render_template('facturacion/medico_centro_form.html', relacion=None, medicos=medicos, centros=centros)

@app.route('/facturacion/medico-centro/<int:relacion_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_medico_centro_editar(relacion_id):
    """Editar relación médico-centro"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    
    # Obtener la relación actual
    relacion = execute_query('SELECT * FROM medico_centro WHERE id = %s AND tenant_id = %s', (relacion_id, tenant_id))
    if not relacion:
        flash('Relación no encontrada', 'error')
        return redirect(url_for('facturacion_medico_centro'))
    
    if request.method == 'POST':
        medico_id = request.form.get('medico_id')
        centro_medico_id = request.form.get('centro_medico_id')
        es_defecto = 1 if request.form.get('es_defecto') == '1' else 0
        
        if not medico_id or not centro_medico_id:
            flash('Médico y centro médico son obligatorios', 'error')
            return redirect(url_for('facturacion_medico_centro_editar', relacion_id=relacion_id))
        
        # Verificar si ya existe otra relación con estos valores (excluyendo la actual)
        existe = execute_query('''
            SELECT id FROM medico_centro 
            WHERE medico_id = %s AND centro_medico_id = %s AND tenant_id = %s AND id != %s
        ''', (medico_id, centro_medico_id, tenant_id, relacion_id))
        
        if existe:
            flash('Ya existe otra relación con este médico y centro médico', 'error')
            return redirect(url_for('facturacion_medico_centro_editar', relacion_id=relacion_id))
        
        # Si se marca como por defecto, desmarcar otros centros por defecto de este médico
        if es_defecto:
            execute_update('''
                UPDATE medico_centro 
                SET es_defecto = 0 
                WHERE medico_id = %s AND tenant_id = %s AND id != %s
            ''', (medico_id, tenant_id, relacion_id))
        
        # Actualizar la relación
        execute_update('''
            UPDATE medico_centro 
            SET medico_id = %s, centro_medico_id = %s, es_defecto = %s
            WHERE id = %s AND tenant_id = %s
        ''', (medico_id, centro_medico_id, es_defecto, relacion_id, tenant_id))
        
        flash('Relación actualizada exitosamente', 'success')
        return redirect(url_for('facturacion_medico_centro'))
    
    # Cargar médicos y centros ACTIVOS del tenant
    medicos = execute_query('SELECT * FROM medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    centros = execute_query('SELECT * FROM centros_medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    
    return render_template('facturacion/medico_centro_form.html', relacion=relacion, medicos=medicos, centros=centros)

@app.route('/facturacion/medico-centro/<int:relacion_id>/eliminar', methods=['POST'])
@login_required
def facturacion_medico_centro_eliminar(relacion_id):
    """Eliminar relación médico-centro"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_medico_centro'))
    
    tenant_id = get_current_tenant_id()
    if not validate_tenant_access('medico_centro', relacion_id):
        flash('No tienes acceso a esta relación', 'error')
        return redirect(url_for('facturacion_medico_centro'))
    
    execute_update('DELETE FROM medico_centro WHERE id = %s AND tenant_id = %s', (relacion_id, tenant_id))
    flash('Relación eliminada exitosamente', 'success')
    return redirect(url_for('facturacion_medico_centro'))

@app.route('/facturacion/ncf')
@login_required
def facturacion_ncf():
    """Lista de NCF - Filtrado por tenant"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    # Incluir registros con el tenant_id actual o sin tenant_id (registros antiguos)
    ncf_list = execute_query('SELECT * FROM ncf WHERE tenant_id = %s OR tenant_id IS NULL ORDER BY tipo, id DESC', (tenant_id,), fetch='all') or []
    return render_template('facturacion/ncf.html', ncf_list=ncf_list)

@app.route('/facturacion/ncf/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_ncf_nuevo():
    """Crear nuevo NCF"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        prefijo = sanitize_input(request.form.get('prefijo', ''), 20)
        tamano_secuencia = request.form.get('tamano_secuencia', '8')
        ultimo_numero = request.form.get('ultimo_numero', '0')
        fecha_fin = request.form.get('fecha_fin')
        
        if not all([tipo, prefijo, tamano_secuencia]):
            flash('Tipo, Prefijo y Tamaño son obligatorios', 'error')
            return redirect(url_for('facturacion_ncf_nuevo'))
        
        if tipo not in ['B01', 'B02', 'B14', 'B15']:
            flash('Tipo de NCF inválido', 'error')
            return redirect(url_for('facturacion_ncf_nuevo'))
        
        tenant_id = get_current_tenant_id()
        proximo_numero = int(ultimo_numero or 0) + 1
        execute_update('''
            INSERT INTO ncf (tenant_id, tipo, prefijo, ultimo_numero, proximo_numero, tamano_secuencia, fecha_fin, activo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 1)
        ''', (tenant_id, tipo, prefijo, int(ultimo_numero or 0), proximo_numero, int(tamano_secuencia or 8), fecha_fin))
        
        flash(f'NCF {tipo} creado exitosamente', 'success')
        return redirect(url_for('facturacion_ncf'))
    
    return render_template('facturacion/ncf_form.html', ncf=None)

@app.route('/facturacion/ncf/<int:ncf_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_ncf_editar(ncf_id):
    """Editar NCF"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    # Incluir registros con el tenant_id actual o sin tenant_id (registros antiguos)
    ncf = execute_query('SELECT * FROM ncf WHERE id = %s AND (tenant_id = %s OR tenant_id IS NULL)', (ncf_id, tenant_id))
    if not ncf:
        flash('NCF no encontrado', 'error')
        return redirect(url_for('facturacion_ncf'))
    
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        prefijo = sanitize_input(request.form.get('prefijo', ''), 20)
        tamano_secuencia = request.form.get('tamano_secuencia', '8')
        ultimo_numero = request.form.get('ultimo_numero', '0')
        fecha_fin = request.form.get('fecha_fin')
        activo = 1 if request.form.get('activo') == '1' else 0
        
        if not all([tipo, prefijo, tamano_secuencia]):
            flash('Tipo, Prefijo y Tamaño son obligatorios', 'error')
            return redirect(url_for('facturacion_ncf_editar', ncf_id=ncf_id))
        
        if tipo not in ['B01', 'B02', 'B14', 'B15']:
            flash('Tipo de NCF inválido', 'error')
            return redirect(url_for('facturacion_ncf_editar', ncf_id=ncf_id))
        
        tenant_id = get_current_tenant_id()
        proximo_numero = int(ultimo_numero or 0) + 1
        execute_update('''
            UPDATE ncf 
            SET tipo = %s, prefijo = %s, ultimo_numero = %s, proximo_numero = %s, tamano_secuencia = %s, fecha_fin = %s, activo = %s
            WHERE id = %s AND tenant_id = %s
        ''', (tipo, prefijo, int(ultimo_numero or 0), proximo_numero, int(tamano_secuencia or 8), fecha_fin, activo, ncf_id, tenant_id))
        
        flash(f'NCF actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_ncf'))
    
    return render_template('facturacion/ncf_form.html', ncf=ncf)

@app.route('/facturacion/ncf/<int:ncf_id>/eliminar', methods=['POST'])
@login_required
def facturacion_ncf_eliminar(ncf_id):
    """Eliminar NCF"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_ncf'))
    
    tenant_id = get_current_tenant_id()
    if not validate_tenant_access('ncf', ncf_id):
        flash('No tienes acceso a este NCF', 'error')
        return redirect(url_for('facturacion_ncf'))
    
    execute_update('DELETE FROM ncf WHERE id = %s AND tenant_id = %s', (ncf_id, tenant_id))
    flash('NCF eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_ncf'))

@app.route('/facturacion/pacientes')
@login_required
def facturacion_pacientes():
    """Lista de pacientes - Filtrado por tenant"""
    tenant_id = get_current_tenant_id()
    search = request.args.get('search', '').strip()
    
    query = '''
        SELECT p.*, a.nombre as ars_nombre 
        FROM pacientes p 
        LEFT JOIN ars a ON p.ars_id = a.id 
        WHERE p.tenant_id = %s
    '''
    params = [tenant_id]
    
    if search:
        query += ' AND (p.nombre LIKE %s OR p.nss LIKE %s OR p.cedula LIKE %s)'
        search_pattern = f'%{search}%'
        params.extend([search_pattern, search_pattern, search_pattern])
    
    query += ' ORDER BY p.nombre'
    
    pacientes_list = execute_query(query, tuple(params), fetch='all') or []
    return render_template('facturacion/pacientes.html', pacientes_list=pacientes_list, search=search)

@app.route('/facturacion/pacientes/<int:paciente_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_pacientes_editar(paciente_id):
    """Editar paciente"""
    tenant_id = get_current_tenant_id()
    paciente = execute_query('''
        SELECT p.*, a.nombre as ars_nombre 
        FROM pacientes p 
        LEFT JOIN ars a ON p.ars_id = a.id 
        WHERE p.id = %s AND p.tenant_id = %s
    ''', (paciente_id, tenant_id))
    
    if not paciente:
        flash('Paciente no encontrado', 'error')
        return redirect(url_for('facturacion_pacientes'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 200)
        cedula = sanitize_input(request.form.get('cedula', ''), 20)
        nss = sanitize_input(request.form.get('nss', ''), 50)
        telefono = sanitize_input(request.form.get('telefono', ''), 20)
        email = request.form.get('email', '').strip().lower()
        direccion = request.form.get('direccion', '').strip()
        fecha_nacimiento = request.form.get('fecha_nacimiento') or None
        sexo = request.form.get('sexo') or None
        ars_id = request.form.get('ars_id') or None
        tipo_afiliacion = request.form.get('tipo_afiliacion') or None
        
        if not nombre:
            flash('El nombre es obligatorio', 'error')
            return redirect(url_for('facturacion_pacientes_editar', paciente_id=paciente_id))
        
        if ars_id:
            ars_id = int(ars_id)
        else:
            ars_id = None
        
        execute_update('''
            UPDATE pacientes 
            SET nombre = %s, cedula = %s, nss = %s, telefono = %s, email = %s, 
                direccion = %s, fecha_nacimiento = %s, sexo = %s, ars_id = %s, tipo_afiliacion = %s
            WHERE id = %s AND tenant_id = %s
        ''', (nombre, cedula or None, nss or None, telefono or None, email or None, 
              direccion or None, fecha_nacimiento, sexo, ars_id, tipo_afiliacion, paciente_id, tenant_id))
        
        flash('Paciente actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_pacientes'))
    
    # Obtener lista de ARS para el dropdown
    ars_list = execute_query('SELECT * FROM ars WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    
    return render_template('facturacion/paciente_form.html', paciente=paciente, ars_list=ars_list)

@app.route('/facturacion/pacientes/<int:paciente_id>/eliminar', methods=['POST'])
@login_required
def facturacion_pacientes_eliminar(paciente_id):
    """Eliminar paciente"""
    tenant_id = get_current_tenant_id()
    
    # Verificar que el paciente existe y pertenece al tenant
    paciente = execute_query('SELECT id FROM pacientes WHERE id = %s AND tenant_id = %s', (paciente_id, tenant_id))
    if not paciente:
        flash('Paciente no encontrado', 'error')
        return redirect(url_for('facturacion_pacientes'))
    
    # Eliminar el paciente
    execute_update('DELETE FROM pacientes WHERE id = %s AND tenant_id = %s', (paciente_id, tenant_id))
    
    flash('Paciente eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_pacientes'))

@app.route('/facturacion/pacientes-pendientes/<int:paciente_id>/eliminar', methods=['POST'])
@login_required
def facturacion_pacientes_pendientes_eliminar(paciente_id):
    """Eliminar paciente pendiente"""
    tenant_id = get_current_tenant_id()
    
    # Verificar si existe la columna tenant_id en pacientes_pendientes
    tiene_tenant_id = False
    try:
        check_tenant = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'pacientes_pendientes' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id = check_tenant and check_tenant.get('count', 0) > 0
    except:
        tiene_tenant_id = False
    
    # Verificar que el registro existe
    if tiene_tenant_id:
        paciente = execute_query('SELECT id FROM pacientes_pendientes WHERE id = %s AND tenant_id = %s', (paciente_id, tenant_id))
        if not paciente:
            flash('Registro no encontrado', 'error')
            return redirect(url_for('facturacion_pacientes_pendientes'))
        
        # Eliminar el registro
        execute_update('DELETE FROM pacientes_pendientes WHERE id = %s AND tenant_id = %s', (paciente_id, tenant_id))
    else:
        # Si no existe tenant_id, eliminar sin verificar tenant
        paciente = execute_query('SELECT id FROM pacientes_pendientes WHERE id = %s', (paciente_id,))
        if not paciente:
            flash('Registro no encontrado', 'error')
            return redirect(url_for('facturacion_pacientes_pendientes'))
        
        # Eliminar el registro
        execute_update('DELETE FROM pacientes_pendientes WHERE id = %s', (paciente_id,))
    
    flash('Registro eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_pacientes_pendientes'))

@app.route('/api/facturacion/pacientes-pendientes/<int:paciente_id>', methods=['GET'])
@login_required
def api_facturacion_pacientes_pendientes_get(paciente_id):
    """Obtener datos de un paciente pendiente para editar"""
    tenant_id = get_current_tenant_id()
    
    # Verificar si existe la columna tenant_id
    tiene_tenant_id = False
    try:
        check_tenant = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'pacientes_pendientes' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id = check_tenant and check_tenant.get('count', 0) > 0
    except:
        tiene_tenant_id = False
    
    # Obtener el registro
    if tiene_tenant_id:
        paciente = execute_query('''
            SELECT pp.*, a.nombre as ars_nombre, m.nombre as medico_nombre
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            LEFT JOIN medicos m ON pp.medico_id = m.id
            WHERE pp.id = %s AND pp.tenant_id = %s
        ''', (paciente_id, tenant_id))
    else:
        paciente = execute_query('''
            SELECT pp.*, a.nombre as ars_nombre, m.nombre as medico_nombre
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            LEFT JOIN medicos m ON pp.medico_id = m.id
            WHERE pp.id = %s
        ''', (paciente_id,))
    
    if not paciente:
        return jsonify({'error': 'Registro no encontrado'}), 404
    
    # Extraer solo el servicio sin la autorización
    servicio_completo = paciente.get('servicios_realizados', '') or ''
    servicio = servicio_completo.split(' - Autorización:')[0].strip() if ' - Autorización:' in servicio_completo else servicio_completo.strip()
    autorizacion = ''
    if ' - Autorización:' in servicio_completo:
        partes = servicio_completo.split(' - Autorización:')
        if len(partes) > 1:
            autorizacion = partes[1].strip()
    
    # Formatear fecha para input type="date" (YYYY-MM-DD)
    fecha_servicio = paciente.get('fecha_servicio', '')
    if fecha_servicio:
        if isinstance(fecha_servicio, str):
            # Si es string, verificar formato y convertir si es necesario
            if '/' in fecha_servicio:
                # Formato MM/DD/YYYY o DD/MM/YYYY
                partes = fecha_servicio.split('/')
                if len(partes) == 3:
                    fecha_servicio = f"{partes[2]}-{partes[0].zfill(2)}-{partes[1].zfill(2)}"
            elif fecha_servicio.count('-') == 2 and len(fecha_servicio.split('-')[0]) == 2:
                # Formato DD-MM-YYYY
                partes = fecha_servicio.split('-')
                fecha_servicio = f"{partes[2]}-{partes[1]}-{partes[0]}"
        else:
            # Si es objeto date/datetime, convertir a string YYYY-MM-DD
            from datetime import date, datetime
            if isinstance(fecha_servicio, (date, datetime)):
                fecha_servicio = fecha_servicio.strftime('%Y-%m-%d')
    
    return jsonify({
        'id': paciente['id'],
        'nombre_paciente': paciente.get('nombre_paciente', ''),
        'nss': paciente.get('nss', ''),
        'fecha_servicio': fecha_servicio,
        'servicio': servicio,
        'autorizacion': autorizacion,
        'monto_estimado': float(paciente.get('monto_estimado', 0)),
        'ars_id': paciente.get('ars_id'),
        'ars_nombre': paciente.get('ars_nombre', ''),
        'medico_id': paciente.get('medico_id'),
        'medico_nombre': paciente.get('medico_nombre', ''),
        'centro_medico_id': paciente.get('centro_medico_id'),
        'observaciones': paciente.get('observaciones', '')
    })

@app.route('/api/facturacion/pacientes-pendientes/<int:paciente_id>', methods=['PUT'])
@login_required
def api_facturacion_pacientes_pendientes_update(paciente_id):
    """Actualizar un paciente pendiente"""
    try:
        tenant_id = get_current_tenant_id()
        
        if not request.is_json:
            return jsonify({'error': 'Content-Type debe ser application/json'}), 400
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No se recibieron datos'}), 400
        
        # Validar datos
        nombre_paciente = sanitize_input(data.get('nombre_paciente', ''), 200)
        nss = sanitize_input(data.get('nss', ''), 50)
        fecha_servicio = data.get('fecha_servicio', '')
        servicio_completo = sanitize_input(data.get('servicio', ''), 500)  # Ya viene con autorización si existe
        
        try:
            monto_estimado = float(data.get('monto_estimado', 0))
        except (ValueError, TypeError):
            monto_estimado = 0.0
        
        ars_id = data.get('ars_id') or None
        medico_id = data.get('medico_id') or None
        centro_medico_id = data.get('centro_medico_id') or None
        
        # Manejar observaciones de forma segura
        observaciones = data.get('observaciones')
        if observaciones:
            observaciones = str(observaciones).strip()
            if not observaciones:
                observaciones = None
        else:
            observaciones = None
        
        if not nombre_paciente:
            return jsonify({'error': 'El nombre del paciente es obligatorio'}), 400
        
        if not fecha_servicio:
            return jsonify({'error': 'La fecha de servicio es obligatoria'}), 400
        
        # El servicio ya viene completo con autorización desde el frontend
        servicios_realizados = servicio_completo
        
        # Verificar si existe la columna tenant_id
        tiene_tenant_id = False
        try:
            check_tenant = execute_query('''
                SELECT COUNT(*) as count 
                FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE()
                AND TABLE_NAME = 'pacientes_pendientes' 
                AND COLUMN_NAME = 'tenant_id'
            ''')
            tiene_tenant_id = check_tenant and check_tenant.get('count', 0) > 0
        except Exception as e:
            print(f"Error al verificar tenant_id: {e}")
            tiene_tenant_id = False
        
        # Convertir IDs a enteros si existen
        try:
            if ars_id:
                ars_id = int(ars_id)
        except (ValueError, TypeError):
            ars_id = None
            
        try:
            if medico_id:
                medico_id = int(medico_id)
        except (ValueError, TypeError):
            medico_id = None
            
        try:
            if centro_medico_id:
                centro_medico_id = int(centro_medico_id)
        except (ValueError, TypeError):
            centro_medico_id = None
        
        # Actualizar
        if tiene_tenant_id:
            execute_update('''
                UPDATE pacientes_pendientes 
                SET nombre_paciente = %s, nss = %s, fecha_servicio = %s, 
                    servicios_realizados = %s, monto_estimado = %s, 
                    ars_id = %s, medico_id = %s, centro_medico_id = %s, observaciones = %s
                WHERE id = %s AND tenant_id = %s
            ''', (nombre_paciente, nss or None, fecha_servicio, servicios_realizados, monto_estimado,
                  ars_id, medico_id, centro_medico_id, observaciones, paciente_id, tenant_id))
        else:
            execute_update('''
                UPDATE pacientes_pendientes 
                SET nombre_paciente = %s, nss = %s, fecha_servicio = %s, 
                    servicios_realizados = %s, monto_estimado = %s, 
                    ars_id = %s, medico_id = %s, centro_medico_id = %s, observaciones = %s
                WHERE id = %s
            ''', (nombre_paciente, nss or None, fecha_servicio, servicios_realizados, monto_estimado,
                  ars_id, medico_id, centro_medico_id, observaciones, paciente_id))
        
        return jsonify({'success': True, 'message': 'Registro actualizado exitosamente'})
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error al actualizar paciente pendiente: {error_trace}")
        return jsonify({'error': f'Error al actualizar: {str(e)}'}), 500
    
    return jsonify({'success': True, 'message': 'Registro actualizado exitosamente'})

@app.route('/facturacion/reclamaciones')
@login_required
def facturacion_reclamaciones():
    """Lista de reclamaciones - Filtrado por tenant"""
    tenant_id = get_current_tenant_id()
    ncf = request.args.get('ncf', '').strip()
    ars_id_filter = request.args.get('ars_id')
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()
    
    query = '''
        SELECT r.*, f.numero_factura, f.nombre_paciente, f.nombre_ars, f.total as total_factura, f.ncf, a.rnc as ars_rnc
        FROM reclamaciones r
        JOIN facturas f ON r.factura_id = f.id
        LEFT JOIN ars a ON f.ars_id = a.id
        WHERE r.tenant_id = %s
    '''
    params = [tenant_id]
    
    if ncf:
        query += ' AND f.ncf LIKE %s'
        params.append(f'%{ncf}%')
    if ars_id_filter:
        query += ' AND f.ars_id = %s'
        params.append(ars_id_filter)
    if fecha_desde:
        query += ' AND r.fecha_reclamacion >= %s'
        params.append(fecha_desde)
    if fecha_hasta:
        query += ' AND r.fecha_reclamacion <= %s'
        params.append(fecha_hasta)
    
    query += ' ORDER BY r.fecha_reclamacion DESC, r.id DESC'
    
    reclamaciones_list = execute_query(query, tuple(params), fetch='all') or []
    ars_list = execute_query('SELECT id, nombre, rnc FROM ars WHERE tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    return render_template('facturacion/reclamaciones.html', reclamaciones_list=reclamaciones_list,
                           ncf=ncf, ars_id=ars_id_filter, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, ars_list=ars_list)

@app.route('/facturacion/reclamaciones/<int:reclamacion_id>')
@login_required
def facturacion_reclamacion_detalle(reclamacion_id: int):
    """Detalle de una reclamación"""
    tenant_id = get_current_tenant_id()
    reclamacion = execute_query('''
        SELECT r.*, 
               f.numero_factura, f.ncf, f.nombre_paciente, f.nombre_ars, f.total as total_factura,
               f.fecha_emision, f.estado as estado_factura
        FROM reclamaciones r
        JOIN facturas f ON r.factura_id = f.id
        WHERE r.id = %s AND r.tenant_id = %s
    ''', (reclamacion_id, tenant_id))
    
    if not reclamacion:
        flash('Reclamación no encontrada', 'error')
        return redirect(url_for('facturacion_reclamaciones'))
    
    return render_template('facturacion/reclamacion_detalle.html', reclamacion=reclamacion)

@app.route('/facturacion/reclamaciones/<int:reclamacion_id>/estado', methods=['POST'])
@login_required
def facturacion_reclamacion_cambiar_estado(reclamacion_id: int):
    """Actualizar estado de reclamación"""
    tenant_id = get_current_tenant_id()
    nuevo_estado = request.form.get('estado')
    observacion_estado = request.form.get('observaciones_estado', '').strip()
    if nuevo_estado not in ['Pendiente', 'Procesada', 'Rechazada']:
        flash('Estado inválido', 'error')
        return redirect(url_for('facturacion_reclamacion_detalle', reclamacion_id=reclamacion_id))
    
    # Verificar pertenencia al tenant y estado actual
    reclamacion_actual = execute_query('SELECT id, estado, factura_id FROM reclamaciones WHERE id = %s AND tenant_id = %s', (reclamacion_id, tenant_id))
    if not reclamacion_actual:
        flash('Reclamación no encontrada', 'error')
        return redirect(url_for('facturacion_reclamaciones'))
    
    estado_actual = reclamacion_actual.get('estado')
    factura_id = reclamacion_actual.get('factura_id')
    if estado_actual == 'Procesada' and nuevo_estado != 'Procesada':
        flash('Una reclamación procesada no puede volver a Pendiente o Rechazada.', 'error')
        return redirect(url_for('facturacion_reclamacion_detalle', reclamacion_id=reclamacion_id))
    
    # Si se intenta marcar como procesada, validar que exista pago asociado a la factura
    if nuevo_estado == 'Procesada' and factura_id:
        pagos_count = execute_query('''
            SELECT COUNT(*) as count
            FROM pago_facturas pf
            JOIN pagos p ON pf.pago_id = p.id
            WHERE pf.factura_id = %s AND p.tenant_id = %s
        ''', (factura_id, tenant_id))
        if not pagos_count or pagos_count.get('count', 0) == 0:
            flash('Debes registrar un pago de la factura antes de marcar la reclamación como Procesada.', 'error')
            return redirect(url_for('facturacion_pagos_nuevo', factura_id=factura_id))
    
    # Si se rechaza, exigir comentario
    if nuevo_estado == 'Rechazada' and not observacion_estado:
        flash('Debes ingresar una observación al rechazar la reclamación.', 'error')
        return redirect(url_for('facturacion_reclamacion_detalle', reclamacion_id=reclamacion_id))
    
    if nuevo_estado == 'Rechazada':
        execute_update('UPDATE reclamaciones SET estado = %s, observaciones = %s WHERE id = %s AND tenant_id = %s',
                       (nuevo_estado, observacion_estado or None, reclamacion_id, tenant_id))
    else:
        execute_update('UPDATE reclamaciones SET estado = %s WHERE id = %s AND tenant_id = %s',
                       (nuevo_estado, reclamacion_id, tenant_id))
    flash('Estado actualizado', 'success')
    if nuevo_estado == 'Procesada':
        return redirect(url_for('facturacion_pagos_nuevo', factura_id=factura_id))
    return redirect(url_for('facturacion_reclamacion_detalle', reclamacion_id=reclamacion_id))

@app.route('/facturacion/reclamaciones/<int:reclamacion_id>/pdf')
@login_required
def facturacion_reclamacion_pdf(reclamacion_id: int):
    """Descargar PDF de la reclamación"""
    if not REPORTLAB_AVAILABLE:
        flash('ReportLab no está disponible', 'error')
        return redirect(url_for('facturacion_reclamacion_detalle', reclamacion_id=reclamacion_id))
    
    tenant_id = get_current_tenant_id()
    reclamacion = execute_query('''
        SELECT r.*, 
               f.numero_factura, f.ncf, f.nombre_paciente, f.nombre_ars, f.total as total_factura,
               f.fecha_emision, f.estado as estado_factura,
               f.nombre_medico, f.nombre_centro_medico
        FROM reclamaciones r
        JOIN facturas f ON r.factura_id = f.id
        WHERE r.id = %s AND r.tenant_id = %s
    ''', (reclamacion_id, tenant_id))
    
    if not reclamacion:
        flash('Reclamación no encontrada', 'error')
        return redirect(url_for('facturacion_reclamaciones'))
    
    # Datos para footer similar a factura
    medico_nombre = reclamacion.get('nombre_medico') or ''
    centro_nombre = reclamacion.get('nombre_centro_medico') or ''
    empresa_info = get_empresa_info(tenant_id)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.6 * inch,
        bottomMargin=0.7 * inch,
    )
    story = []
    styles = getSampleStyleSheet()
    primary_color = colors.HexColor('#111111')
    accent_color = colors.HexColor('#0077b6')
    subtle_gray = colors.HexColor('#f2f2f2')
    
    # Header
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=18, fontName='Helvetica-Bold', textColor=primary_color)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#555'))
    header_table = Table([
        [Paragraph(f"RECLAMACIÓN #{reclamacion_id}", header_style),
         Paragraph(f"Fecha reclamación: {escape(str(reclamacion.get('fecha_reclamacion') or ''))}", sub_style)],
        [Paragraph(f"NCF: {escape(str(reclamacion.get('ncf') or reclamacion.get('numero_factura') or ''))}", sub_style),
         Paragraph(f"Estado: {escape(str(reclamacion.get('estado') or ''))}", sub_style)],
    ], colWidths=[3.7*inch, 3.7*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ('ALIGN', (1,1), (1,1), 'RIGHT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(header_table)
    line = Table([['']], colWidths=[7.4*inch])
    line.setStyle(TableStyle([('LINEBELOW', (0,0), (-1,-1), 1, primary_color)]))
    story.append(line)
    story.append(Spacer(1, 0.2*inch))
    
    # Info blocks
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#222'))
    bold_style = ParagraphStyle('Bold', parent=info_style, fontName='Helvetica-Bold')
    card_data = [
        [Paragraph('<b>Factura</b>', bold_style), Paragraph('<b>Reclamación</b>', bold_style)],
        [
            Paragraph(f"NCF: {escape(str(reclamacion.get('ncf') or ''))}<br/>Fecha factura: {escape(str(reclamacion.get('fecha_emision') or ''))}<br/>ARS: {escape(str(reclamacion.get('nombre_ars') or ''))}<br/>Paciente: {escape(str(reclamacion.get('nombre_paciente') or ''))}", info_style),
            Paragraph(f"Monto reclamado: ${reclamacion.get('monto_reclamado'):,.2f}<br/>Total factura: ${reclamacion.get('total_factura'):,.2f}<br/>Estado factura: {escape(str(reclamacion.get('estado_factura') or ''))}<br/>Estado reclamación: {escape(str(reclamacion.get('estado') or ''))}", info_style),
        ]
    ]
    card = Table(card_data, colWidths=[3.7*inch, 3.7*inch])
    card.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,1), (-1,1), subtle_gray),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
        ('VALIGN', (0,1), (-1,-1), 'TOP'),
    ]))
    story.append(card)
    
    # Observaciones
    if reclamacion.get('observaciones'):
        story.append(Spacer(1, 0.25*inch))
        obs_title = ParagraphStyle('ObsTitle', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=primary_color)
        obs_body = ParagraphStyle('ObsBody', parent=styles['Normal'], fontSize=9, leading=12)
        story.append(Paragraph("Observaciones", obs_title))
        story.append(Paragraph(escape(str(reclamacion.get('observaciones'))), obs_body))
    
    story.append(Spacer(1, 0.35*inch))
    
    # Footer similar a factura
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#666'))
    footer_bold_style = ParagraphStyle('FooterBold', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, textColor=accent_color, fontName='Helvetica-Bold')
    footer_lines = []
    if medico_nombre:
        footer_lines.append(Paragraph(escape(str(medico_nombre)), footer_bold_style))
    if centro_nombre:
        footer_lines.append(Paragraph(escape(str(centro_nombre)), footer_style))
    extra = []
    if empresa_info:
        if empresa_info.get('telefono'):
            extra.append(f"Tel: {empresa_info.get('telefono')}")
        if empresa_info.get('email'):
            extra.append(f"Email: {empresa_info.get('email')}")
    if extra:
        footer_lines.append(Paragraph(' | '.join(extra), footer_style))
    
    if footer_lines:
        footer_table = Table([[line] for line in footer_lines], colWidths=[7.4*inch])
        footer_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(footer_table)
    
    doc.build(story)
    buffer.seek(0)
    filename = f"reclamacion_{reclamacion_id}.pdf"
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

@app.route('/facturacion/reclamaciones/nueva', methods=['GET', 'POST'])
@login_required
def facturacion_reclamaciones_nueva():
    """Crear nueva reclamación"""
    tenant_id = get_current_tenant_id()
    
    if request.method == 'POST':
        factura_id = request.form.get('factura_id')
        monto_reclamado = request.form.get('monto_reclamado')
        fecha_reclamacion = request.form.get('fecha_reclamacion')
        observaciones = request.form.get('observaciones', '').strip()
        
        if not all([factura_id, monto_reclamado, fecha_reclamacion]):
            flash('Factura, monto y fecha son obligatorios', 'error')
            return redirect(url_for('facturacion_reclamaciones_nueva'))
        
        # Verificar que la factura existe y pertenece al tenant
        factura = execute_query('''
            SELECT f.id, f.total, f.fecha_emision,
                   (f.total - COALESCE(SUM(pf.monto_aplicado), 0)) AS balance_pendiente
            FROM facturas f
            LEFT JOIN pago_facturas pf ON f.id = pf.factura_id
            WHERE f.id = %s AND f.tenant_id = %s AND f.estado != 'Anulada'
            GROUP BY f.id
        ''', (factura_id, tenant_id))
        if not factura:
            flash('Factura no encontrada', 'error')
            return redirect(url_for('facturacion_reclamaciones_nueva'))
        
        try:
            monto_reclamado = float(monto_reclamado)
            if monto_reclamado <= 0:
                flash('El monto debe ser mayor a cero', 'error')
                return redirect(url_for('facturacion_reclamaciones_nueva'))
            # Validar antigüedad (<= 90 días)
            fecha_emision = factura.get('fecha_emision')
            if fecha_emision:
                try:
                    fecha_emision_dt = datetime.strptime(str(fecha_emision), "%Y-%m-%d")
                    if (datetime.now() - fecha_emision_dt).days > 90:
                        flash('Solo se permiten reclamaciones de facturas con menos de 90 días.', 'error')
                        return redirect(url_for('facturacion_reclamaciones_nueva'))
                except Exception:
                    pass
            # Validar balance pendiente
            balance_pendiente = float(factura.get('balance_pendiente') or 0)
            # Sumar reclamaciones previas de la misma factura
            reclamado_previo = execute_query('''
                SELECT COALESCE(SUM(monto_reclamado), 0) AS total_reclamado
                FROM reclamaciones
                WHERE factura_id = %s AND tenant_id = %s
            ''', (factura_id, tenant_id))
            total_reclamado = float(reclamado_previo.get('total_reclamado') or 0)
            if monto_reclamado + total_reclamado > balance_pendiente:
                flash(f'El monto reclamado supera el balance pendiente considerando reclamaciones previas (${balance_pendiente:,.2f}).', 'error')
                return redirect(url_for('facturacion_reclamaciones_nueva'))
            if monto_reclamado > balance_pendiente:
                flash(f'El monto reclamado excede el balance pendiente (${balance_pendiente:,.2f}).', 'error')
                return redirect(url_for('facturacion_reclamaciones_nueva'))
        except ValueError:
            flash('Monto inválido', 'error')
            return redirect(url_for('facturacion_reclamaciones_nueva'))
        
        execute_update('''
            INSERT INTO reclamaciones (factura_id, monto_reclamado, fecha_reclamacion, observaciones, tenant_id, created_by, estado)
            VALUES (%s, %s, %s, %s, %s, %s, 'Pendiente')
        ''', (factura_id, monto_reclamado, fecha_reclamacion, observaciones or None, tenant_id, current_user.id))
        
        flash('Reclamación creada exitosamente', 'success')
        return redirect(url_for('facturacion_reclamaciones'))
    
    # Filtros (ARS requerido para listar facturas)
    search = ''  # desactivado: solo filtramos por ARS y NCF
    ncf_search = request.args.get('ncf', '').strip()
    ars_id = request.args.get('ars_id')
    
    facturas_list = []
    if ars_id:
        query = '''
            SELECT f.id, f.numero_factura, f.nombre_paciente, f.nombre_ars, f.total, f.fecha_emision, f.estado,
                   (f.total - COALESCE(SUM(pf.monto_aplicado), 0)) AS balance_pendiente,
                   (SELECT COUNT(*) FROM reclamaciones r WHERE r.factura_id = f.id AND r.tenant_id = f.tenant_id) AS reclamaciones_count,
                   (SELECT r.monto_reclamado FROM reclamaciones r WHERE r.factura_id = f.id AND r.tenant_id = f.tenant_id ORDER BY r.fecha_reclamacion DESC LIMIT 1) AS ultima_reclamacion_monto,
                   (SELECT r.fecha_reclamacion FROM reclamaciones r WHERE r.factura_id = f.id AND r.tenant_id = f.tenant_id ORDER BY r.fecha_reclamacion DESC LIMIT 1) AS ultima_reclamacion_fecha,
                   (SELECT r.id FROM reclamaciones r WHERE r.factura_id = f.id AND r.tenant_id = f.tenant_id ORDER BY r.fecha_reclamacion DESC LIMIT 1) AS ultima_reclamacion_id
        FROM facturas f
            LEFT JOIN pago_facturas pf ON f.id = pf.factura_id
            WHERE f.tenant_id = %s
              AND f.estado != 'Anulada'
              AND f.ars_id = %s
              AND f.fecha_emision >= DATE_SUB(CURDATE(), INTERVAL 90 DAY)
        '''
        params = [tenant_id, ars_id]
        if ncf_search:
            query += ' AND f.ncf LIKE %s'
            params.append(f'%{ncf_search}%')
        query += '''
            GROUP BY f.id
            HAVING (f.total - COALESCE(SUM(pf.monto_aplicado), 0)) > 0
        ORDER BY f.fecha_emision DESC, f.numero_factura DESC
        LIMIT 100
        '''
        facturas_list = execute_query(query, tuple(params), fetch='all') or []
    
    ars_list = execute_query('SELECT id, nombre FROM ars WHERE tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    
    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    return render_template('facturacion/reclamacion_form.html',
                           facturas_list=facturas_list,
                           fecha_actual=fecha_actual,
                           ars_list=ars_list,
                           search=search,
                           ncf=ncf_search,
                           ars_id=ars_id)

@app.route('/facturacion/pagos')
@login_required
def facturacion_pagos():
    """Lista de pagos - Filtrado por tenant"""
    tenant_id = get_current_tenant_id()
    ars_id_filter = request.args.get('ars_id')
    ncf_filter = request.args.get('ncf', '').strip()
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()
    
    query = '''
        SELECT 
               p.id,
               CONCAT('PAGO-', LPAD(p.id, 6, '0')) AS numero_pago,
               p.fecha_pago,
               p.metodo_pago,
               p.monto AS monto_total,
               COUNT(pf.id) AS cantidad_facturas,
               GROUP_CONCAT(f.numero_factura SEPARATOR ', ') AS facturas_numeros,
               GROUP_CONCAT(DISTINCT f.ncf SEPARATOR ', ') AS ncf_list,
               GROUP_CONCAT(DISTINCT a.nombre SEPARATOR ', ') AS ars_list
        FROM pagos p
        LEFT JOIN pago_facturas pf ON p.id = pf.pago_id
        LEFT JOIN facturas f ON pf.factura_id = f.id
        LEFT JOIN ars a ON f.ars_id = a.id
        WHERE p.tenant_id = %s
    '''
    params = [tenant_id]
    if ars_id_filter:
        query += ' AND f.ars_id = %s'
        params.append(ars_id_filter)
    if ncf_filter:
        query += ' AND f.ncf LIKE %s'
        params.append(f'%{ncf_filter}%')
    if fecha_desde:
        query += ' AND p.fecha_pago >= %s'
        params.append(fecha_desde)
    if fecha_hasta:
        query += ' AND p.fecha_pago <= %s'
        params.append(fecha_hasta)
    
    query += '''
        GROUP BY p.id
        ORDER BY p.fecha_pago DESC, p.id DESC
    '''
    pagos_list = execute_query(query, tuple(params), fetch='all') or []
    ars_list = execute_query('SELECT id, nombre FROM ars WHERE tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    return render_template('facturacion/pagos.html', pagos_list=pagos_list,
                           ars_list=ars_list, ars_id=ars_id_filter, ncf=ncf_filter,
                           fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)

@app.route('/facturacion/pagos/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_pagos_nuevo():
    """Crear nuevo pago"""
    tenant_id = get_current_tenant_id()
    factura_id_param = request.args.get('factura_id')
    
    if request.method == 'POST':
        fecha_pago = request.form.get('fecha_pago')
        metodo_pago = request.form.get('metodo_pago')
        referencia = request.form.get('referencia', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        facturas_ids = request.form.getlist('facturas_ids[]')
        montos = request.form.getlist('montos[]')
        monto_pago_ingresado = request.form.get('monto_pago')
        
        if not all([fecha_pago, metodo_pago]):
            flash('Fecha y método de pago son obligatorios', 'error')
            return redirect(url_for('facturacion_pagos_nuevo'))
        
        if not facturas_ids or not montos:
            flash('Debe seleccionar al menos una factura', 'error')
            return redirect(url_for('facturacion_pagos_nuevo'))
        
        # Validar y calcular monto total
        monto_total = 0.0
        facturas_data = []
        for i, factura_id in enumerate(facturas_ids):
            if i < len(montos) and montos[i]:
                try:
                    monto = float(montos[i])
                    if monto > 0:
                        # Verificar que la factura existe y pertenece al tenant
                        factura = execute_query('SELECT id, total FROM facturas WHERE id = %s AND tenant_id = %s', (factura_id, tenant_id))
                        if factura:
                            monto_total += monto
                            facturas_data.append((factura_id, monto))
                except ValueError:
                    continue
        
        if monto_total <= 0:
            flash('El monto total debe ser mayor a cero', 'error')
            return redirect(url_for('facturacion_pagos_nuevo'))
        
        try:
            monto_pago_num = float(monto_pago_ingresado or 0)
        except ValueError:
            flash('Monto de pago inválido', 'error')
            return redirect(url_for('facturacion_pagos_nuevo'))
        
        if monto_pago_num < monto_total:
            flash('El monto del pago no puede ser menor que el total aplicado a documentos.', 'error')
            return redirect(url_for('facturacion_pagos_nuevo'))
        
        # Seleccionar una factura principal para el registro legacy (tabla pagos requiere factura_id)
        factura_principal_id = facturas_data[0][0]
        
        # Crear el pago (schema actual: no tiene numero_pago ni monto_total)
        pago_id = execute_update('''
            INSERT INTO pagos (tenant_id, factura_id, fecha_pago, monto, metodo_pago, referencia, observaciones, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (tenant_id, factura_principal_id, fecha_pago, monto_pago_num, metodo_pago, referencia or None, observaciones or None, current_user.id))
        
        # Crear relaciones pago-facturas
        for factura_id, monto in facturas_data:
            execute_update('''
                INSERT INTO pago_facturas (pago_id, factura_id, monto_aplicado)
                VALUES (%s, %s, %s)
            ''', (pago_id, factura_id, monto))
            
            # Actualizar estado de la factura a Pagada si el monto aplicado es igual o mayor al total
            factura = execute_query('SELECT total FROM facturas WHERE id = %s', (factura_id,))
            if factura and monto >= float(factura['total']):
                execute_update('UPDATE facturas SET estado = "Pagada" WHERE id = %s', (factura_id,))
        
        flash('Pago registrado exitosamente', 'success')
        return redirect(url_for('facturacion_pagos'))
    
    # Obtener facturas disponibles para pagar, con info de reclamaciones
    filtro_factura = ""
    params_facturas = [tenant_id]
    if factura_id_param:
        filtro_factura = " AND f.id = %s "
        params_facturas.append(factura_id_param)
    facturas_list = execute_query(f'''
        SELECT f.id, f.numero_factura, f.nombre_paciente, f.nombre_ars, f.total, f.fecha_emision, f.estado,
               COALESCE(SUM(pf.monto_aplicado), 0) as monto_pagado,
               (SELECT COALESCE(SUM(r.monto_reclamado), 0) FROM reclamaciones r WHERE r.factura_id = f.id AND r.tenant_id = f.tenant_id) as monto_reclamado,
               (SELECT COALESCE(SUM(r.monto_reclamado), 0) FROM reclamaciones r WHERE r.factura_id = f.id AND r.tenant_id = f.tenant_id AND r.estado = 'Procesada') as monto_reclamado_procesado,
               (SELECT r.monto_reclamado FROM reclamaciones r WHERE r.factura_id = f.id AND r.tenant_id = f.tenant_id AND r.estado = 'Procesada' ORDER BY r.fecha_reclamacion DESC LIMIT 1) as ultima_reclamacion_monto,
               (SELECT COUNT(*) FROM reclamaciones r WHERE r.factura_id = f.id AND r.tenant_id = f.tenant_id) as reclamaciones_count,
               (SELECT r.id FROM reclamaciones r WHERE r.factura_id = f.id AND r.tenant_id = f.tenant_id ORDER BY r.fecha_reclamacion DESC LIMIT 1) as ultima_reclamacion_id
        FROM facturas f
        LEFT JOIN pago_facturas pf ON f.id = pf.factura_id
        WHERE f.tenant_id = %s AND f.estado != 'Anulada' {filtro_factura}
        GROUP BY f.id
        HAVING (f.total - COALESCE(SUM(pf.monto_aplicado), 0)) > 0
        ORDER BY f.fecha_emision DESC, f.numero_factura DESC
        LIMIT 100
    ''', tuple(params_facturas), fetch='all') or []
    
    ars_pago = None
    if facturas_list:
        ars_pago = facturas_list[0].get('nombre_ars')
    
    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    return render_template('facturacion/pago_form.html', facturas_list=facturas_list, fecha_actual=fecha_actual, factura_id_param=factura_id_param, ars_pago=ars_pago)

@app.route('/facturacion/pacientes/exportar-excel')
@login_required
def facturacion_pacientes_exportar_excel():
    """Exportar lista de pacientes a Excel"""
    if not OPENPYXL_AVAILABLE:
        flash('La funcionalidad de Excel no está disponible', 'error')
        return redirect(url_for('facturacion_pacientes'))
    
    tenant_id = get_current_tenant_id()
    search = request.args.get('search', '').strip()
    
    # Obtener pacientes con el mismo filtro que la vista
    query = '''
        SELECT p.*, a.nombre as ars_nombre 
        FROM pacientes p 
        LEFT JOIN ars a ON p.ars_id = a.id 
        WHERE p.tenant_id = %s
    '''
    params = [tenant_id]
    
    if search:
        query += ' AND (p.nombre LIKE %s OR p.nss LIKE %s OR p.cedula LIKE %s)'
        search_pattern = f'%{search}%'
        params.extend([search_pattern, search_pattern, search_pattern])
    
    query += ' ORDER BY p.nombre'
    
    pacientes_list = execute_query(query, tuple(params), fetch='all') or []
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pacientes"
    
    # Estilos para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = ['NSS', 'Nombre Completo', 'Cédula', 'Teléfono', 'Email', 'Fecha de Nacimiento', 
               'Sexo', 'ARS', 'Tipo de Afiliación', 'Dirección']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Datos
    for row_num, paciente in enumerate(pacientes_list, 2):
        ws.cell(row=row_num, column=1, value=paciente.get('nss') or '')
        ws.cell(row=row_num, column=2, value=paciente.get('nombre') or '')
        ws.cell(row=row_num, column=3, value=paciente.get('cedula') or '')
        ws.cell(row=row_num, column=4, value=paciente.get('telefono') or '')
        ws.cell(row=row_num, column=5, value=paciente.get('email') or '')
        ws.cell(row=row_num, column=6, value=paciente.get('fecha_nacimiento') or '')
        ws.cell(row=row_num, column=7, value=paciente.get('sexo') or '')
        ws.cell(row=row_num, column=8, value=paciente.get('ars_nombre') or 'Sin ARS')
        ws.cell(row=row_num, column=9, value=paciente.get('tipo_afiliacion') or '')
        ws.cell(row=row_num, column=10, value=paciente.get('direccion') or '')
    
    # Ajustar ancho de columnas
    column_widths = [15, 30, 15, 15, 25, 15, 10, 20, 15, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre del archivo con fecha
    fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'pacientes_{fecha_actual}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/facturacion/historico')
@login_required
def facturacion_historico():
    """Histórico de facturas - Filtrado por tenant"""
    tenant_id = get_current_tenant_id()
    
    # Verificar si existe tenant_id en facturas
    tiene_tenant_id_facturas = False
    try:
        check_tenant_facturas = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'facturas' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id_facturas = check_tenant_facturas and check_tenant_facturas.get('count', 0) > 0
    except:
        tiene_tenant_id_facturas = False
    
    if tiene_tenant_id_facturas:
        # Paginación para mejor rendimiento
        page = validate_int(request.args.get('page', 1), min_value=1, default=1)
        per_page = validate_int(request.args.get('per_page', 50), min_value=1, max_value=100, default=50)
        offset = (page - 1) * per_page
        
        facturas = execute_query('''
            SELECT f.*, f.fecha_emision as fecha_factura, a.nombre as nombre_ars, m.nombre as medico_nombre
            FROM facturas f 
            LEFT JOIN ars a ON f.ars_id = a.id 
            LEFT JOIN medicos m ON f.medico_id = m.id
            WHERE f.tenant_id = %s
            ORDER BY f.id DESC 
            LIMIT %s OFFSET %s
        ''', (tenant_id, per_page, offset), fetch='all') or []
        
        # Obtener total para paginación
        total_facturas = execute_query('''
            SELECT COUNT(*) as total FROM facturas WHERE tenant_id = %s
        ''', (tenant_id,))
        total_pages = (total_facturas.get('total', 0) + per_page - 1) // per_page if total_facturas else 0
    else:
        facturas = execute_query('''
            SELECT f.*, f.fecha_emision as fecha_factura, a.nombre as nombre_ars, m.nombre as medico_nombre
            FROM facturas f 
            LEFT JOIN ars a ON f.ars_id = a.id 
            LEFT JOIN medicos m ON f.medico_id = m.id
            ORDER BY f.id DESC 
            LIMIT 100
        ''', fetch='all') or []
    
    return render_template('facturacion/historico.html', 
                         facturas=facturas,
                         page=page,
                         per_page=per_page,
                         total_pages=total_pages,
                         total_facturas=total_facturas.get('total', 0) if total_facturas else 0)

@app.route('/facturacion/facturas/<int:factura_id>/ver')
@login_required
def facturacion_ver_factura(factura_id):
    """Ver factura generada"""
    # Validar que el ID sea válido
    if not validate_int(factura_id, min_value=1):
        flash('ID de factura inválido', 'error')
        return redirect(url_for('facturacion_historico'))
    
    # Validar acceso al tenant
    if not validate_tenant_access('facturas', factura_id):
        flash('No tienes acceso a esta factura', 'error')
        return redirect(url_for('facturacion_historico'))
    
    tenant_id = get_current_tenant_id()
    
    # Verificar si existe tenant_id en facturas
    tiene_tenant_id_facturas = False
    try:
        check_tenant_facturas = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'facturas' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id_facturas = check_tenant_facturas and check_tenant_facturas.get('count', 0) > 0
    except:
        tiene_tenant_id_facturas = False
    
    # Obtener factura
    if tiene_tenant_id_facturas:
        factura = execute_query('''
            SELECT f.*, a.nombre as nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur
            FROM facturas f
            LEFT JOIN ars a ON f.ars_id = a.id
            LEFT JOIN medicos m ON f.medico_id = m.id
            WHERE f.id = %s AND f.tenant_id = %s
        ''', (factura_id, tenant_id))
    else:
        factura = execute_query('''
            SELECT f.*, a.nombre as nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur
            FROM facturas f
            LEFT JOIN ars a ON f.ars_id = a.id
            LEFT JOIN medicos m ON f.medico_id = m.id
            WHERE f.id = %s
        ''', (factura_id,))
    
    if not factura:
        flash('Factura no encontrada', 'error')
        return redirect(url_for('facturacion_historico'))
    
    # Obtener detalles de la factura (pacientes/servicios)
    detalles = execute_query('''
        SELECT * FROM factura_detalles 
        WHERE factura_id = %s
        ORDER BY id
    ''', (factura_id,), fetch='all') or []
    
    # Procesar detalles para mostrar como pacientes
    pacientes = []
    for detalle in detalles:
        # Extraer información del detalle
        descripcion = detalle.get('descripcion', '')
        # Intentar extraer autorización de la descripción si está presente
        autorizacion = ''
        descripcion_servicio = descripcion
        if ' - Autorización:' in descripcion:
            partes = descripcion.split(' - Autorización:')
            descripcion_servicio = partes[0].strip()
            autorizacion = partes[1].strip() if len(partes) > 1 else ''
        
        paciente = {
            'nombre_paciente': factura.get('nombre_paciente', 'N/A'),
            'nss': factura.get('nss_paciente', ''),
            'fecha_servicio': factura.get('fecha_emision', ''),
            'autorizacion': autorizacion,
            'descripcion_servicio': descripcion_servicio if descripcion_servicio else '',
            'monto_estimado': float(detalle.get('precio_unitario', 0) or 0),
            'monto': float(detalle.get('precio_unitario', 0) or 0)
        }
        pacientes.append(paciente)
    
    # Obtener centro médico
    centro_medico = None
    if factura.get('centro_medico_id'):
        centro_medico = execute_query('SELECT * FROM centros_medicos WHERE id = %s AND tenant_id = %s', 
                                     (factura['centro_medico_id'], tenant_id))
    
    if not centro_medico:
        centro_medico = execute_query('SELECT * FROM centros_medicos WHERE tenant_id = %s LIMIT 1', (tenant_id,))
    
    if not centro_medico:
        centro_medico = {
            'nombre': 'Centro Médico',
            'direccion': ''
        }
    
    # Calcular totales
    subtotal = float(factura.get('subtotal', 0) or 0)
    total = float(factura.get('total', 0) or 0)
    
    # Obtener información de la empresa para determinar tipo
    empresa_info = get_empresa_info(tenant_id)
    tipo_empresa = empresa_info.get('tipo_empresa') if empresa_info else None
    
    # Obtener médico o empresa según tipo (igual que en vista_previa)
    medico_factura = None
    if tipo_empresa == 'centro_salud':
        medico_factura = {
            'id': empresa_info.get('id'),
            'nombre': empresa_info.get('razon_social', empresa_info.get('nombre', 'N/A')),
            'especialidad': 'Centro de Salud',
            'cedula': empresa_info.get('rnc', ''),
            'telefono': empresa_info.get('telefono', ''),
            'email': empresa_info.get('email', '')
        }
    else:
        # Obtener datos completos del médico
        medico_completo = execute_query('SELECT * FROM medicos WHERE id = %s', (factura.get('medico_id'),))
        if medico_completo:
            medico_factura = {
                'id': medico_completo.get('id'),
                'nombre': medico_completo.get('nombre', factura.get('medico_nombre', 'N/A')),
                'especialidad': medico_completo.get('especialidad', factura.get('medico_especialidad', '')),
                'cedula': medico_completo.get('cedula', factura.get('medico_cedula', '')),
                'exequatur': medico_completo.get('exequatur', factura.get('medico_exequatur', '')),
                'telefono': medico_completo.get('telefono', ''),
                'email': medico_completo.get('email', '')
            }
        else:
            medico_factura = {
                'id': factura.get('medico_id'),
                'nombre': factura.get('medico_nombre', 'N/A'),
                'especialidad': factura.get('medico_especialidad', ''),
                'cedula': factura.get('medico_cedula', ''),
                'exequatur': factura.get('medico_exequatur', ''),
                'telefono': '',
                'email': ''
            }
    
    # Obtener NCF completo y descripción
    ncf_numero = factura.get('ncf', '')
    ncf_prefijo = ncf_numero[:3] if len(ncf_numero) >= 3 else ''
    ncf_obj = execute_query('SELECT * FROM ncf WHERE prefijo = %s AND tenant_id = %s LIMIT 1', (ncf_prefijo, tenant_id))
    if not ncf_obj:
        ncf_obj = execute_query('SELECT * FROM ncf WHERE prefijo = %s LIMIT 1', (ncf_prefijo,))
    
    ncf_tipos_descripciones = {
        'B01': 'Factura de Crédito Fiscal',
        'B02': 'Factura de Consumo',
        'B14': 'Registro Único de Ingresos',
        'B15': 'GUBERNAMENTAL'
    }
    ncf_tipo_descripcion = ncf_tipos_descripciones.get(ncf_obj.get('tipo', '') if ncf_obj else '', '') if ncf_obj else ''
    ncf_fecha_fin = ncf_obj.get('fecha_fin', '') if ncf_obj else ''
    ncf_completo = ncf_numero
    
    # Intentar obtener pacientes desde pacientes_pendientes que fueron facturados
    fecha_factura = factura.get('fecha_emision', '')
    pacientes_pendientes_facturados = []
    try:
        tiene_tenant_id_pp = False
        try:
            check_tenant_pp = execute_query('''
                SELECT COUNT(*) as count 
                FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE()
                AND TABLE_NAME = 'pacientes_pendientes' 
                AND COLUMN_NAME = 'tenant_id'
            ''')
            tiene_tenant_id_pp = check_tenant_pp and check_tenant_pp.get('count', 0) > 0
        except:
            tiene_tenant_id_pp = False
        
        if tiene_tenant_id_pp:
            pacientes_pendientes_facturados = execute_query('''
                SELECT pp.*, a.nombre as ars_nombre
                FROM pacientes_pendientes pp
                LEFT JOIN ars a ON pp.ars_id = a.id
                WHERE pp.estado = 'Facturado' 
                AND pp.ars_id = %s 
                AND DATE(pp.updated_at) = DATE(%s)
                AND pp.tenant_id = %s
                ORDER BY pp.id
            ''', (factura.get('ars_id'), fecha_factura, tenant_id), fetch='all') or []
        else:
            pacientes_pendientes_facturados = execute_query('''
                SELECT pp.*, a.nombre as ars_nombre
                FROM pacientes_pendientes pp
                LEFT JOIN ars a ON pp.ars_id = a.id
                WHERE pp.estado = 'Facturado' 
                AND pp.ars_id = %s 
                AND DATE(pp.updated_at) = DATE(%s)
                ORDER BY pp.id
            ''', (factura.get('ars_id'), fecha_factura), fetch='all') or []
    except Exception as e:
        logger.error(f"Error al obtener pacientes_pendientes_facturados: {str(e)}")
        pacientes_pendientes_facturados = []
    
    # Procesar pacientes desde detalles y pacientes_pendientes (igual que en vista_previa)
    pacientes_procesados = []
    if pacientes_pendientes_facturados and len(pacientes_pendientes_facturados) == len(detalles):
        # Si encontramos pacientes_pendientes que coinciden, usarlos
        for idx, (detalle, pp) in enumerate(zip(detalles, pacientes_pendientes_facturados), 1):
            servicio_completo = pp.get('servicios_realizados', '') or ''
            if ' - Autorización:' in servicio_completo:
                partes = servicio_completo.split(' - Autorización:')
                descripcion_servicio = partes[0].strip()
                autorizacion = partes[1].strip() if len(partes) > 1 else ''
            else:
                descripcion_servicio = servicio_completo.strip()
                autorizacion = ''
            
            paciente = {
                'nombre_paciente': pp.get('nombre_paciente', factura.get('nombre_paciente', 'N/A')),
                'nss': pp.get('nss', factura.get('nss_paciente', '')),
                'fecha_servicio': pp.get('fecha_servicio', fecha_factura),
                'autorizacion': autorizacion,
                'descripcion_servicio': descripcion_servicio if descripcion_servicio else detalle.get('descripcion', ''),
                'monto_estimado': float(detalle.get('precio_unitario', 0) or 0),
                'monto': float(detalle.get('precio_unitario', 0) or 0)
            }
            pacientes_procesados.append(paciente)
    else:
        # Si no encontramos pacientes_pendientes, usar datos de la factura
        for idx, detalle in enumerate(detalles, 1):
            descripcion = detalle.get('descripcion', '')
            # Intentar extraer autorización de la descripción si está presente
            autorizacion = ''
            descripcion_servicio = descripcion
            if ' - Autorización:' in descripcion:
                partes = descripcion.split(' - Autorización:')
                descripcion_servicio = partes[0].strip()
                autorizacion = partes[1].strip() if len(partes) > 1 else ''
            
            paciente = {
                'nombre_paciente': factura.get('nombre_paciente', 'N/A'),
                'nss': factura.get('nss_paciente', ''),
                'fecha_servicio': fecha_factura,
                'autorizacion': autorizacion,
                'descripcion_servicio': descripcion_servicio if descripcion_servicio else '',
                'monto_estimado': float(detalle.get('precio_unitario', 0) or 0),
                'monto': float(detalle.get('precio_unitario', 0) or 0)
            }
            pacientes_procesados.append(paciente)
    
    # Preparar datos de ARS
    ars = {
        'id': factura.get('ars_id'),
        'nombre': factura.get('nombre_ars', 'N/A'),
        'rnc': factura.get('ars_rnc', '')
    }
    
    # Preparar datos de NCF
    ncf = {
        'id': ncf_obj.get('id') if ncf_obj else None,
        'prefijo': ncf_prefijo,
        'tipo': ncf_obj.get('tipo', '') if ncf_obj else '',
        'fecha_fin': ncf_fecha_fin
    }
    
    return render_template('facturacion/ver_factura.html',
                          factura=factura,
                          pacientes=pacientes_procesados,
                          centro_medico=centro_medico,
                          subtotal=subtotal,
                          total=total,
                          ars=ars,
                          ncf=ncf,
                          ncf_completo=ncf_completo,
                          ncf_tipo_descripcion=ncf_tipo_descripcion,
                          medico=medico_factura,
                          tipo_empresa=tipo_empresa,
                          empresa_info=empresa_info,
                          fecha_factura=fecha_factura)

@app.route('/facturacion/facturas/<int:factura_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_editar_factura(factura_id):
    """Editar factura generada"""
    tenant_id = get_current_tenant_id()
    
    # Verificar si existe tenant_id en facturas
    tiene_tenant_id_facturas = False
    try:
        check_tenant_facturas = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'facturas' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id_facturas = check_tenant_facturas and check_tenant_facturas.get('count', 0) > 0
    except:
        tiene_tenant_id_facturas = False
    
    # Obtener factura
    if tiene_tenant_id_facturas:
        factura = execute_query('''
            SELECT f.*, a.nombre as nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur
            FROM facturas f
            LEFT JOIN ars a ON f.ars_id = a.id
            LEFT JOIN medicos m ON f.medico_id = m.id
            WHERE f.id = %s AND f.tenant_id = %s
        ''', (factura_id, tenant_id))
    else:
        factura = execute_query('''
            SELECT f.*, a.nombre as nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur
            FROM facturas f
            LEFT JOIN ars a ON f.ars_id = a.id
            LEFT JOIN medicos m ON f.medico_id = m.id
            WHERE f.id = %s
        ''', (factura_id,))
    
    if not factura:
        flash('Factura no encontrada', 'error')
        return redirect(url_for('facturacion_historico'))
    
    # Calcular días transcurridos desde la creación
    from datetime import datetime, date
    fecha_creacion = factura.get('created_at')
    if isinstance(fecha_creacion, str):
        try:
            fecha_creacion = datetime.strptime(fecha_creacion, '%Y-%m-%d %H:%M:%S').date()
        except:
            fecha_creacion = date.today()
    elif isinstance(fecha_creacion, datetime):
        fecha_creacion = fecha_creacion.date()
    elif isinstance(fecha_creacion, date):
        pass
    else:
        fecha_creacion = date.today()
    
    fecha_actual = date.today()
    dias_transcurridos = (fecha_actual - fecha_creacion).days
    dias_restantes = 30 - dias_transcurridos
    
    # Verificar si se puede editar (menos de 30 días)
    if dias_transcurridos >= 30:
        flash('Esta factura no se puede editar. Han pasado más de 30 días desde su creación.', 'error')
        return redirect(url_for('facturacion_historico'))
    
    # Si es POST, procesar la actualización
    if request.method == 'POST':
        # Aquí se procesaría la actualización de la factura
        # Por ahora, solo redirigir
        flash('Funcionalidad de edición en desarrollo', 'info')
        return redirect(url_for('facturacion_historico'))
    
    # Obtener detalles de la factura (pacientes/servicios)
    detalles = execute_query('''
        SELECT * FROM factura_detalles 
        WHERE factura_id = %s
        ORDER BY id
    ''', (factura_id,), fetch='all') or []
    
    # Procesar detalles para mostrar como pacientes
    pacientes = []
    for detalle in detalles:
        # Extraer información del servicio desde la descripción
        descripcion_servicio = detalle.get('descripcion', '')
        servicio_nombre = descripcion_servicio
        if ' - Autorización:' in descripcion_servicio:
            servicio_nombre = descripcion_servicio.split(' - Autorización:')[0].strip()
        
        paciente = {
            'id': detalle.get('id'),
            'nombre_paciente': factura.get('nombre_paciente', 'N/A'),
            'nss': factura.get('nss_paciente', ''),
            'fecha_servicio': factura.get('fecha_emision', ''),
            'autorizacion': '',
            'descripcion_servicio': descripcion_servicio,
            'servicio_nombre': servicio_nombre,
            'medico_nombre': factura.get('medico_nombre', 'N/A'),
            'monto_estimado': float(detalle.get('precio_unitario', 0) or 0),
            'monto': float(detalle.get('precio_unitario', 0) or 0)
        }
        pacientes.append(paciente)
    
    # Agregar campos adicionales a factura para el template
    factura['fecha_factura'] = factura.get('fecha_emision', '')
    factura['ncf_numero'] = factura.get('ncf', '')
    
    return render_template('facturacion/editar_factura.html',
                          factura=factura,
                          pacientes_factura=pacientes,  # Cambiado de pacientes a pacientes_factura
                          pacientes_disponibles=[],  # Lista vacía por ahora, se puede poblar después
                          dias_transcurridos=dias_transcurridos,
                          dias_restantes=dias_restantes)

def generar_pdf_factura_vista_previa(factura_id, tenant_id=None):
    """Generar PDF de factura con el mismo formato que la vista previa"""
    if not REPORTLAB_AVAILABLE:
        return None
    
    if tenant_id is None:
        tenant_id = get_current_tenant_id()
    
    # Verificar si existe tenant_id en facturas
    tiene_tenant_id_facturas = False
    try:
        check_tenant_facturas = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'facturas' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id_facturas = check_tenant_facturas and check_tenant_facturas.get('count', 0) > 0
    except:
        tiene_tenant_id_facturas = False
    
    # Obtener factura
    if tiene_tenant_id_facturas:
        factura = execute_query('''
            SELECT f.*, a.nombre as nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur, m.id as medico_id
            FROM facturas f
            LEFT JOIN ars a ON f.ars_id = a.id
            LEFT JOIN medicos m ON f.medico_id = m.id
            WHERE f.id = %s AND f.tenant_id = %s
        ''', (factura_id, tenant_id))
    else:
        factura = execute_query('''
            SELECT f.*, a.nombre as nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur, m.id as medico_id
            FROM facturas f
            LEFT JOIN ars a ON f.ars_id = a.id
            LEFT JOIN medicos m ON f.medico_id = m.id
            WHERE f.id = %s
        ''', (factura_id,))
    
    if not factura:
        logger.error(f"Factura {factura_id} no encontrada para tenant {tenant_id}")
        print(f"ERROR: Factura {factura_id} no encontrada para tenant {tenant_id}")
        return None
    
    logger.info(f"Iniciando generación de PDF para factura {factura_id}")
    print(f"INFO: Factura {factura_id} encontrada. Datos básicos: ncf={factura.get('ncf')}, fecha_emision={factura.get('fecha_emision')}, ars_id={factura.get('ars_id')}")
    
    # Obtener información de la empresa para determinar tipo
    empresa_info = get_empresa_info(tenant_id)
    tipo_empresa = empresa_info.get('tipo_empresa') if empresa_info else None
    
    # Obtener médico o empresa según tipo
    medico_factura = None
    if tipo_empresa == 'centro_salud':
        medico_factura = {
            'id': empresa_info.get('id'),
            'nombre': empresa_info.get('razon_social', empresa_info.get('nombre', 'N/A')),
            'especialidad': 'Centro de Salud',
            'cedula': empresa_info.get('rnc', '')
        }
    else:
        medico_factura = {
            'id': factura.get('medico_id'),
            'nombre': factura.get('medico_nombre', 'N/A'),
            'especialidad': factura.get('medico_especialidad', ''),
            'cedula': factura.get('medico_cedula', '')
        }
    
    # Obtener ARS
    ars = {
        'id': factura.get('ars_id'),
        'nombre': factura.get('nombre_ars', 'N/A'),
        'rnc': factura.get('ars_rnc', '')
    }
    
    # Obtener NCF
    ncf_numero = factura.get('ncf', '')
    ncf_prefijo = ncf_numero[:3] if len(ncf_numero) >= 3 else ''
    ncf_obj = execute_query('SELECT * FROM ncf WHERE prefijo = %s AND tenant_id = %s LIMIT 1', (ncf_prefijo, tenant_id))
    if not ncf_obj:
        ncf_obj = execute_query('SELECT * FROM ncf WHERE prefijo = %s LIMIT 1', (ncf_prefijo,))
    
    ncf_tipos_descripciones = {
        'B01': 'Factura de Crédito Fiscal',
        'B02': 'Factura de Consumo',
        'B14': 'Registro Único de Ingresos',
        'B15': 'GUBERNAMENTAL'
    }
    ncf_tipo_descripcion = ncf_tipos_descripciones.get(ncf_obj.get('tipo', '') if ncf_obj else '', '') if ncf_obj else ''
    ncf_fecha_fin = ncf_obj.get('fecha_fin', '') if ncf_obj else ''
    
    # Obtener fecha de factura para usar en consultas
    fecha_factura = factura.get('fecha_emision', '')
    
    # Obtener detalles de la factura (pacientes/servicios)
    detalles = execute_query('''
        SELECT * FROM factura_detalles 
        WHERE factura_id = %s
        ORDER BY id
    ''', (factura_id,), fetch='all') or []
    
    # Intentar obtener pacientes desde pacientes_pendientes que fueron facturados
    # Buscar pacientes_pendientes con estado 'Facturado' que coincidan con esta factura
    # Por fecha y ARS como aproximación
    pacientes_pendientes_facturados = []
    try:
        # Verificar si existe tenant_id en pacientes_pendientes
        tiene_tenant_id_pp = False
        try:
            check_tenant_pp = execute_query('''
                SELECT COUNT(*) as count 
                FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE()
                AND TABLE_NAME = 'pacientes_pendientes' 
                AND COLUMN_NAME = 'tenant_id'
            ''')
            tiene_tenant_id_pp = check_tenant_pp and check_tenant_pp.get('count', 0) > 0
        except:
            tiene_tenant_id_pp = False
        
        # Buscar pacientes pendientes facturados en la misma fecha y ARS
        if tiene_tenant_id_pp:
            pacientes_pendientes_facturados = execute_query('''
                SELECT pp.*, a.nombre as ars_nombre
                FROM pacientes_pendientes pp
                LEFT JOIN ars a ON pp.ars_id = a.id
                WHERE pp.estado = 'Facturado' 
                AND pp.ars_id = %s 
                AND DATE(pp.updated_at) = DATE(%s)
                AND pp.tenant_id = %s
                ORDER BY pp.id
            ''', (factura.get('ars_id'), fecha_factura, tenant_id), fetch='all') or []
        else:
            pacientes_pendientes_facturados = execute_query('''
                SELECT pp.*, a.nombre as ars_nombre
                FROM pacientes_pendientes pp
                LEFT JOIN ars a ON pp.ars_id = a.id
                WHERE pp.estado = 'Facturado' 
                AND pp.ars_id = %s 
                AND DATE(pp.updated_at) = DATE(%s)
                ORDER BY pp.id
            ''', (factura.get('ars_id'), fecha_factura), fetch='all') or []
    except Exception as e:
        logger.error(f"Error al obtener pacientes_pendientes_facturados: {str(e)}")
        pacientes_pendientes_facturados = []
    
    # Validar que hay detalles antes de procesar
    if not detalles:
        logger.error(f"Factura {factura_id} no tiene detalles asociados. Query ejecutada: SELECT * FROM factura_detalles WHERE factura_id = {factura_id}")
        return None
    
    logger.info(f"Factura {factura_id} tiene {len(detalles)} detalles. Primer detalle: {detalles[0] if detalles else 'N/A'}")
    
    # Procesar pacientes desde detalles y pacientes_pendientes
    pacientes = []
    if pacientes_pendientes_facturados and len(pacientes_pendientes_facturados) == len(detalles):
        # Si encontramos pacientes_pendientes que coinciden, usarlos
        for idx, (detalle, pp) in enumerate(zip(detalles, pacientes_pendientes_facturados), 1):
            servicio_completo = pp.get('servicios_realizados', '') or ''
            if ' - Autorización:' in servicio_completo:
                partes = servicio_completo.split(' - Autorización:')
                descripcion_servicio = partes[0].strip()
                autorizacion = partes[1].strip() if len(partes) > 1 else ''
            else:
                descripcion_servicio = servicio_completo.strip()
                autorizacion = ''
            
            paciente = {
                'nombre_paciente': pp.get('nombre_paciente', factura.get('nombre_paciente', 'N/A')),
                'nss': pp.get('nss', factura.get('nss_paciente', '')),
                'fecha_servicio': pp.get('fecha_servicio', factura.get('fecha_emision', '')),
                'autorizacion': autorizacion,
                'descripcion_servicio': descripcion_servicio if descripcion_servicio else detalle.get('descripcion', ''),
                'monto_estimado': float(detalle.get('precio_unitario', 0) or 0),
                'monto': float(detalle.get('precio_unitario', 0) or 0)
            }
            pacientes.append(paciente)
    else:
        # Si no encontramos pacientes_pendientes, usar datos de la factura (como en ver_factura)
        for idx, detalle in enumerate(detalles, 1):
            descripcion = detalle.get('descripcion', '')
            # Intentar extraer autorización de la descripción si está presente
            autorizacion = ''
            descripcion_servicio = descripcion
            if ' - Autorización:' in descripcion:
                partes = descripcion.split(' - Autorización:')
                descripcion_servicio = partes[0].strip()
                autorizacion = partes[1].strip() if len(partes) > 1 else ''
            
            paciente = {
                'nombre_paciente': factura.get('nombre_paciente', 'N/A'),
                'nss': factura.get('nss_paciente', ''),
                'fecha_servicio': fecha_factura,
                'autorizacion': autorizacion,
                'descripcion_servicio': descripcion_servicio if descripcion_servicio else '',
                'monto_estimado': float(detalle.get('precio_unitario', 0) or 0),
                'monto': float(detalle.get('precio_unitario', 0) or 0)
            }
            pacientes.append(paciente)
    
    # Obtener centro médico
    centro_medico = None
    if factura.get('centro_medico_id'):
        centro_medico = execute_query('SELECT * FROM centros_medicos WHERE id = %s AND tenant_id = %s', 
                                     (factura['centro_medico_id'], tenant_id))
    
    if not centro_medico:
        centro_medico = execute_query('SELECT * FROM centros_medicos WHERE tenant_id = %s LIMIT 1', (tenant_id,))
    
    if not centro_medico:
        centro_medico = {
            'nombre': 'Centro Médico',
            'direccion': ''
        }
    
    # Calcular totales
    subtotal = float(factura.get('subtotal', 0) or 0)
    total = float(factura.get('total', 0) or 0)
    
    # Obtener datos completos del médico para el footer y remitente (antes de generar PDF)
    medico_completo = None
    if tipo_empresa != 'centro_salud' and medico_factura.get('id'):
        medico_completo = execute_query('SELECT * FROM medicos WHERE id = %s', (medico_factura.get('id'),))
        # Actualizar medico_factura con datos completos si están disponibles
        if medico_completo:
            medico_factura['nombre'] = medico_completo.get('nombre', medico_factura.get('nombre', 'N/A'))
            medico_factura['especialidad'] = medico_completo.get('especialidad', medico_factura.get('especialidad', ''))
            medico_factura['cedula'] = medico_completo.get('cedula', medico_factura.get('cedula', ''))
            medico_factura['exequatur'] = medico_completo.get('exequatur', '')
    
    # Validar datos mínimos necesarios antes de generar PDF
    if not pacientes:
        logger.error(f"Factura {factura_id} no tiene pacientes procesados después de procesar {len(detalles)} detalles")
        logger.error(f"Detalles procesados: {detalles}")
        logger.error(f"Pacientes pendientes encontrados: {len(pacientes_pendientes_facturados)}")
        return None
    
    # Validar que tenemos datos esenciales
    if not fecha_factura:
        logger.error(f"Factura {factura_id} no tiene fecha_emision. factura['fecha_emision'] = {factura.get('fecha_emision')}")
        return None
    
    if not ars.get('nombre'):
        logger.error(f"Factura {factura_id} no tiene nombre de ARS. ars = {ars}, factura['nombre_ars'] = {factura.get('nombre_ars')}")
        return None
    
    if not ncf_numero:
        logger.error(f"Factura {factura_id} no tiene número de NCF. factura['ncf'] = {factura.get('ncf')}")
        return None
    
    logger.info(f"Iniciando generación de PDF para factura {factura_id}: {len(pacientes)} pacientes, subtotal: {subtotal}, total: {total}")
    logger.info(f"Datos validados: fecha_factura={fecha_factura}, ars={ars.get('nombre')}, ncf={ncf_numero}, tipo_empresa={tipo_empresa}")
    
    # Generar PDF
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, 
                                leftMargin=0.5*inch, rightMargin=0.5*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        primary_color = colors.black
        subtle_gray = colors.HexColor('#d9d9d9')
        
        # Datos auxiliares
        numero_factura = factura.get('numero_factura', factura_id)
        proveedor_rnc = ''
        if tipo_empresa == 'centro_salud':
            proveedor_rnc = empresa_info.get('rnc', '') if empresa_info else ''
        else:
            proveedor_rnc = medico_factura.get('cedula', '')
        cliente_rnc = ars.get('rnc') or factura.get('rnc_paciente') or ''
        
        # Título principal con número de factura y NCF
        title_style = ParagraphStyle('TituloFactura', parent=styles['Normal'], fontSize=16, fontName='Helvetica-Bold')
        ncf_style = ParagraphStyle('NCF', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold')
        title_table = Table([
            [Paragraph(f"FACTURA #{escape(str(numero_factura))}", title_style),
             Paragraph(escape(str(ncf_numero)), ncf_style)]
        ], colWidths=[4.0*inch, 3.5*inch])
        title_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(title_table)
        story.append(Spacer(1, 0.15*inch))
        
        # Datos de proveedor / cliente
        info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=8.5, leading=11)
        info_table = Table([
            [Paragraph(f"De: {escape(str(medico_factura.get('nombre', '')))}", info_style),
             Paragraph(f"Fecha: {escape(str(fecha_factura))}", info_style)],
            [Paragraph(f"RNC: {escape(str(proveedor_rnc or ''))}", info_style),
             Paragraph(f"Cliente: {escape(str(ars.get('nombre', '')))}", info_style)],
            ['', Paragraph(f"RNC: {escape(str(cliente_rnc or ''))}", info_style)],
        ], colWidths=[3.75*inch, 3.75*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        story.append(info_table)
        
        # Línea divisoria
        line_div = Table([['']], colWidths=[7.5*inch])
        line_div.setStyle(TableStyle([('LINEBELOW', (0, 0), (-1, -1), 1.0, primary_color)]))
        story.append(line_div)
        story.append(Spacer(1, 0.15*inch))
        
        # Tabla de pacientes/servicios
        if pacientes:
            header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold')
            cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10)
            tabla_headers = ['No.', 'NOMBRES PACIENTE', 'NSS', 'FECHA', 'AUTORIZACIÓN', 'SERVICIO', 'V/UNITARIO']
            tabla_data = [tabla_headers]
            for idx, paciente in enumerate(pacientes, 1):
                monto = float(paciente.get('monto') or paciente.get('monto_estimado', 0) or 0)
                tabla_data.append([
                    str(idx),
                    Paragraph(escape(str(paciente.get('nombre_paciente', ''))), cell_style),
                    escape(str(paciente.get('nss', ''))),
                    escape(str(paciente.get('fecha_servicio', ''))),
                    escape(str(paciente.get('autorizacion', ''))),
                    escape(str(paciente.get('descripcion_servicio', ''))),
                    Paragraph(f"{monto:,.2f}", ParagraphStyle('MontoCell', parent=cell_style, alignment=TA_RIGHT))
                ])
            
            tabla = Table(tabla_data, colWidths=[0.35*inch, 2.2*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1.4*inch, 0.85*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.black),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),
                ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cfcfcf')),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(tabla)
        
        story.append(Spacer(1, 0.3*inch))
        
        # Totales
        total_label_style = ParagraphStyle('TotalLabel', parent=styles['Normal'], fontSize=9, alignment=TA_RIGHT, fontName='Helvetica-Bold')
        total_value_style = ParagraphStyle('TotalValue', parent=styles['Normal'], fontSize=9, alignment=TA_RIGHT)
        total_final_style = ParagraphStyle('TotalFinal', parent=styles['Normal'], fontSize=12, alignment=TA_RIGHT, fontName='Helvetica-Bold')
        
        totales_data = [
            [Paragraph('Subtotal:', total_label_style), Paragraph(f"{subtotal:,.2f}", total_value_style)],
            [Paragraph('ITBIS:', total_label_style), Paragraph('*E', total_value_style)],
            [Paragraph('TOTAL:', total_label_style), Paragraph(f"{total:,.2f}", total_final_style)],
        ]
        totales_table = Table(totales_data, colWidths=[1.5*inch, 1.2*inch])
        totales_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('LINEABOVE', (0, 2), (-1, 2), 1, primary_color),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        totales_container = Table([[totales_table]], colWidths=[7.5*inch])
        totales_container.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ]))
        story.append(totales_container)
        
        story.append(Spacer(1, 0.35*inch))
        
        # Footer sencillo (ajustado)
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)
        footer_bold_style = ParagraphStyle('FooterBold', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, fontName='Helvetica-Bold')
        footer_data = [
            [Paragraph(escape(str(medico_factura.get('nombre', ''))), footer_bold_style)],
        ]
        footer_line2 = []
        if centro_medico and centro_medico.get('nombre'):
            footer_line2.append(centro_medico.get('nombre'))
        if centro_medico and centro_medico.get('direccion'):
            footer_line2.append(centro_medico.get('direccion'))
        if footer_line2:
            footer_data.append([Paragraph(' - '.join(footer_line2), footer_style)])
        footer_line3 = []
        if empresa_info and empresa_info.get('telefono'):
            footer_line3.append(f"Tel: {empresa_info.get('telefono')}")
        if empresa_info and empresa_info.get('email'):
            footer_line3.append(f"Email: {empresa_info.get('email')}")
        if footer_line3:
            footer_data.append([Paragraph(' | '.join(footer_line3), footer_style)])
        footer_table = Table(footer_data, colWidths=[7.5*inch])
        footer_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        story.append(footer_table)
        
        # Construir PDF
        try:
            logger.info(f"Construyendo PDF para factura {factura_id} con {len(story)} elementos en story")
            doc.build(story)
            logger.info(f"PDF construido exitosamente para factura {factura_id}")
            
            buffer.seek(0)
            buffer_content = buffer.read()
            buffer_size = len(buffer_content) if buffer_content else 0
            if buffer_size == 0:
                logger.error(f"PDF generado para factura {factura_id} está vacío (buffer_size=0)")
                return None
            
            pdf_buffer = BytesIO(buffer_content)
            pdf_buffer.seek(0)
            logger.info(f"PDF generado exitosamente para factura {factura_id}, tamaño: {buffer_size} bytes")
            return pdf_buffer
        except Exception as build_error:
            import traceback
            error_trace = traceback.format_exc()
            logger.error(f"Error al construir PDF para factura {factura_id}: {error_trace}")
            print(f"Error al construir PDF para factura {factura_id}: {str(build_error)}")
            print(f"Traceback: {error_trace}")
            return None
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Error al generar PDF para factura {factura_id}: {error_trace}")
        logger.error(f"Tipo de error: {type(e).__name__}, Mensaje: {str(e)}")
        print(f"ERROR CRÍTICO al generar PDF para factura {factura_id}: {str(e)}")
        print(f"Traceback completo: {error_trace}")
        print(f"DEBUG - factura encontrada: {factura is not None}")
        print(f"DEBUG - detalles encontrados: {len(detalles) if detalles else 0}")
        print(f"DEBUG - pacientes procesados: {len(pacientes) if 'pacientes' in locals() else 'N/A'}")
        return None

@app.route('/facturacion/facturas/<int:factura_id>/pdf')
@login_required
def facturacion_descargar_pdf(factura_id):
    """Descargar PDF de factura"""
    if not REPORTLAB_AVAILABLE:
        flash('ReportLab no está disponible. Por favor, instale la librería reportlab.', 'error')
        return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
    
    tenant_id = get_current_tenant_id()
    
    try:
        # Verificar que la factura existe
        tiene_tenant_id_facturas = False
        try:
            check_tenant_facturas = execute_query('''
                SELECT COUNT(*) as count 
                FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE()
                AND TABLE_NAME = 'facturas' 
                AND COLUMN_NAME = 'tenant_id'
            ''')
            tiene_tenant_id_facturas = check_tenant_facturas and check_tenant_facturas.get('count', 0) > 0
        except:
            tiene_tenant_id_facturas = False
        
        if tiene_tenant_id_facturas:
            factura_check = execute_query('SELECT id FROM facturas WHERE id = %s AND tenant_id = %s', (factura_id, tenant_id))
        else:
            factura_check = execute_query('SELECT id FROM facturas WHERE id = %s', (factura_id,))
        
        if not factura_check:
            flash('La factura no existe o no tiene permisos para acceder a ella.', 'error')
            return redirect(url_for('facturacion_historico'))
        
        logger.info(f"Iniciando descarga de PDF para factura {factura_id}, tenant_id={tenant_id}")
        print(f"=== INICIANDO DESCARGA PDF FACTURA {factura_id} ===")
        
        # Verificar datos de la factura antes de generar PDF
        if tiene_tenant_id_facturas:
            factura_data = execute_query('SELECT id, ncf, fecha_emision, ars_id, subtotal, total FROM facturas WHERE id = %s AND tenant_id = %s', (factura_id, tenant_id))
        else:
            factura_data = execute_query('SELECT id, ncf, fecha_emision, ars_id, subtotal, total FROM facturas WHERE id = %s', (factura_id,))
        
        if factura_data:
            logger.info(f"Datos de factura {factura_id}: ncf={factura_data.get('ncf')}, fecha={factura_data.get('fecha_emision')}, ars_id={factura_data.get('ars_id')}")
            print(f"Factura encontrada: ncf={factura_data.get('ncf')}, fecha={factura_data.get('fecha_emision')}, ars_id={factura_data.get('ars_id')}")
        else:
            logger.error(f"No se encontraron datos básicos de factura {factura_id}")
            print(f"ERROR: No se encontraron datos básicos de factura {factura_id}")
        
        detalles_check = execute_query('SELECT COUNT(*) as count FROM factura_detalles WHERE factura_id = %s', (factura_id,))
        detalles_count = detalles_check.get('count', 0) if detalles_check else 0
        logger.info(f"Factura {factura_id} tiene {detalles_count} detalles")
        print(f"Detalles encontrados: {detalles_count}")
        
        if detalles_count == 0:
            logger.error(f"Factura {factura_id} no tiene detalles. No se puede generar PDF.")
            print(f"ERROR: Factura {factura_id} no tiene detalles")
            flash('Error: La factura no tiene detalles asociados. No se puede generar el PDF.', 'error')
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
        
        print(f"Llamando a generar_pdf_factura_vista_previa para factura {factura_id}...")
        buffer = generar_pdf_factura_vista_previa(factura_id, tenant_id)
        
        if buffer is None:
            logger.error(f"No se pudo generar PDF para factura {factura_id}")
            print(f"ERROR: generar_pdf_factura_vista_previa retornó None para factura {factura_id}")
            flash('Error al generar PDF. Verifique que la factura tiene datos válidos y detalles asociados.', 'error')
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
        
        print(f"PDF generado exitosamente. Buffer tipo: {type(buffer)}")
        
        # Verificar que el buffer tiene contenido y prepararlo para envío
        try:
            # Asegurarse de que el buffer esté al inicio
            buffer.seek(0)
            
            # Verificar que el buffer tiene contenido sin leerlo (para no consumirlo)
            buffer_size = len(buffer.getvalue()) if hasattr(buffer, 'getvalue') else 0
            
            if buffer_size == 0:
                logger.error(f"PDF generado para factura {factura_id} está vacío en facturacion_descargar_pdf")
                flash('Error: El PDF generado está vacío.', 'error')
                return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
            
            logger.info(f"PDF listo para descarga: factura {factura_id}, tamaño: {buffer_size} bytes")
            
            # Asegurarse de que el buffer esté al inicio antes de enviarlo
            buffer.seek(0)
            
            factura = execute_query('SELECT numero_factura FROM facturas WHERE id = %s', (factura_id,))
            filename = f"factura_{factura_id}_{factura.get('numero_factura', '') if factura else ''}.pdf"
            
            logger.info(f"Enviando PDF: factura {factura_id}, filename={filename}")
            return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        except Exception as buffer_error:
            import traceback
            error_trace = traceback.format_exc()
            logger.error(f"Error al preparar/enviar buffer para factura {factura_id}: {str(buffer_error)}")
            logger.error(f"Traceback: {error_trace}")
            flash(f'Error al generar/enviar PDF: {str(buffer_error)}', 'error')
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Error en facturacion_descargar_pdf para factura {factura_id}: {error_trace}")
        flash(f'Error al generar PDF: {str(e)}', 'error')
        return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))

@app.route('/facturacion/facturas/<int:factura_id>/enviar-email', methods=['POST'])
@login_required
def facturacion_enviar_email(factura_id):
    """Enviar factura por email"""
    tenant_id = get_current_tenant_id()
    
    # Obtener email del destinatario
    destinatario = request.form.get('destinatario', '').strip()
    if not destinatario:
        flash('Debe especificar un email destinatario', 'error')
        return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
    
    # Validar email
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_pattern, destinatario):
        flash('Email inválido', 'error')
        return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
    
    # Verificar si existe tenant_id en facturas
    tiene_tenant_id_facturas = False
    try:
        check_tenant_facturas = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'facturas' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id_facturas = check_tenant_facturas and check_tenant_facturas.get('count', 0) > 0
    except:
        tiene_tenant_id_facturas = False
    
    # Obtener factura
    if tiene_tenant_id_facturas:
        factura = execute_query('''
            SELECT f.*, a.nombre as nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur, m.email as medico_email
            FROM facturas f
            LEFT JOIN ars a ON f.ars_id = a.id
            LEFT JOIN medicos m ON f.medico_id = m.id
            WHERE f.id = %s AND f.tenant_id = %s
        ''', (factura_id, tenant_id))
    else:
        factura = execute_query('''
            SELECT f.*, a.nombre as nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur, m.email as medico_email
            FROM facturas f
            LEFT JOIN ars a ON f.ars_id = a.id
            LEFT JOIN medicos m ON f.medico_id = m.id
            WHERE f.id = %s
        ''', (factura_id,))
    
    if not factura:
        flash('Factura no encontrada', 'error')
        return redirect(url_for('facturacion_historico'))
    
    # Si SendGrid está disponible, enviar email
    if SENDGRID_AVAILABLE and REPORTLAB_AVAILABLE:
        try:
            # Generar PDF usando la función auxiliar
            buffer = generar_pdf_factura_vista_previa(factura_id, tenant_id)
            
            if not buffer:
                flash('Error al generar PDF', 'error')
                return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
            
            pdf_data = buffer.getvalue()
            
            # Enviar email con SendGrid
            sendgrid_api_key = os.getenv('SENDGRID_API_KEY')
            sendgrid_from_email = os.getenv('SENDGRID_FROM_EMAIL', 'noreply@facturacion.com')
            
            if not sendgrid_api_key:
                flash('Configuración de email no disponible. Contacte al administrador.', 'error')
                return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
            
            message = Mail(
                from_email=sendgrid_from_email,
                to_emails=destinatario,
                subject=f"Factura #{factura.get('numero_factura', factura_id)} - {factura.get('nombre_ars', 'N/A')}",
                html_content=f"""
                <html>
                <body>
                    <h2>Factura #{factura.get('numero_factura', factura_id)}</h2>
                    <p><strong>Fecha:</strong> {factura.get('fecha_emision', '')}</p>
                    <p><strong>NCF:</strong> {factura.get('ncf', '')}</p>
                    <p><strong>Cliente:</strong> {factura.get('nombre_ars', 'N/A')}</p>
                    <p><strong>Total:</strong> RD$ {factura.get('total', 0):,.2f}</p>
                    <p>Se adjunta el PDF de la factura.</p>
                </body>
                </html>
                """
            )
            
            # Adjuntar PDF
            encoded_pdf = base64.b64encode(pdf_data).decode()
            attachment = {
                'content': encoded_pdf,
                'filename': f"factura_{factura_id}_{factura.get('numero_factura', '')}.pdf",
                'type': 'application/pdf',
                'disposition': 'attachment'
            }
            message.attachment = attachment
            
            sg = SendGridAPIClient(sendgrid_api_key)
            response = sg.send(message)
            
            if response.status_code in [200, 202]:
                flash(f'Factura enviada exitosamente a {destinatario}', 'success')
            else:
                flash(f'Error al enviar email. Código: {response.status_code}', 'error')
            
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"Error al enviar email: {error_trace}")
            flash(f'Error al enviar email: {str(e)}', 'error')
    else:
        flash('El servicio de envío de emails no está disponible. Contacte al administrador.', 'error')
    
    return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))

@app.route('/facturacion/dashboard')
@login_required
def facturacion_dashboard():
    """Dashboard de facturación"""
    from datetime import datetime, timedelta
    
    # Fechas por defecto (último mes)
    fecha_hasta = request.args.get('fecha_hasta', datetime.now().strftime('%Y-%m-%d'))
    fecha_desde = request.args.get('fecha_desde', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    
    # Estadísticas básicas
    total_facturas = 0
    total_facturado = 0.0
    monto_pendiente = 0.0
    ars_pendientes_nombres = []
    
    try:
        tenant_id = get_current_tenant_id()
        
        # Total de facturas
        result = execute_query('''
            SELECT COUNT(*) as total FROM facturas 
            WHERE tenant_id = %s
        ''', (tenant_id,))
        total_facturas = result['total'] if result else 0
        
        # Total facturado
        result = execute_query('''
            SELECT COALESCE(SUM(total), 0) as total FROM facturas 
            WHERE tenant_id = %s
        ''', (tenant_id,))
        total_facturado = float(result['total']) if result and result['total'] else 0.0
        
        # Monto pendiente (usar monto estimado de pacientes_pendientes si existe)
        try:
            result = execute_query('''
                SELECT COALESCE(SUM(monto_estimado), 0) as total 
                FROM pacientes_pendientes
                WHERE estado = 'Pendiente'
            ''')
            monto_pendiente = float(result['total']) if result and result['total'] else 0.0
        except Exception:
            monto_pendiente = 0.0
        
        # ARS pendientes
        result = execute_query('''
            SELECT DISTINCT a.nombre 
            FROM pacientes_pendientes pp 
            JOIN ars a ON pp.ars_id = a.id 
            WHERE pp.estado = 'Pendiente'
        ''', fetch='all')
        ars_pendientes_nombres = [r['nombre'] for r in result] if result else []
        
    except Exception as e:
        print(f"Error en dashboard: {e}")
    
    # Facturación por mes
    facturacion_por_mes = []
    try:
        result = execute_query('''
            SELECT DATE_FORMAT(fecha_emision, '%%Y-%%m') as mes, 
                   SUM(total) as total_monto
            FROM facturas
            WHERE fecha_emision BETWEEN %s AND %s
            GROUP BY DATE_FORMAT(fecha_emision, '%%Y-%%m')
            ORDER BY mes
        ''', (fecha_desde, fecha_hasta), fetch='all')
        facturacion_por_mes = [{'mes': r['mes'], 'total_monto': float(r['total_monto'])} for r in result] if result else []
    except:
        facturacion_por_mes = []
    
    # Facturación por ARS y mes
    facturacion_ars_mes = []
    try:
        result = execute_query('''
            SELECT DATE_FORMAT(f.fecha_emision, '%%Y-%%m') as mes,
                   a.nombre as nombre_ars,
                   SUM(f.total) as total_monto
            FROM facturas f
            JOIN ars a ON f.ars_id = a.id
            WHERE f.fecha_emision BETWEEN %s AND %s
            GROUP BY DATE_FORMAT(f.fecha_emision, '%%Y-%%m'), a.nombre
            ORDER BY mes, a.nombre
        ''', (fecha_desde, fecha_hasta), fetch='all')
        facturacion_ars_mes = [{'mes': r['mes'], 'nombre_ars': r['nombre_ars'], 'total_monto': float(r['total_monto'])} for r in result] if result else []
    except:
        facturacion_ars_mes = []
    
    # Listas para filtros
    tenant_id = get_current_tenant_id()
    ars_list = execute_query('SELECT * FROM ars WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    medicos_factura_list = execute_query('SELECT * FROM medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    medicos_consulta_list = medicos_factura_list  # Usar la misma lista
    
    return render_template('facturacion/dashboard.html',
                          total_facturas=total_facturas,
                          total_facturado=total_facturado,
                          monto_pendiente=monto_pendiente,
                          ars_pendientes_nombres=ars_pendientes_nombres,
                          facturacion_por_mes=facturacion_por_mes,
                          facturacion_ars_mes=facturacion_ars_mes,
                          ars_list=ars_list,
                          medicos_factura_list=medicos_factura_list,
                          medicos_consulta_list=medicos_consulta_list,
                          fecha_desde=fecha_desde,
                          fecha_hasta=fecha_hasta,
                          es_administrador=(current_user.perfil == 'Administrador'),
                          ars_ids_seleccionados=[],
                          medico_factura_ids_seleccionados=[],
                          medico_consulta_ids_seleccionados=[])

@app.route('/facturacion/facturas/nueva', methods=['GET', 'POST'])
@login_required
def facturacion_facturas_nueva():
    """Agregar pacientes para facturar"""
    tenant_id = get_current_tenant_id()
    
    if request.method == 'POST':
        # Obtener datos del formulario
        # Validar y sanitizar entrada
        medico_id = validate_int(request.form.get('medico_id'), min_value=1)
        ars_id = validate_int(request.form.get('ars_id'), min_value=1)
        centro_medico_id = validate_int(request.form.get('centro_medico_id'), min_value=1) if request.form.get('centro_medico_id') else None
        lineas_json = request.form.get('lineas_json', '').strip()
        
        if not medico_id or not ars_id or not lineas_json:
            flash('Faltan datos obligatorios (Médico, ARS o pacientes)', 'error')
            return redirect(url_for('facturacion_facturas_nueva'))
        
        # Validar que los IDs pertenezcan al tenant
        tenant_id = get_current_tenant_id()
        if not validate_tenant_access('medicos', medico_id) or \
           not validate_tenant_access('ars', ars_id):
            flash('No tienes acceso a uno o más de los recursos seleccionados', 'error')
            return redirect(url_for('facturacion_facturas_nueva'))
        
        if centro_medico_id and not validate_tenant_access('centros_medicos', centro_medico_id):
            flash('Centro médico no válido', 'error')
            return redirect(url_for('facturacion_facturas_nueva'))
        
        # Validar JSON
        try:
            import json
            lineas = json.loads(lineas_json)
            # Limitar número de pacientes por request (prevenir DoS)
            if len(lineas) > 1000:
                flash('Demasiados pacientes en una sola operación (máximo 1000)', 'error')
                return redirect(url_for('facturacion_facturas_nueva'))
        except json.JSONDecodeError as e:
            logger.error(f"Error al parsear JSON: {e}")
            flash('Error al procesar los datos de los pacientes', 'error')
            return redirect(url_for('facturacion_facturas_nueva'))
        
        if not lineas or len(lineas) == 0:
            flash('Debe agregar al menos un paciente', 'error')
            return redirect(url_for('facturacion_facturas_nueva'))
        
        # Procesar cada línea (paciente)
        pacientes_agregados = 0
        for linea in lineas:
            nss = sanitize_input(linea.get('nss', ''), 50)
            nombre = sanitize_input(linea.get('nombre', ''), 200)
            fecha = linea.get('fecha')
            autorizacion = sanitize_input(linea.get('autorizacion', ''), 50)
            servicio = sanitize_input(linea.get('servicio', ''), 200)
            monto = float(linea.get('monto', 0))
            
            if not nss or not nombre:
                continue
            
            # Buscar si el paciente ya existe (por NSS + ARS)
            paciente_existente = execute_query('''
                SELECT id FROM pacientes 
                WHERE nss = %s AND ars_id = %s AND tenant_id = %s
            ''', (nss, ars_id, tenant_id))
            
            paciente_id = None
            if paciente_existente:
                paciente_id = paciente_existente['id']
                # Actualizar datos del paciente si es necesario
                execute_update('''
                    UPDATE pacientes 
                    SET nombre = %s, updated_at = NOW()
                    WHERE id = %s
                ''', (nombre, paciente_id))
            else:
                # Crear nuevo paciente
                # Intentar insertar con tenant_id y created_by, si falla intentar sin ellas
                try:
                    paciente_id = execute_update('''
                        INSERT INTO pacientes (tenant_id, nombre, nss, ars_id, created_by)
                        VALUES (%s, %s, %s, %s, %s)
                    ''', (tenant_id, nombre, nss, ars_id, current_user.id))
                except Exception as e:
                    error_msg = str(e).lower()
                    # Si falla por columnas desconocidas, intentar sin ellas
                    if 'unknown column' in error_msg:
                        if 'tenant_id' in error_msg and 'created_by' in error_msg:
                            # Intentar sin ambas columnas
                            paciente_id = execute_update('''
                                INSERT INTO pacientes (nombre, nss, ars_id)
                                VALUES (%s, %s, %s)
                            ''', (nombre, nss, ars_id))
                        elif 'created_by' in error_msg:
                            # Intentar sin created_by
                            try:
                                paciente_id = execute_update('''
                                    INSERT INTO pacientes (tenant_id, nombre, nss, ars_id)
                                    VALUES (%s, %s, %s, %s)
                                ''', (tenant_id, nombre, nss, ars_id))
                            except:
                                # Si también falla tenant_id, intentar sin ambas
                                paciente_id = execute_update('''
                                    INSERT INTO pacientes (nombre, nss, ars_id)
                                    VALUES (%s, %s, %s)
                                ''', (nombre, nss, ars_id))
                        elif 'tenant_id' in error_msg:
                            # Intentar sin tenant_id
                            try:
                                paciente_id = execute_update('''
                                    INSERT INTO pacientes (nombre, nss, ars_id, created_by)
                                    VALUES (%s, %s, %s, %s)
                                ''', (nombre, nss, ars_id, current_user.id))
                            except:
                                # Si también falla created_by, intentar sin ambas
                                paciente_id = execute_update('''
                                    INSERT INTO pacientes (nombre, nss, ars_id)
                                    VALUES (%s, %s, %s)
                                ''', (nombre, nss, ars_id))
                        else:
                            raise
                    else:
                        raise  # Re-lanzar si es otro error
            
            # Crear registro en pacientes_pendientes
            servicios_realizados = f"{servicio} - Autorización: {autorizacion}" if autorizacion else servicio
            
            # Intentar insertar con todas las columnas, si falla intentar sin las opcionales
            try:
                # Intentar insertar con tenant_id y created_by
                execute_update('''
                    INSERT INTO pacientes_pendientes 
                    (tenant_id, paciente_id, nombre_paciente, nss, ars_id, fecha_servicio, 
                     servicios_realizados, monto_estimado, estado, medico_id, centro_medico_id, created_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s, %s)
                ''', (tenant_id, paciente_id, nombre, nss, ars_id, fecha, servicios_realizados, monto, medico_id, centro_medico_id, current_user.id))
            except Exception as e:
                error_msg = str(e).lower()
                # Si falla por columnas desconocidas, intentar sin ellas
                if 'unknown column' in error_msg:
                    if 'created_by' in error_msg:
                        # Intentar sin created_by
                        try:
                            execute_update('''
                                INSERT INTO pacientes_pendientes 
                                (tenant_id, paciente_id, nombre_paciente, nss, ars_id, fecha_servicio, 
                                 servicios_realizados, monto_estimado, estado, medico_id, centro_medico_id)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s)
                            ''', (tenant_id, paciente_id, nombre, nss, ars_id, fecha, servicios_realizados, monto, medico_id, centro_medico_id))
                        except Exception as e2:
                            # Si también falla tenant_id, intentar sin ambos
                            if 'tenant_id' in str(e2).lower():
                                execute_update('''
                                    INSERT INTO pacientes_pendientes 
                                    (paciente_id, nombre_paciente, nss, ars_id, fecha_servicio, 
                                     servicios_realizados, monto_estimado, estado, medico_id, centro_medico_id)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s)
                                ''', (paciente_id, nombre, nss, ars_id, fecha, servicios_realizados, monto, medico_id, centro_medico_id))
                            else:
                                raise
                    elif 'tenant_id' in error_msg:
                        # Intentar sin tenant_id
                        execute_update('''
                            INSERT INTO pacientes_pendientes 
                            (paciente_id, nombre_paciente, nss, ars_id, fecha_servicio, 
                             servicios_realizados, monto_estimado, estado, medico_id, centro_medico_id, created_by)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s, %s)
                        ''', (paciente_id, nombre, nss, ars_id, fecha, servicios_realizados, monto, medico_id, centro_medico_id, current_user.id))
                else:
                    raise  # Re-lanzar si es otro error
            
            pacientes_agregados += 1
        
        flash(f'{pacientes_agregados} paciente(s) agregado(s) como pendientes de facturación', 'success')
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    # GET: Mostrar formulario
    ars_list = execute_query('SELECT * FROM ars WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    medicos = execute_query('SELECT * FROM medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    
    # Obtener relaciones médico-centro para poblar el dropdown de centros médicos
    centros_medicos = execute_query('''
        SELECT 
            mc.medico_id,
            mc.centro_medico_id as centro_id,
            cm.nombre as centro_nombre,
            mc.es_defecto
        FROM medico_centro mc
        INNER JOIN centros_medicos cm ON mc.centro_medico_id = cm.id
        WHERE mc.tenant_id = %s AND cm.activo = 1
        ORDER BY mc.medico_id, mc.es_defecto DESC, cm.nombre
    ''', (tenant_id,), fetch='all') or []
    
    # Obtener servicios para el datalist
    servicios_list = execute_query('''
        SELECT descripcion, precio_base 
        FROM servicios 
        WHERE tenant_id = %s AND activo = 1 
        ORDER BY descripcion
    ''', (tenant_id,), fetch='all') or []
    
    return render_template('facturacion/facturas_form.html', 
                         ars_list=ars_list, 
                         medicos=medicos, 
                         centros_medicos=centros_medicos,
                         servicios_list=servicios_list)

@app.route('/facturacion/pacientes-pendientes')
@login_required
def facturacion_pacientes_pendientes():
    """Estado de facturación - Pacientes pendientes"""
    # Obtener filtros de la query string
    medico_id_filtro = request.args.get('medico_id', '')
    ars_id_filtro = request.args.get('ars_id', '')
    estado_filtro = request.args.get('estado', 'pendiente')  # Por defecto 'pendiente'
    
    # Construir query con filtros - consultar tabla pacientes_pendientes
    tenant_id = get_current_tenant_id()
    
    # Verificar si existe la columna tenant_id en pacientes_pendientes
    tiene_tenant_id = False
    try:
        check_tenant = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'pacientes_pendientes' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id = check_tenant and check_tenant.get('count', 0) > 0
    except:
        tiene_tenant_id = False
    
    # Construir query base
    query = '''
        SELECT pp.*, 
               a.nombre as nombre_ars,
               m.nombre as medico_nombre,
               m.especialidad as medico_especialidad,
               pp.servicios_realizados as descripcion_servicio,
               pp.monto_estimado as monto
        FROM pacientes_pendientes pp
        LEFT JOIN ars a ON pp.ars_id = a.id
        LEFT JOIN medicos m ON pp.medico_id = m.id
        WHERE 1=1
    '''
    params = []
    
    # Filtro por tenant_id (solo si existe la columna)
    if tiene_tenant_id:
        query += ' AND pp.tenant_id = %s'
        params.append(tenant_id)
    else:
        # Si no existe tenant_id, filtrar por médicos y ARS que pertenezcan al tenant
        query += ' AND (m.tenant_id = %s OR m.tenant_id IS NULL)'
        params.append(tenant_id)
    
    # Filtro por estado
    if estado_filtro:
        # Convertir 'pendiente' a 'Pendiente' y 'facturado' a 'Facturado'
        estado_db = estado_filtro.capitalize()
        if estado_db == 'Facturado':
            query += ' AND pp.estado = %s'
            params.append('Facturado')
        elif estado_db == 'Pendiente':
            query += ' AND pp.estado = %s'
            params.append('Pendiente')
    
    # Filtro por médico
    if medico_id_filtro:
        query += ' AND pp.medico_id = %s'
        params.append(int(medico_id_filtro))
    
    # Filtro por ARS
    if ars_id_filtro:
        query += ' AND pp.ars_id = %s'
        params.append(int(ars_id_filtro))
    
    query += ' ORDER BY pp.fecha_servicio DESC, pp.id DESC'
    
    pendientes = execute_query(query, tuple(params), fetch='all') or []
    
    # Obtener listas para filtros
    medicos = execute_query('SELECT * FROM medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    ars_list = execute_query('SELECT * FROM ars WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    
    # Obtener nombres para mostrar en los badges de filtros activos
    medico_seleccionado = None
    if medico_id_filtro:
        medico = execute_query('SELECT nombre FROM medicos WHERE id = %s', (medico_id_filtro,))
        if medico:
            medico_seleccionado = medico['nombre']
    
    ars_seleccionada = None
    if ars_id_filtro:
        ars = execute_query('SELECT nombre FROM ars WHERE id = %s', (ars_id_filtro,))
        if ars:
            ars_seleccionada = ars['nombre']
    
    # Obtener servicios para el combobox
    servicios_list = execute_query('''
        SELECT descripcion, precio_base 
        FROM servicios 
        WHERE tenant_id = %s AND activo = 1 
        ORDER BY descripcion
    ''', (tenant_id,), fetch='all') or []
    
    return render_template('facturacion/pacientes_pendientes.html', 
                          pendientes=pendientes,
                          medicos=medicos,
                          ars_list=ars_list,
                          servicios_list=servicios_list,
                          medico_id_filtro=medico_id_filtro,
                          ars_id_filtro=ars_id_filtro,
                          estado_filtro=estado_filtro,
                          medico_seleccionado=medico_seleccionado,
                          ars_seleccionada=ars_seleccionada)

@app.route('/facturacion/pacientes-pendientes/pdf')
@login_required
def facturacion_pacientes_pendientes_pdf():
    """Descargar PDF de pacientes pendientes"""
    flash('Funcionalidad de PDF en desarrollo', 'info')
    return redirect(url_for('facturacion_pacientes_pendientes'))

@app.route('/descargar-plantilla-excel')
@login_required
def descargar_plantilla_excel():
    """Descargar plantilla Excel para importar pacientes"""
    try:
        if not OPENPYXL_AVAILABLE:
            flash('La funcionalidad de Excel no está disponible', 'error')
            return redirect(url_for('facturacion_facturas_nueva'))
        
        tenant_id = get_current_tenant_id()
        if tenant_id is None:
            flash('Error al obtener el tenant', 'error')
            return redirect(url_for('facturacion_facturas_nueva'))
        
        # Obtener tema del usuario
        TEMAS = {
            'cyan': {'primary': '#06B6D4', 'primary_dark': '#0891B2'},
            'ocean': {'primary': '#0EA5E9', 'primary_dark': '#0284C7'},
            'emerald': {'primary': '#10B981', 'primary_dark': '#059669'},
            'teal': {'primary': '#14B8A6', 'primary_dark': '#0D9488'},
            'coral': {'primary': '#FF6B6B', 'primary_dark': '#EE5A52'},
            'sunset': {'primary': '#F59E0B', 'primary_dark': '#D97706'},
            'rose': {'primary': '#F43F5E', 'primary_dark': '#E11D48'},
            'amber': {'primary': '#F59E0B', 'primary_dark': '#D97706'},
            'indigo': {'primary': '#6366F1', 'primary_dark': '#4F46E5'},
            'purple': {'primary': '#A855F7', 'primary_dark': '#9333EA'},
            'violet': {'primary': '#8B5CF6', 'primary_dark': '#7C3AED'},
            'slate': {'primary': '#64748B', 'primary_dark': '#475569'},
            'navy': {'primary': '#1E3A8A', 'primary_dark': '#1E40AF'},
            'forest': {'primary': '#166534', 'primary_dark': '#14532D'},
            'wine': {'primary': '#7F1D1D', 'primary_dark': '#991B1B'},
            'bronze': {'primary': '#92400E', 'primary_dark': '#78350F'}
        }
        
        tema_actual = 'cyan'  # Default
        if current_user.is_authenticated and hasattr(current_user, 'tema_color') and current_user.tema_color:
            tema_actual = current_user.tema_color
        
        # Asegurar que tema_actual sea válido
        if tema_actual not in TEMAS:
            tema_actual = 'cyan'
        
        tema = TEMAS.get(tema_actual, TEMAS['cyan'])
        if not tema or 'primary' not in tema:
            tema = TEMAS['cyan']
        
        color_primary = tema.get('primary', '#06B6D4')
        if not color_primary:
            color_primary = '#06B6D4'
        
        # Convertir color hexadecimal a formato de openpyxl (sin #)
        color_hex_clean = color_primary.lstrip('#')
        if not color_hex_clean:
            color_hex_clean = '06B6D4'
        
        # Convertir color hexadecimal a RGB
        def hex_to_rgb(hex_color):
            if not hex_color:
                hex_color = '#06B6D4'
            hex_color = hex_color.lstrip('#')
            if not hex_color or len(hex_color) < 6:
                hex_color = '06B6D4'
            try:
                return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
            except Exception:
                return (6, 182, 212)  # Color cyan por defecto
        
        rgb_color = hex_to_rgb(color_primary)
        
        # Obtener servicios activos del tenant
        try:
            servicios_result = execute_query(
                'SELECT descripcion FROM servicios WHERE tenant_id = %s AND activo = 1 ORDER BY descripcion', 
                (tenant_id,), fetch='all'
            )
            if servicios_result is None:
                servicios_list = []
            elif isinstance(servicios_result, list):
                servicios_list = [s for s in servicios_result if s and isinstance(s, dict)]
            else:
                servicios_list = []
        except Exception as e:
            servicios_list = []
        
        # Crear workbook
        wb = Workbook()
        
        # Eliminar hoja por defecto (si existe)
        if wb.active is not None:
            try:
                wb.remove(wb.active)
            except Exception:
                pass  # Si hay error al eliminar, continuar
    
        # ========== HOJA 1: INSTRUCCIONES ==========
        ws_instrucciones = wb.create_sheet("Instrucciones", 0)
        if ws_instrucciones is None:
            raise ValueError("No se pudo crear la hoja de Instrucciones")
        
        # Título
        title_cell = ws_instrucciones.cell(row=1, column=1, value="INSTRUCCIONES PARA CARGAR PACIENTES")
        title_cell.font = Font(bold=True, size=16, color="8B5A9F")
        ws_instrucciones.merge_cells('A1:D1')
        
        # Instrucciones numeradas
        instrucciones = [
            "Complete la hoja \"Pacientes\" con los datos de los pacientes",
            "NSS: Solo números y guiones (ej: 001-234-5678)",
            "NOMBRE: Nombre completo del paciente",
            "FECHA: Formato AAAA-MM-DD (ej: 2025-10-16)",
            "AUTORIZACIÓN: Solo números, debe ser única para cada paciente",
            "SERVICIO: Seleccione de la lista desplegable (se alimenta de la hoja \"Servicios\")",
            "MONTO: Cantidad en pesos (solo números)"
        ]
        
        row = 3
        if instrucciones and isinstance(instrucciones, list):
            for i, instruccion in enumerate(instrucciones, 1):
                if instruccion is not None:
                    num_cell = ws_instrucciones.cell(row=row, column=1, value=f"{i}.")
                    num_cell.font = Font(bold=True)
                    text_cell = ws_instrucciones.cell(row=row, column=2, value=instruccion)
                    ws_instrucciones.merge_cells(f'B{row}:D{row}')
                    row += 1
        
        # Sección IMPORTANTE
        row += 1
        importante_cell = ws_instrucciones.cell(row=row, column=1, value="IMPORTANTE:")
        importante_cell.font = Font(bold=True, size=12, color="8B5A9F")
        row += 1
        
        importantes = [
            "Los encabezados están protegidos y NO se pueden modificar",
            "La columna SERVICIO tiene lista desplegable - haga clic en la flecha para seleccionar",
            "Cada autorización debe ser única",
            "Complete directamente desde la fila 2"
        ]
        
        if importantes and isinstance(importantes, list):
            for importante in importantes:
                if importante is not None:
                    bullet_cell = ws_instrucciones.cell(row=row, column=1, value="•")
                    bullet_cell.font = Font(bold=True)
                    text_cell = ws_instrucciones.cell(row=row, column=2, value=importante)
                    ws_instrucciones.merge_cells(f'B{row}:D{row}')
                    row += 1
        
        # Sección NOTA
        row += 1
        nota_title = ws_instrucciones.cell(row=row, column=1, value="NOTA:")
        nota_title.font = Font(bold=True, size=12, color="8B5A9F")
        row += 1
        nota_text = ws_instrucciones.cell(row=row, column=1, value="Antes de cargar, debe seleccionar el Médico y ARS en la página web")
        ws_instrucciones.merge_cells(f'A{row}:D{row}')
        
        # Ajustar ancho de columnas
        ws_instrucciones.column_dimensions['A'].width = 5
        ws_instrucciones.column_dimensions['B'].width = 80
        ws_instrucciones.column_dimensions['C'].width = 10
        ws_instrucciones.column_dimensions['D'].width = 10
        
        # Proteger hoja de instrucciones
        try:
            if ws_instrucciones is not None:
                # Iterar sobre las filas que tienen datos
                rows_iter = ws_instrucciones.iter_rows(min_row=1, max_row=100)
                if rows_iter is not None:
                    for row in rows_iter:
                        if row is not None:
                            for cell in row:
                                if cell is not None:
                                    cell.protection = Protection(locked=True)
            if ws_instrucciones is not None:
                ws_instrucciones.protection.sheet = True
        except Exception as e:
            # Si hay error, solo activar la protección sin bloquear celdas
            try:
                if ws_instrucciones is not None:
                    ws_instrucciones.protection.sheet = True
            except:
                pass
        
        # ========== HOJA 2: SERVICIOS ==========
        ws_servicios = wb.create_sheet("Servicios", 1)
        if ws_servicios is None:
            raise ValueError("No se pudo crear la hoja de Servicios")
        
        # Encabezado (usar color del tema del usuario)
        header_cell = ws_servicios.cell(row=1, column=1, value="SERVICIOS DISPONIBLES")
        header_cell.font = Font(bold=True, size=14, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=color_hex_clean, end_color=color_hex_clean, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_servicios.merge_cells('A1:B1')
        
        # Encabezados de tabla
        ws_servicios.cell(row=2, column=1, value="Servicio").font = Font(bold=True)
        ws_servicios.cell(row=2, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Agregar servicios
        if servicios_list and isinstance(servicios_list, list) and len(servicios_list) > 0:
            try:
                for idx, servicio in enumerate(servicios_list, 3):
                    if servicio and isinstance(servicio, dict):
                        descripcion = servicio.get('descripcion', '')
                        if descripcion:
                            ws_servicios.cell(row=idx, column=1, value=descripcion)
            except Exception as e:
                ws_servicios.cell(row=3, column=1, value="Error al cargar servicios")
        else:
            ws_servicios.cell(row=3, column=1, value="No hay servicios disponibles")
        
        # Ajustar ancho
        ws_servicios.column_dimensions['A'].width = 50
        ws_servicios.column_dimensions['B'].width = 10
        
        # Proteger hoja de servicios
        try:
            if ws_servicios is not None:
                # Iterar sobre las filas que tienen datos
                rows_iter = ws_servicios.iter_rows(min_row=1, max_row=100)
                if rows_iter is not None:
                    for row in rows_iter:
                        if row is not None:
                            for cell in row:
                                if cell is not None:
                                    cell.protection = Protection(locked=True)
            if ws_servicios is not None:
                ws_servicios.protection.sheet = True
        except Exception as e:
            # Si hay error, solo activar la protección sin bloquear celdas
            try:
                if ws_servicios is not None:
                    ws_servicios.protection.sheet = True
            except:
                pass
        
        # ========== HOJA 3: PACIENTES ==========
        ws = wb.create_sheet("Pacientes", 2)
        if ws is None:
            raise ValueError("No se pudo crear la hoja de Pacientes")
        
        # Estilos para encabezados (usar color del tema del usuario)
        header_fill = PatternFill(start_color=color_hex_clean, end_color=color_hex_clean, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Primero, desbloquear TODAS las celdas por defecto
        # Luego bloquearemos solo la fila 1 (encabezado)
        if ws is not None:
            try:
                # Desbloquear todas las celdas primero (fila 2 en adelante, hasta fila 1000)
                for row_num in range(2, 1001):
                    for col_num in range(1, 7):  # 6 columnas (A-F)
                        try:
                            cell = ws.cell(row=row_num, column=col_num)
                            if cell is not None:
                                cell.protection = Protection(locked=False)
                        except Exception:
                            pass
            except Exception:
                pass
        
        # Encabezados (bloqueados) - Solo la fila 1 estará protegida
        headers = ['NSS', 'Nombre Completo', 'Fecha', 'Autorización', 'Servicio', 'Monto']
        if headers and isinstance(headers, list):
            for col_num, header in enumerate(headers, 1):
                if header is not None:
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.protection = Protection(locked=True)  # Bloquear solo el encabezado
        
        # No agregar fila de ejemplo - el usuario llenará desde la fila 2
        
        # Validación de datos: Lista desplegable para SERVICIO (columna E)
        servicios_nombres = []
        if servicios_list and isinstance(servicios_list, list) and len(servicios_list) > 0:
            try:
                # Validar cada elemento antes de procesarlo
                for s in servicios_list:
                    if s is not None and isinstance(s, dict):
                        descripcion = s.get('descripcion', '')
                        if descripcion:
                            servicios_nombres.append(descripcion)
            except Exception as e:
                servicios_nombres = []
            
            if servicios_nombres and len(servicios_nombres) > 0:
                # Crear referencia a la hoja Servicios
                servicios_range = f"Servicios!$A$3:$A${2 + len(servicios_nombres)}"
                dv_servicio = DataValidation(type="list", formula1=servicios_range, allow_blank=False)
                dv_servicio.error = "Seleccione un servicio de la lista"
                dv_servicio.errorTitle = "Servicio inválido"
                dv_servicio.prompt = "Seleccione un servicio de la lista desplegable"
                dv_servicio.promptTitle = "Seleccionar Servicio"
                # Aplicar a toda la columna E (Servicio) desde la fila 2
                ws.add_data_validation(dv_servicio)
                dv_servicio.add(f"E2:E1048576")  # Aplicar a toda la columna E desde fila 2
        
        # Validación de datos: Autorización única (columna D)
        # Usar fórmula personalizada para verificar que no haya duplicados
        # COUNTIF($D:$D, D2) debe ser igual a 1 (solo una ocurrencia)
        dv_autorizacion = DataValidation(
            type="custom",
            formula1="COUNTIF($D:$D,D2)=1",
            allow_blank=False
        )
        dv_autorizacion.error = "Esta autorización ya existe. Cada autorización debe ser única."
        dv_autorizacion.errorTitle = "Autorización duplicada"
        dv_autorizacion.prompt = "Ingrese una autorización única (solo números)"
        dv_autorizacion.promptTitle = "Autorización"
        ws.add_data_validation(dv_autorizacion)
        dv_autorizacion.add(f"D2:D1048576")  # Aplicar a toda la columna D desde fila 2
        
        # Validación de datos: Solo números para MONTO (columna F)
        dv_monto = DataValidation(type="decimal", operator="greaterThan", formula1=0, allow_blank=False)
        dv_monto.error = "El monto debe ser un número mayor a cero"
        dv_monto.errorTitle = "Monto inválido"
        dv_monto.prompt = "Ingrese solo números (ej: 500.00)"
        dv_monto.promptTitle = "Monto"
        ws.add_data_validation(dv_monto)
        dv_monto.add(f"F2:F1048576")  # Aplicar a toda la columna F desde fila 2
        
        # Ajustar ancho de columnas
        column_widths = [15, 35, 12, 15, 30, 12]
        if column_widths and isinstance(column_widths, list):
            for col_num, width in enumerate(column_widths, 1):
                if width is not None:
                    try:
                        ws.column_dimensions[get_column_letter(col_num)].width = width
                    except Exception:
                        pass
        
        # Proteger la hoja - Solo el encabezado (fila 1) estará protegido
        # Las celdas de datos (fila 2 en adelante) permanecerán desbloqueadas
        try:
            if ws is not None and hasattr(ws, 'protection') and ws.protection is not None:
                ws.protection.sheet = True
                ws.protection.password = None
                ws.protection.formatCells = False
                ws.protection.formatColumns = False
                ws.protection.formatRows = False
                ws.protection.insertColumns = True
                ws.protection.insertRows = True
                ws.protection.insertHyperlinks = True
                ws.protection.deleteColumns = True
                ws.protection.deleteRows = True
                ws.protection.selectLockedCells = True
                ws.protection.sort = True
                ws.protection.autoFilter = True
                ws.protection.pivotTables = True
                ws.protection.selectUnlockedCells = True
            elif ws is not None:
                # Si protection no existe, solo activar la protección básica
                ws.protection.sheet = True
        except Exception as e:
            # Si hay error con la protección, continuar sin ella (no es crítico)
            try:
                if ws is not None:
                    ws.protection.sheet = True
            except:
                pass
    
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo
        filename = 'plantilla_pacientes.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except TypeError as e:
        import traceback
        error_details = traceback.format_exc()
        flash(f'Error al generar la plantilla: {str(e)}. Detalles: {error_details[:200]}', 'error')
        return redirect(url_for('facturacion_facturas_nueva'))
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        flash(f'Error inesperado al generar la plantilla: {str(e)}. Detalles: {error_details[:200]}', 'error')
        return redirect(url_for('facturacion_facturas_nueva'))

@app.route('/facturacion/procesar-excel', methods=['POST'])
@login_required
def facturacion_procesar_excel():
    """Procesar archivo Excel y devolver pacientes en formato JSON"""
    try:
        if not OPENPYXL_AVAILABLE:
            return jsonify({
                'error': True,
                'mensaje': 'La funcionalidad de Excel no está disponible',
                'errores': ['OpenPyXL no está instalado'],
                'total_errores': 1
            }), 400
        
        # Verificar que se haya enviado un archivo
        if 'archivo_excel' not in request.files:
            return jsonify({
                'error': True,
                'mensaje': 'No se recibió ningún archivo',
                'errores': ['Debe seleccionar un archivo Excel'],
                'total_errores': 1
            }), 400
        
        file = request.files['archivo_excel']
        if file.filename == '':
            return jsonify({
                'error': True,
                'mensaje': 'No se seleccionó ningún archivo',
                'errores': ['Debe seleccionar un archivo Excel'],
                'total_errores': 1
            }), 400
        
        # Validar extensión
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({
                'error': True,
                'mensaje': 'Formato de archivo inválido',
                'errores': ['El archivo debe ser .xlsx o .xls'],
                'total_errores': 1
            }), 400
        
        # Leer el archivo Excel
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
        
        # Buscar la hoja "Pacientes"
        if 'Pacientes' not in wb.sheetnames:
            return jsonify({
                'error': True,
                'mensaje': 'Hoja "Pacientes" no encontrada',
                'errores': ['El archivo Excel debe contener una hoja llamada "Pacientes"'],
                'total_errores': 1
            }), 400
        
        ws = wb['Pacientes']
        
        # Obtener tenant_id
        tenant_id = get_current_tenant_id()
        if not tenant_id:
            return jsonify({
                'error': True,
                'mensaje': 'Error al obtener el tenant',
                'errores': ['No se pudo identificar la empresa'],
                'total_errores': 1
            }), 400
        
        # Obtener servicios válidos del tenant
        servicios_result = execute_query(
            'SELECT descripcion FROM servicios WHERE tenant_id = %s AND activo = 1',
            (tenant_id,), fetch='all'
        ) or []
        servicios_validos = [s['descripcion'].upper() for s in servicios_result if s and s.get('descripcion')]
        
        # Leer datos desde la fila 2 (la fila 1 es el encabezado)
        pacientes = []
        errores = []
        autorizaciones_vistas = set()
        numero_fila = 1
        
        for row in ws.iter_rows(min_row=2, values_only=False):
            numero_fila += 1
            
            # Obtener valores de las celdas
            nss = str(row[0].value).strip() if row[0].value else ''  # Columna A
            nombre = str(row[1].value).strip() if row[1].value else ''  # Columna B
            # Fecha puede venir como datetime de Excel o como string
            fecha_raw = row[2].value if row[2].value else ''
            if fecha_raw:
                # Si es datetime de Excel, convertir a string primero
                from datetime import datetime, date
                if isinstance(fecha_raw, (datetime, date)):
                    fecha = fecha_raw.strftime('%Y-%m-%d')
                else:
                    fecha = str(fecha_raw).strip()
            else:
                fecha = ''
            autorizacion = str(row[3].value).strip() if row[3].value else ''  # Columna D
            servicio = str(row[4].value).strip() if row[4].value else ''  # Columna E
            monto = row[5].value if row[5].value else ''  # Columna F
            
            # Si la fila está vacía, saltarla
            if not nss and not nombre and not fecha and not autorizacion and not servicio and not monto:
                continue
            
            # Validaciones
            errores_fila = []
            
            # Validar NSS
            if not nss:
                errores_fila.append(f'Fila {numero_fila}: NSS es obligatorio')
            elif len(nss) > 50:
                errores_fila.append(f'Fila {numero_fila}: NSS muy largo (máximo 50 caracteres)')
            
            # Validar Nombre
            if not nombre:
                errores_fila.append(f'Fila {numero_fila}: Nombre es obligatorio')
            elif len(nombre) > 200:
                errores_fila.append(f'Fila {numero_fila}: Nombre muy largo (máximo 200 caracteres)')
            
            # Validar y normalizar Fecha
            if not fecha:
                errores_fila.append(f'Fila {numero_fila}: Fecha es obligatoria')
            else:
                try:
                    from datetime import datetime, date
                    fecha_normalizada = None
                    
                    # Si ya está en formato AAAA-MM-DD, validar y usar directamente
                    try:
                        datetime.strptime(fecha, '%Y-%m-%d')
                        fecha_normalizada = fecha  # Ya está en el formato correcto
                    except ValueError:
                        # Si no está en formato AAAA-MM-DD, intentar normalizar
                        # Normalizar separadores: convertir "/" a "-"
                        fecha_str = fecha.replace('/', '-').strip()
                        
                        # Intentar parsear diferentes formatos
                        formatos_fecha = [
                            '%Y-%m-%d',      # AAAA-MM-DD (formato estándar)
                            '%d-%m-%Y',      # DD-MM-AAAA
                            '%m-%d-%Y',      # MM-DD-AAAA
                            '%Y/%m/%d',      # AAAA/MM/DD (por si acaso quedó algún /)
                            '%d/%m/%Y',      # DD/MM/AAAA
                            '%m/%d/%Y',      # MM/DD/AAAA
                        ]
                        
                        fecha_obj = None
                        for formato in formatos_fecha:
                            try:
                                fecha_obj = datetime.strptime(fecha_str, formato)
                                break
                            except ValueError:
                                continue
                        
                        if fecha_obj:
                            # Convertir a formato estándar AAAA-MM-DD
                            fecha_normalizada = fecha_obj.strftime('%Y-%m-%d')
                        else:
                            # Si no se pudo parsear, intentar con el valor original
                            raise ValueError(f'No se pudo parsear la fecha: {fecha}')
                    
                    if fecha_normalizada:
                        # Validar que el formato sea correcto (AAAA-MM-DD)
                        datetime.strptime(fecha_normalizada, '%Y-%m-%d')
                        fecha = fecha_normalizada
                    else:
                        raise ValueError(f'Fecha no válida: {fecha}')
                        
                except Exception as e:
                    errores_fila.append(f'Fila {numero_fila}: Fecha inválida "{fecha}" (formato esperado: AAAA-MM-DD o DD/MM/AAAA)')
            
            # Validar Autorización
            if not autorizacion:
                errores_fila.append(f'Fila {numero_fila}: Autorización es obligatoria')
            elif len(autorizacion) > 50:
                errores_fila.append(f'Fila {numero_fila}: Autorización muy larga (máximo 50 caracteres)')
            elif autorizacion.upper() in autorizaciones_vistas:
                errores_fila.append(f'Fila {numero_fila}: Autorización duplicada ({autorizacion})')
            else:
                autorizaciones_vistas.add(autorizacion.upper())
            
            # Validar Servicio
            if not servicio:
                errores_fila.append(f'Fila {numero_fila}: Servicio es obligatorio')
            elif servicios_validos and servicio.upper() not in servicios_validos:
                errores_fila.append(f'Fila {numero_fila}: Servicio "{servicio}" no existe. Servicios válidos: {", ".join(servicios_validos[:5])}...')
            
            # Validar Monto
            try:
                if monto == '' or monto is None:
                    errores_fila.append(f'Fila {numero_fila}: Monto es obligatorio')
                else:
                    monto_float = float(monto)
                    if monto_float <= 0:
                        errores_fila.append(f'Fila {numero_fila}: Monto debe ser mayor a cero')
            except (ValueError, TypeError):
                errores_fila.append(f'Fila {numero_fila}: Monto inválido (debe ser un número)')
            
            # Si hay errores en esta fila, agregarlos y continuar
            if errores_fila:
                errores.extend(errores_fila)
                continue
            
            # Si no hay errores, agregar el paciente
            pacientes.append({
                'nss': nss,
                'nombre': nombre.upper(),
                'fecha': fecha,
                'autorizacion': autorizacion.upper(),
                'servicio': servicio.upper(),
                'monto': float(monto)
            })
        
        # Si hay errores, devolverlos
        if errores:
            return jsonify({
                'error': True,
                'mensaje': f'Se encontraron {len(errores)} error(es) en el archivo',
                'errores': errores,
                'total_errores': len(errores),
                'pacientes': []
            }), 400
        
        # Si no hay pacientes, devolver error
        if not pacientes:
            return jsonify({
                'error': True,
                'mensaje': 'No se encontraron pacientes válidos en el archivo',
                'errores': ['El archivo Excel no contiene datos válidos en la hoja "Pacientes"'],
                'total_errores': 1,
                'pacientes': []
            }), 400
        
        # Si todo está bien, devolver los pacientes
        return jsonify({
            'error': False,
            'mensaje': f'Se procesaron {len(pacientes)} paciente(s) correctamente',
            'pacientes': pacientes,
            'total': len(pacientes)
        }), 200
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return jsonify({
            'error': True,
            'mensaje': f'Error al procesar el archivo: {str(e)}',
            'errores': [f'Error inesperado: {str(e)}'],
            'total_errores': 1,
            'detalles': error_details[:500] if 'FLASK_ENV' in os.environ and os.environ['FLASK_ENV'] == 'development' else None
        }), 500

@app.route('/facturacion/generar', methods=['GET', 'POST'])
@login_required
def facturacion_generar():
    """Generar factura"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    # Si es POST, redirigir al step 2
    if request.method == 'POST':
        # Validar y sanitizar entrada
        ars_id = validate_int(request.form.get('ars_id'), min_value=1)
        ncf_id = validate_int(request.form.get('ncf_id'), min_value=1)
        medico_factura_id = validate_int(request.form.get('medico_factura_id'), min_value=1)
        fecha_factura = request.form.get('fecha_factura', '').strip()
        
        # Validar fecha
        if fecha_factura:
            try:
                datetime.strptime(fecha_factura, '%Y-%m-%d')
            except ValueError:
                flash('Fecha inválida', 'error')
                return redirect(url_for('facturacion_generar'))
        
        if not all([ars_id, ncf_id, medico_factura_id, fecha_factura]):
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('facturacion_generar'))
        
        # Validar que los IDs pertenezcan al tenant
        tenant_id = get_current_tenant_id()
        if not validate_tenant_access('ars', ars_id) or \
           not validate_tenant_access('ncf', ncf_id) or \
           not validate_tenant_access('medicos', medico_factura_id):
            flash('No tienes acceso a uno o más de los recursos seleccionados', 'error')
            return redirect(url_for('facturacion_generar'))
        
        # Redirigir al step 2 con los parámetros
        return redirect(url_for('facturacion_generar_step2', 
                              ars_id=ars_id, 
                              ncf_id=ncf_id, 
                              medico_factura_id=medico_factura_id,
                              fecha_factura=fecha_factura))
    
    tenant_id = get_current_tenant_id()
    
    # Obtener ARS activas
    ars_list = execute_query('''
        SELECT * FROM ars 
        WHERE activo = 1 AND tenant_id = %s 
        ORDER BY nombre
    ''', (tenant_id,), fetch='all') or []
    
    # Obtener NCF activos
    ncf_list = execute_query('''
        SELECT * FROM ncf 
        WHERE activo = 1 AND tenant_id = %s 
        ORDER BY tipo, prefijo
    ''', (tenant_id,), fetch='all') or []
    
    # Obtener información de la empresa para determinar tipo
    empresa_info = get_empresa_info(tenant_id)
    tipo_empresa = empresa_info.get('tipo_empresa') if empresa_info else None
    
    # Obtener médicos activos o razón social según tipo de empresa
    medicos_habilitados = []
    if tipo_empresa == 'centro_salud':
        # Si es centro de salud, usar la razón social de la empresa
        if empresa_info and empresa_info.get('razon_social'):
            medicos_habilitados = [{
                'id': empresa_info.get('id'),
                'nombre': empresa_info.get('razon_social'),
                'especialidad': 'Centro de Salud'
            }]
    else:
        # Si es médico o no tiene tipo definido, usar médicos
        medicos_habilitados = execute_query('''
            SELECT * FROM medicos 
            WHERE activo = 1 AND tenant_id = %s 
            ORDER BY nombre
        ''', (tenant_id,), fetch='all') or []
    
    # Obtener pacientes pendientes
    tiene_tenant_id = False
    try:
        check_tenant = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'pacientes_pendientes' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id = check_tenant and check_tenant.get('count', 0) > 0
    except:
        tiene_tenant_id = False
    
    if tiene_tenant_id:
        pendientes = execute_query('''
            SELECT pp.*, a.nombre as ars_nombre 
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            WHERE pp.estado = 'Pendiente' AND pp.tenant_id = %s
            ORDER BY pp.created_at
        ''', (tenant_id,), fetch='all') or []
    else:
        pendientes = execute_query('''
            SELECT pp.*, a.nombre as ars_nombre 
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            WHERE pp.estado = 'Pendiente'
            ORDER BY pp.created_at
        ''', fetch='all') or []
    
    # Obtener fecha actual en formato YYYY-MM-DD
    from datetime import date
    fecha_actual = date.today().strftime('%Y-%m-%d')
    
    return render_template('facturacion/generar_factura.html', 
                          pendientes=pendientes,
                          ars_list=ars_list,
                          ncf_list=ncf_list,
                          medicos_habilitados=medicos_habilitados,
                          fecha_actual=fecha_actual,
                          tipo_empresa=tipo_empresa)

@app.route('/facturacion/generar/step2', methods=['GET', 'POST'])
@login_required
def facturacion_generar_step2():
    """Generar factura - Paso 2: Selección de pacientes"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    # Si es POST, redirigir a vista previa
    if request.method == 'POST':
        pacientes_ids_json = request.form.get('pacientes_ids')
        ars_id = request.form.get('ars_id')
        ncf_id = request.form.get('ncf_id')
        medico_factura_id = request.form.get('medico_factura_id')
        fecha_factura = request.form.get('fecha_factura')
        
        if not all([pacientes_ids_json, ars_id, ncf_id, medico_factura_id, fecha_factura]):
            flash('Faltan datos obligatorios', 'error')
            return redirect(url_for('facturacion_generar'))
        
        try:
            import json
            pacientes_ids = json.loads(pacientes_ids_json)
        except json.JSONDecodeError:
            flash('Error al procesar los IDs de pacientes', 'error')
            return redirect(url_for('facturacion_generar'))
        
        if not pacientes_ids or len(pacientes_ids) == 0:
            flash('Debe seleccionar al menos un paciente', 'error')
            return redirect(url_for('facturacion_generar_step2',
                                  ars_id=ars_id,
                                  ncf_id=ncf_id,
                                  medico_factura_id=medico_factura_id,
                                  fecha_factura=fecha_factura))
        
        # Redirigir a vista previa
        return redirect(url_for('facturacion_vista_previa',
                              pacientes_ids=','.join(map(str, pacientes_ids)),
                              ars_id=ars_id,
                              ncf_id=ncf_id,
                              medico_factura_id=medico_factura_id,
                              fecha_factura=fecha_factura))
    
    # Obtener parámetros de la URL (GET)
    ars_id = request.args.get('ars_id')
    ncf_id = request.args.get('ncf_id')
    medico_factura_id = request.args.get('medico_factura_id')
    fecha_factura = request.args.get('fecha_factura')
    
    if not all([ars_id, ncf_id, medico_factura_id, fecha_factura]):
        flash('Faltan parámetros obligatorios', 'error')
        return redirect(url_for('facturacion_generar'))
    
    tenant_id = get_current_tenant_id()
    
    # Obtener ARS
    ars = execute_query('SELECT * FROM ars WHERE id = %s AND tenant_id = %s', (ars_id, tenant_id))
    if not ars:
        flash('ARS no encontrada', 'error')
        return redirect(url_for('facturacion_generar'))
    
    # Obtener NCF
    ncf = execute_query('SELECT * FROM ncf WHERE id = %s AND tenant_id = %s', (ncf_id, tenant_id))
    if not ncf:
        flash('NCF no encontrado', 'error')
        return redirect(url_for('facturacion_generar'))
    
    # Obtener información de la empresa para determinar tipo
    empresa_info = get_empresa_info(tenant_id)
    tipo_empresa = empresa_info.get('tipo_empresa') if empresa_info else None
    
    # Obtener médico o empresa según tipo
    medico_factura = None
    medico_factura_nombre = 'N/A'
    
    if tipo_empresa == 'centro_salud':
        # Si es centro de salud, buscar en empresas
        empresa_factura = execute_query('SELECT * FROM empresas WHERE id = %s', (medico_factura_id,))
        if empresa_factura and empresa_factura.get('id') == tenant_id:
            medico_factura = {
                'id': empresa_factura.get('id'),
                'nombre': empresa_factura.get('razon_social', empresa_factura.get('nombre', 'N/A')),
                'especialidad': 'Centro de Salud'
            }
            medico_factura_nombre = empresa_factura.get('razon_social', empresa_factura.get('nombre', 'N/A'))
        else:
            flash('Empresa no encontrada', 'error')
            return redirect(url_for('facturacion_generar'))
    else:
        # Si es médico, buscar en medicos
        medico_factura = execute_query('SELECT * FROM medicos WHERE id = %s AND tenant_id = %s', (medico_factura_id, tenant_id))
        if not medico_factura:
            flash('Médico no encontrado', 'error')
            return redirect(url_for('facturacion_generar'))
        medico_factura_nombre = medico_factura.get('nombre', 'N/A')
    
    # Obtener pacientes pendientes filtrados por ARS
    tiene_tenant_id = False
    try:
        check_tenant = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'pacientes_pendientes' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id = check_tenant and check_tenant.get('count', 0) > 0
    except:
        tiene_tenant_id = False
    
    if tiene_tenant_id:
        pendientes_raw = execute_query('''
            SELECT pp.*, a.nombre as ars_nombre, m.nombre as medico_nombre
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            LEFT JOIN medicos m ON pp.medico_id = m.id
            WHERE pp.estado = 'Pendiente' AND pp.ars_id = %s AND pp.tenant_id = %s
            ORDER BY pp.created_at
        ''', (ars_id, tenant_id), fetch='all') or []
    else:
        pendientes_raw = execute_query('''
            SELECT pp.*, a.nombre as ars_nombre, m.nombre as medico_nombre
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            LEFT JOIN medicos m ON pp.medico_id = m.id
            WHERE pp.estado = 'Pendiente' AND pp.ars_id = %s
            ORDER BY pp.created_at
        ''', (ars_id,), fetch='all') or []
    
    # Procesar los datos para extraer autorización y servicio
    pendientes = []
    for p in pendientes_raw:
        servicio_completo = p.get('servicios_realizados', '') or ''
        # Extraer servicio y autorización
        if ' - Autorización:' in servicio_completo:
            partes = servicio_completo.split(' - Autorización:')
            descripcion_servicio = partes[0].strip()
            autorizacion = partes[1].strip() if len(partes) > 1 else ''
        else:
            descripcion_servicio = servicio_completo.strip()
            autorizacion = ''
        
        p['descripcion_servicio'] = descripcion_servicio
        p['autorizacion'] = autorizacion
        p['paciente_nombre_completo'] = p.get('nombre_paciente', '')
        pendientes.append(p)
    
    # Obtener todos los médicos para el filtro
    medicos = execute_query('SELECT id, nombre FROM medicos WHERE activo = 1 AND tenant_id = %s ORDER BY nombre', (tenant_id,), fetch='all') or []
    
    return render_template('facturacion/generar_factura_step2.html',
                          ars=ars,
                          ncf=ncf,
                          medico_factura_id=medico_factura_id,
                          medico_factura_nombre=medico_factura_nombre,
                          fecha_factura=fecha_factura,
                          pendientes=pendientes,
                          medicos=medicos)

@app.route('/facturacion/vista-previa')
@login_required
def facturacion_vista_previa():
    """Vista previa de factura antes de generar"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    # Obtener parámetros
    pacientes_ids_str = request.args.get('pacientes_ids', '')
    ars_id = request.args.get('ars_id')
    ncf_id = request.args.get('ncf_id')
    medico_factura_id = request.args.get('medico_factura_id')
    fecha_factura = request.args.get('fecha_factura')
    
    if not all([pacientes_ids_str, ars_id, ncf_id, medico_factura_id, fecha_factura]):
        flash('Faltan parámetros obligatorios', 'error')
        return redirect(url_for('facturacion_generar'))
    
    # Convertir IDs de pacientes
    try:
        pacientes_ids = [int(id) for id in pacientes_ids_str.split(',') if id.strip()]
    except ValueError:
        flash('Error en los IDs de pacientes', 'error')
        return redirect(url_for('facturacion_generar'))
    
    if not pacientes_ids:
        flash('Debe seleccionar al menos un paciente', 'error')
        return redirect(url_for('facturacion_generar'))
    
    tenant_id = get_current_tenant_id()
    
    # Obtener datos de ARS, NCF y Médico
    ars = execute_query('SELECT * FROM ars WHERE id = %s AND tenant_id = %s', (ars_id, tenant_id))
    if not ars:
        flash('ARS no encontrada', 'error')
        return redirect(url_for('facturacion_generar'))
    
    ncf = execute_query('SELECT * FROM ncf WHERE id = %s AND tenant_id = %s', (ncf_id, tenant_id))
    if not ncf:
        flash('NCF no encontrado', 'error')
        return redirect(url_for('facturacion_generar'))
    
    # Obtener información de la empresa para determinar tipo
    empresa_info = get_empresa_info(tenant_id)
    tipo_empresa = empresa_info.get('tipo_empresa') if empresa_info else None
    
    # Obtener médico o empresa según tipo
    medico_factura = None
    empresa_factura = None
    if tipo_empresa == 'centro_salud':
        # Si es centro de salud, buscar en empresas
        empresa_factura = execute_query('SELECT * FROM empresas WHERE id = %s', (medico_factura_id,))
        if empresa_factura and empresa_factura.get('id') == tenant_id:
            medico_factura = {
                'id': empresa_factura.get('id'),
                'nombre': empresa_factura.get('razon_social', empresa_factura.get('nombre', 'N/A')),
                'especialidad': 'Centro de Salud',
                'cedula': empresa_factura.get('rnc', '')
            }
        else:
            flash('Empresa no encontrada', 'error')
            return redirect(url_for('facturacion_generar'))
    else:
        # Si es médico, buscar en medicos
        medico_factura = execute_query('SELECT * FROM medicos WHERE id = %s AND tenant_id = %s', (medico_factura_id, tenant_id))
        if not medico_factura:
            flash('Médico no encontrado', 'error')
            return redirect(url_for('facturacion_generar'))
    
    # Obtener pacientes seleccionados
    tiene_tenant_id = False
    try:
        check_tenant = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'pacientes_pendientes' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id = check_tenant and check_tenant.get('count', 0) > 0
    except:
        tiene_tenant_id = False
    
    placeholders = ','.join(['%s'] * len(pacientes_ids))
    if tiene_tenant_id:
        pacientes_query = f'''
            SELECT pp.*, a.nombre as ars_nombre, m.nombre as medico_nombre
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            LEFT JOIN medicos m ON pp.medico_id = m.id
            WHERE pp.id IN ({placeholders}) AND pp.tenant_id = %s
            ORDER BY pp.fecha_servicio
        '''
        pacientes_raw = execute_query(pacientes_query, tuple(pacientes_ids) + (tenant_id,), fetch='all') or []
    else:
        pacientes_query = f'''
            SELECT pp.*, a.nombre as ars_nombre, m.nombre as medico_nombre
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            LEFT JOIN medicos m ON pp.medico_id = m.id
            WHERE pp.id IN ({placeholders})
            ORDER BY pp.fecha_servicio
        '''
        pacientes_raw = execute_query(pacientes_query, tuple(pacientes_ids), fetch='all') or []
    
    # Procesar pacientes
    pacientes = []
    for p in pacientes_raw:
        servicio_completo = p.get('servicios_realizados', '') or ''
        if ' - Autorización:' in servicio_completo:
            partes = servicio_completo.split(' - Autorización:')
            descripcion_servicio = partes[0].strip()
            autorizacion = partes[1].strip() if len(partes) > 1 else ''
        else:
            descripcion_servicio = servicio_completo.strip()
            autorizacion = ''
        
        p['descripcion_servicio'] = descripcion_servicio
        p['autorizacion'] = autorizacion
        p['paciente_nombre_completo'] = p.get('nombre_paciente', '')
        pacientes.append(p)
    
    # Obtener centro médico del médico (si tiene uno asociado)
    centro_medico = None
    if medico_factura.get('centro_medico_id'):
        centro_medico = execute_query('SELECT * FROM centros_medicos WHERE id = %s AND tenant_id = %s', 
                                     (medico_factura['centro_medico_id'], tenant_id))
    
    # Si no tiene centro médico asociado, obtener el primero del tenant
    if not centro_medico:
        centro_medico = execute_query('SELECT * FROM centros_medicos WHERE tenant_id = %s LIMIT 1', (tenant_id,))
    
    # Si aún no hay centro médico, crear uno por defecto
    if not centro_medico:
        centro_medico = {
            'nombre': 'Centro Médico',
            'direccion': ''
        }
    
    # Calcular subtotal y total
    subtotal = sum(float(p.get('monto_estimado', 0) or 0) for p in pacientes)
    total = subtotal  # Por ahora el total es igual al subtotal (ITBIS es exento)
    
    # Generar número de NCF completo para mostrar en vista previa
    proximo_numero = ncf.get('ultimo_numero', 0) + 1
    tamano_secuencia = ncf.get('tamano_secuencia', 8)
    ncf_completo = f"{ncf.get('prefijo', '')}{proximo_numero:0{tamano_secuencia}d}"
    
    # Mapeo de tipos de NCF a descripciones
    ncf_tipos_descripciones = {
        'B01': 'Factura de Crédito Fiscal',
        'B02': 'Factura de Consumo',
        'B14': 'Registro Único de Ingresos',
        'B15': 'GUBERNAMENTAL'
    }
    ncf_tipo_descripcion = ncf_tipos_descripciones.get(ncf.get('tipo', ''), ncf.get('tipo', ''))
    
    return render_template('facturacion/vista_previa_factura.html',
                          pacientes=pacientes,
                          pacientes_ids=','.join(map(str, pacientes_ids)),
                          ars=ars,
                          ncf=ncf,
                          ncf_completo=ncf_completo,  # Número completo del NCF
                          ncf_tipo_descripcion=ncf_tipo_descripcion,  # Descripción del tipo de NCF
                          medico=medico_factura,  # Cambiado de medico_factura a medico
                          medico_factura=medico_factura,  # Mantener también por si acaso
                          medico_factura_id=medico_factura_id,
                          fecha_factura=fecha_factura,
                          ars_id=ars_id,
                          ncf_id=ncf_id,
                          centro_medico=centro_medico,
                          subtotal=subtotal,
                          total=total,
                          tipo_empresa=tipo_empresa,
                          empresa_info=empresa_info)

@app.route('/facturacion/generar/final', methods=['POST'])
@login_required
def facturacion_generar_final():
    """Generar factura final en la base de datos"""
    if current_user.perfil not in ['Administrador', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    tenant_id = get_current_tenant_id()
    
    # Obtener datos del formulario
    pacientes_ids_str = request.form.get('pacientes_ids', '')
    ars_id = request.form.get('ars_id')
    ncf_id = request.form.get('ncf_id')
    medico_factura_id = request.form.get('medico_factura_id')
    fecha_factura = request.form.get('fecha_factura')
    
    if not all([pacientes_ids_str, ars_id, ncf_id, medico_factura_id, fecha_factura]):
        flash('Faltan datos obligatorios', 'error')
        return redirect(url_for('facturacion_generar'))
    
    # Convertir IDs de pacientes
    try:
        pacientes_ids = [int(id) for id in pacientes_ids_str.split(',') if id.strip()]
    except ValueError:
        flash('Error en los IDs de pacientes', 'error')
        return redirect(url_for('facturacion_generar'))
    
    if not pacientes_ids:
        flash('Debe seleccionar al menos un paciente', 'error')
        return redirect(url_for('facturacion_generar'))
    
    # Obtener datos de ARS, NCF y Médico
    ars = execute_query('SELECT * FROM ars WHERE id = %s AND tenant_id = %s', (ars_id, tenant_id))
    if not ars:
        flash('ARS no encontrada', 'error')
        return redirect(url_for('facturacion_generar'))
    
    ncf = execute_query('SELECT * FROM ncf WHERE id = %s AND tenant_id = %s', (ncf_id, tenant_id))
    if not ncf:
        flash('NCF no encontrado', 'error')
        return redirect(url_for('facturacion_generar'))
    
    # Obtener información de la empresa para determinar tipo
    empresa_info = get_empresa_info(tenant_id)
    tipo_empresa = empresa_info.get('tipo_empresa') if empresa_info else None
    
    # Obtener médico o empresa según tipo
    medico_factura = None
    empresa_factura = None
    if tipo_empresa == 'centro_salud':
        # Si es centro de salud, buscar en empresas
        empresa_factura = execute_query('SELECT * FROM empresas WHERE id = %s', (medico_factura_id,))
        if empresa_factura and empresa_factura.get('id') == tenant_id:
            medico_factura = {
                'id': empresa_factura.get('id'),
                'nombre': empresa_factura.get('razon_social', empresa_factura.get('nombre', 'N/A')),
                'especialidad': 'Centro de Salud',
                'cedula': empresa_factura.get('rnc', '')
            }
        else:
            flash('Empresa no encontrada', 'error')
            return redirect(url_for('facturacion_generar'))
    else:
        # Si es médico, buscar en medicos
        medico_factura = execute_query('SELECT * FROM medicos WHERE id = %s AND tenant_id = %s', (medico_factura_id, tenant_id))
        if not medico_factura:
            flash('Médico no encontrado', 'error')
            return redirect(url_for('facturacion_generar'))
    
    # Obtener pacientes seleccionados
    tiene_tenant_id = False
    try:
        check_tenant = execute_query('''
            SELECT COUNT(*) as count 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'pacientes_pendientes' 
            AND COLUMN_NAME = 'tenant_id'
        ''')
        tiene_tenant_id = check_tenant and check_tenant.get('count', 0) > 0
    except:
        tiene_tenant_id = False
    
    placeholders = ','.join(['%s'] * len(pacientes_ids))
    if tiene_tenant_id:
        pacientes_raw = execute_query(f'''
            SELECT pp.*, a.nombre as ars_nombre, m.nombre as medico_nombre
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            LEFT JOIN medicos m ON pp.medico_id = m.id
            WHERE pp.id IN ({placeholders}) AND pp.tenant_id = %s
            ORDER BY pp.fecha_servicio
        ''', tuple(pacientes_ids) + (tenant_id,), fetch='all') or []
    else:
        pacientes_raw = execute_query(f'''
            SELECT pp.*, a.nombre as ars_nombre, m.nombre as medico_nombre
            FROM pacientes_pendientes pp
            LEFT JOIN ars a ON pp.ars_id = a.id
            LEFT JOIN medicos m ON pp.medico_id = m.id
            WHERE pp.id IN ({placeholders})
            ORDER BY pp.fecha_servicio
        ''', tuple(pacientes_ids), fetch='all') or []
    
    if not pacientes_raw:
        flash('No se encontraron los pacientes seleccionados', 'error')
        return redirect(url_for('facturacion_generar'))
    
    # Calcular totales
    subtotal = sum(float(p.get('monto_estimado', 0) or 0) for p in pacientes_raw)
    total = subtotal
    
    # Generar número de NCF
    # Obtener el próximo número del NCF
    proximo_numero = ncf.get('ultimo_numero', 0) + 1
    tamano_secuencia = ncf.get('tamano_secuencia', 8)
    ncf_numero = f"{ncf['prefijo']}{proximo_numero:0{tamano_secuencia}d}"
    
    # Generar número de factura
    from datetime import datetime
    fecha_actual = datetime.now()
    numero_factura = f"FAC-{fecha_actual.strftime('%Y%m%d')}-{proximo_numero:04d}"
    
    try:
        # Obtener centro médico
        centro_medico_id = None
        centro_medico_nombre = None
        if medico_factura.get('centro_medico_id'):
            centro_medico = execute_query('SELECT * FROM centros_medicos WHERE id = %s AND tenant_id = %s', 
                                         (medico_factura['centro_medico_id'], tenant_id))
            if centro_medico:
                centro_medico_id = centro_medico['id']
                centro_medico_nombre = centro_medico.get('nombre', '')
        
        # Crear factura
        # Verificar si existe tenant_id en facturas
        tiene_tenant_id_facturas = False
        try:
            check_tenant_facturas = execute_query('''
                SELECT COUNT(*) as count 
                FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE()
                AND TABLE_NAME = 'facturas' 
                AND COLUMN_NAME = 'tenant_id'
            ''')
            tiene_tenant_id_facturas = check_tenant_facturas and check_tenant_facturas.get('count', 0) > 0
        except:
            tiene_tenant_id_facturas = False
        
        # Usar el primer paciente como referencia para datos generales
        primer_paciente = pacientes_raw[0]
        
        if tiene_tenant_id_facturas:
            factura_id = execute_update('''
                INSERT INTO facturas 
                (tenant_id, numero_factura, ncf, fecha_emision, paciente_id, nombre_paciente, 
                 cedula_paciente, nss_paciente, ars_id, nombre_ars, medico_id, nombre_medico,
                 centro_medico_id, nombre_centro_medico, subtotal, itbis, total, estado, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s)
            ''', (tenant_id, numero_factura, ncf_numero, fecha_factura, 
                  primer_paciente.get('paciente_id'), primer_paciente.get('nombre_paciente', ''),
                  primer_paciente.get('cedula'), primer_paciente.get('nss'),
                  ars_id, ars.get('nombre', ''), medico_factura_id, medico_factura.get('nombre', ''),
                  centro_medico_id, centro_medico_nombre, subtotal, 0, total, current_user.id))
        else:
            factura_id = execute_update('''
                INSERT INTO facturas 
                (numero_factura, ncf, fecha_emision, paciente_id, nombre_paciente, 
                 cedula_paciente, nss_paciente, ars_id, nombre_ars, medico_id, nombre_medico,
                 centro_medico_id, nombre_centro_medico, subtotal, itbis, total, estado, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s)
            ''', (numero_factura, ncf_numero, fecha_factura, 
                  primer_paciente.get('paciente_id'), primer_paciente.get('nombre_paciente', ''),
                  primer_paciente.get('cedula'), primer_paciente.get('nss'),
                  ars_id, ars.get('nombre', ''), medico_factura_id, medico_factura.get('nombre', ''),
                  centro_medico_id, centro_medico_nombre, subtotal, 0, total, current_user.id))
        
        # Crear detalles de factura para cada paciente
        for paciente in pacientes_raw:
            servicio_completo = paciente.get('servicios_realizados', '') or ''
            if ' - Autorización:' in servicio_completo:
                descripcion_servicio = servicio_completo.split(' - Autorización:')[0].strip()
            else:
                descripcion_servicio = servicio_completo.strip()
            
            monto = float(paciente.get('monto_estimado', 0) or 0)
            
            execute_update('''
                INSERT INTO factura_detalles 
                (factura_id, descripcion, cantidad, precio_unitario, subtotal)
                VALUES (%s, %s, 1, %s, %s)
            ''', (factura_id, descripcion_servicio, monto, monto))
        
        # Actualizar estado de pacientes_pendientes a 'Facturado'
        if tiene_tenant_id:
            execute_update(f'''
                UPDATE pacientes_pendientes 
                SET estado = 'Facturado' 
                WHERE id IN ({placeholders}) AND tenant_id = %s
            ''', tuple(pacientes_ids) + (tenant_id,))
        else:
            execute_update(f'''
                UPDATE pacientes_pendientes 
                SET estado = 'Facturado' 
                WHERE id IN ({placeholders})
            ''', tuple(pacientes_ids))
        
        # Actualizar número del NCF
        execute_update('''
            UPDATE ncf 
            SET ultimo_numero = %s 
            WHERE id = %s
        ''', (proximo_numero, ncf_id))
        
        flash(f'Factura {numero_factura} generada exitosamente', 'success')
        return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error al generar factura: {error_trace}")
        flash(f'Error al generar la factura: {str(e)}', 'error')
        return redirect(url_for('facturacion_generar'))

# Rutas alternativas del template base (redirigen al menú principal)
@app.route('/services')
@app.route('/about')
@app.route('/contact')
@app.route('/request-appointment')
def redirects():
    """Rutas alternativas - redirigen al login o menú si está autenticado"""
    if current_user.is_authenticated:
        return redirect(url_for('facturacion_menu'))
    return redirect(url_for('login'))

@app.route('/admin/usuarios')
@login_required
def admin_usuarios():
    """Listar usuarios - Filtra por tenant del usuario actual"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    # Obtener usuarios del mismo tenant
    tenant_id = get_current_tenant_id()
    usuarios = execute_query('''
        SELECT u.*, e.nombre as empresa_nombre
        FROM usuarios u
        LEFT JOIN empresas e ON u.tenant_id = e.id
        WHERE u.tenant_id = %s 
        ORDER BY u.created_at DESC
    ''', (tenant_id,), fetch='all')
    
    # Obtener info de licencias
    empresa = get_empresa_info()
    
    return render_template('usuarios/lista.html', usuarios=usuarios, empresa=empresa)

@app.route('/admin/usuarios/nuevo', methods=['GET', 'POST'])
@login_required
def admin_usuarios_nuevo():
    """Crear nuevo usuario"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 100)
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password_nuevo', '')
        perfil = request.form.get('perfil', '')
        
        if not nombre or not email or not password or not perfil:
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        if not validate_email(email):
            flash('Email inválido', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        password_errors = validar_password_segura(password)
        if password_errors:
            flash(f'Contraseña no válida: {", ".join(password_errors)}', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        if perfil not in ['Administrador', 'Nivel 2', 'Registro de Facturas']:
            flash('Perfil inválido', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        # Obtener tenant_id del usuario actual
        tenant_id = get_current_tenant_id()
        
        # VALIDAR LICENCIAS DISPONIBLES
        licencia_ok, licencias_disponibles, mensaje = check_license_available(tenant_id)
        if not licencia_ok:
            flash(f'No se puede crear usuario: {mensaje}', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        # Verificar email en el mismo tenant
        existe = execute_query(
            'SELECT id FROM usuarios WHERE email = %s AND tenant_id = %s', 
            (email, tenant_id)
        )
        
        if existe:
            flash('Ya existe un usuario con ese email en tu empresa', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        password_hash = generate_password_hash(password)
        execute_update('''
            INSERT INTO usuarios (tenant_id, nombre, email, password_hash, perfil, activo, password_temporal)
            VALUES (%s, %s, %s, %s, %s, 1, 1)
        ''', (tenant_id, nombre, email, password_hash, perfil))
        
        flash(f'Usuario {nombre} creado exitosamente ({licencias_disponibles - 1} licencias restantes)', 'success')
        return redirect(url_for('admin_usuarios'))
    
    return render_template('usuarios/form.html', usuario=None)

@app.route('/admin/usuarios/<int:usuario_id>/editar', methods=['GET', 'POST'])
@login_required
def admin_usuarios_editar(usuario_id):
    """Editar usuario"""
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos', 'error')
        return redirect(url_for('facturacion_menu'))
    
    usuario = execute_query('SELECT * FROM usuarios WHERE id = %s', (usuario_id,))
    
    if not usuario:
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('admin_usuarios'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 100)
        email = request.form.get('email', '').strip().lower()
        perfil = request.form.get('perfil', '')
        activo = request.form.get('activo') == '1'
        cambiar_password = request.form.get('cambiar_password') == '1'
        password = request.form.get('password', '')
        
        if not nombre or not email or not perfil:
            flash('Nombre, email y perfil son obligatorios', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        if not validate_email(email):
            flash('Email inválido', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        if perfil not in ['Administrador', 'Nivel 2', 'Registro de Facturas']:
            flash('Perfil inválido', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        if usuario_id == current_user.id and not activo:
            flash('No puedes desactivar tu propia cuenta', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        existe = execute_query('SELECT id FROM usuarios WHERE email = %s AND id != %s',
                             (email, usuario_id))
        
        if existe:
            flash('Ya existe otro usuario con ese email', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        if cambiar_password and password:
            if len(password) < 8:
                flash('La contraseña debe tener al menos 8 caracteres', 'error')
                return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
            
            password_hash = generate_password_hash(password)
            execute_update('''
                UPDATE usuarios 
                SET nombre = %s, email = %s, password_hash = %s, perfil = %s, activo = %s, password_temporal = 1
                WHERE id = %s
            ''', (nombre, email, password_hash, perfil, activo, usuario_id))
            
            if usuario_id == current_user.id:
                logout_user()
                flash('Tu contraseña ha sido cambiada.', 'warning')
                return redirect(url_for('login'))
            
            flash(f'Usuario {nombre} actualizado con nueva contraseña', 'success')
        else:
            execute_update('''
                UPDATE usuarios 
                SET nombre = %s, email = %s, perfil = %s, activo = %s
                WHERE id = %s
            ''', (nombre, email, perfil, activo, usuario_id))
            
            flash(f'Usuario {nombre} actualizado exitosamente', 'success')
        
        return redirect(url_for('admin_usuarios'))
    
    return render_template('usuarios/form.html', usuario=usuario)

@app.route('/admin/usuarios/<int:usuario_id>/eliminar', methods=['POST'])
@login_required
def admin_usuarios_eliminar(usuario_id):
    """Eliminar usuario - DESHABILITADO"""
    flash('Eliminación deshabilitada. Desactiva el usuario en su lugar.', 'warning')
    return redirect(url_for('admin_usuarios'))

@app.route('/perfil/configuracion', methods=['GET', 'POST'])
@login_required
def perfil_configuracion():
    """Configuración del perfil del usuario"""
    TEMAS_VALIDOS = ['cyan', 'ocean', 'emerald', 'teal', 'coral', 'sunset', 'rose', 'amber', 
                     'indigo', 'purple', 'violet', 'slate', 'navy', 'forest', 'wine', 'bronze']
    
    if request.method == 'POST':
        tema_color = request.form.get('tema_color')
        
        if tema_color not in TEMAS_VALIDOS:
            flash('Tema de color inválido', 'error')
            return redirect(url_for('perfil_configuracion'))
        
        # Actualizar tema del usuario
        execute_update('UPDATE usuarios SET tema_color = %s WHERE id = %s',
                      (tema_color, current_user.id))
        
        # Actualizar el usuario actual en sesión
        current_user.tema_color = tema_color
        
        flash('Tema de color actualizado exitosamente. Recarga la página para ver los cambios.', 'success')
        return redirect(url_for('perfil_configuracion'))
    
    return render_template('perfil/configuracion.html')

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    host = os.getenv('HOST', '0.0.0.0')
    debug = os.getenv('FLASK_ENV') != 'production'
    
    print("\n" + "="*60)
    print(" ARSFLOW GESTION DE FACTRAS MEDICAS")
    print("="*60)
    print(f" Entorno: {'PRODUCCION' if not debug else 'DESARROLLO'}")
    print(f" Host: {host}:{port}")
    print(f" Base de datos: {DATABASE_CONFIG['database']}")
    print("="*60 + "\n")
    
    app.run(host=host, port=port, debug=debug)
